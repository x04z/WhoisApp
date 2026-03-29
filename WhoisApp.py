import streamlit as st
from streamlit_option_menu import option_menu
import pandas as pd
import requests
import time
from concurrent.futures import ThreadPoolExecutor, wait, FIRST_COMPLETED
import socket
import struct
import ipaddress
from urllib.parse import quote
import math
import random
import altair as alt 
alt.data_transformers.disable_max_rows() # 5000行以上の大容量データセットの描画を許可する
import json 
import io 
import re 
import subprocess
import dns.resolver
import dns.reversename
import dns
import zipfile
import datetime
import tempfile
import os
import bisect
import uuid

BACKUP_FILE = "whois_recovery_session.json"
BACKUP_DETAILS_FILE = "whois_recovery_details.json"

def save_recovery_data():
    """ 検索進捗と詳細データをローカルに退避させる """
    if IS_PUBLIC_MODE: return # パブリック環境ではストレージ保護のため無効化
    try:
        session_data = {
            'raw_results': st.session_state.raw_results,
            'targets_cache': st.session_state.targets_cache,
            'deferred_ips': st.session_state.deferred_ips,
            'finished_ips': list(st.session_state.finished_ips),
            'target_freq_map': st.session_state.target_freq_map,
            'cidr_cache': st.session_state.cidr_cache,
            'learned_proxy_isps': st.session_state.learned_proxy_isps,
            'resolved_dns_map': st.session_state.resolved_dns_map,
        }
        
        # 一時ファイルに完全に書き込んでからリネーム(アトミック書き込み)し、クラッシュ時のデータ破損を防ぐ
        tmp_session = BACKUP_FILE + ".tmp"
        tmp_details = BACKUP_DETAILS_FILE + ".tmp"
        
        with open(tmp_session, "w", encoding="utf-8") as f:
            json.dump(session_data, f, ensure_ascii=False)
        with open(tmp_details, "w", encoding="utf-8") as f:
            json.dump(st.session_state.detailed_data, f, ensure_ascii=False)
            
        os.replace(tmp_session, BACKUP_FILE)
        os.replace(tmp_details, BACKUP_DETAILS_FILE)
    except TypeError as e:
        import logging
        logging.error(f"[Recovery Save Error] JSONシリアライズ失敗: {e}")
    except OSError as e:
        import logging
        logging.error(f"[Recovery Save Error] ファイル保存/置換失敗: {e}")

def load_recovery_data():
    """ 中断されたデータを復元し、再開フラグを立てる """
    if IS_PUBLIC_MODE: return False
    try:
        if os.path.exists(BACKUP_FILE) and os.path.exists(BACKUP_DETAILS_FILE):
            with open(BACKUP_FILE, "r", encoding="utf-8") as f:
                session_data = json.load(f)
            with open(BACKUP_DETAILS_FILE, "r", encoding="utf-8") as f:
                detailed_data = json.load(f)
                
            st.session_state.raw_results = session_data['raw_results']
            st.session_state.targets_cache = session_data['targets_cache']
            st.session_state.deferred_ips = session_data['deferred_ips']
            st.session_state.finished_ips = set(session_data['finished_ips'])
            st.session_state.target_freq_map = session_data['target_freq_map']
            st.session_state.cidr_cache = session_data['cidr_cache']
            st.session_state.learned_proxy_isps = session_data['learned_proxy_isps']
            st.session_state.resolved_dns_map = session_data['resolved_dns_map']
            st.session_state.detailed_data = detailed_data
            
            st.session_state.is_searching = True
            st.session_state.cancel_search = False
            return True
    except Exception:
        pass
    return False

def clear_recovery_data():
    """ 正常終了時や新規検索時にバックアップを破棄する """
    if IS_PUBLIC_MODE: return
    try:
        if os.path.exists(BACKUP_FILE): os.remove(BACKUP_FILE)
        if os.path.exists(BACKUP_DETAILS_FILE): os.remove(BACKUP_DETAILS_FILE)
    except OSError as e:
        import logging
        logging.warning(f"バックアップファイルの削除に失敗しました: {e}")

import platform

def open_local_path(path):
    """ ローカルPCでファイルやフォルダをOSの標準機能で開く """
    if IS_PUBLIC_MODE: return
    try:
        if platform.system() == 'Windows':
            os.startfile(path)
        elif platform.system() == 'Darwin': # macOS
            subprocess.Popen(['open', path])
        else: # Linux etc.
            subprocess.Popen(['xdg-open', path])
    except Exception as e:
        st.error(f"ファイルを開けませんでした: {e}")

def save_file_to_local(filename, data, export_dir=None):
    """ ローカル環境専用: ファイルを直接ディスクに保存して絶対パスを返す """
    if not export_dir:
        export_dir = os.path.join(os.getcwd(), "exports")
    
    os.makedirs(export_dir, exist_ok=True)
    filepath = os.path.abspath(os.path.join(export_dir, filename))
    
    mode = "wb" if isinstance(data, bytes) else "w"
    encoding = None if isinstance(data, bytes) else "utf-8"
    
    with open(filepath, mode, encoding=encoding) as f:
        f.write(data)
    return filepath

def render_local_save_ui(button_label, filename, data, key_prefix, button_type="primary"):
    """ ローカル保存ボタンと「開く」アクションを統合したUIコンポーネント """
    save_btn = st.button(button_label, type=button_type, width="stretch", key=f"btn_save_{key_prefix}")
    local_export_dir = st.session_state.get('local_export_dir', None)
    
    if save_btn:
        saved_path = save_file_to_local(filename, data, local_export_dir)
        st.session_state[f'saved_path_{key_prefix}'] = saved_path
        
    # 保存成功後のUI表示
    if st.session_state.get(f'saved_path_{key_prefix}'):
        saved_path = st.session_state[f'saved_path_{key_prefix}']
        st.success(f"📂 **保存完了:** `{saved_path}`")
        
        col_open1, col_open2 = st.columns(2)
        with col_open1:
            if st.button("📄 ファイルを開く (Excel/ブラウザ等)", key=f"btn_open_file_{key_prefix}", width="stretch"):
                open_local_path(saved_path)
        with col_open2:
            if st.button("📁 フォルダを開く (保存先)", key=f"btn_open_dir_{key_prefix}", width="stretch"):
                open_local_path(os.path.dirname(saved_path))

# --- Excelグラフ生成用ライブラリ ---
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import BarChart, Reference, Series
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.axis import ChartLines
from openpyxl.chart.layout import Layout, ManualLayout
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ページ設定
st.set_page_config(layout="wide", page_title="検索大臣", page_icon="🔎")

st.markdown("""
<style>
/* グラフ描画時のフェードインアニメーション */
@keyframes fadeIn {
    from { opacity: 0.5; transform: translateY(2px); }
    to { opacity: 1; transform: translateY(0); }
}
.stAltairChart {
    animation: fadeIn 0.6s ease-out;
}
/* プログレスバーの滑らかな遷移 */
.stProgress > div > div {
    transition: width 0.5s ease-in-out !important;
}
</style>
""", unsafe_allow_html=True)

# ==========================================
#  [Local User Config] API Key Hardcoding
# ==========================================
# ローカルで利用する場合、ここにAPIキーを記述するとGUIでの入力を省略できます。
# 記述例: HARDCODED_IPINFO_KEY = "your_token_here"
HARDCODED_IPINFO_KEY = "" 
HARDCODED_VPNAPI_KEY = ""
HARDCODED_SECURITYTRAILS_KEY = ""
# ==========================================

# ==========================================
# 自動モード判定ロジック (st.secrets利用)
# ==========================================
IS_PUBLIC_MODE = False
try:
    if "ENV_MODE" in st.secrets and st.secrets["ENV_MODE"] == "public":
        IS_PUBLIC_MODE = True
except FileNotFoundError:
    IS_PUBLIC_MODE = False
# ==========================================

# --- 設定 ---
MODE_SETTINGS = {
    "安定性重視 (2.5秒待機/単一スレッド)": {
        "MAX_WORKERS": 1, 
        "DELAY_BETWEEN_REQUESTS": 2.5 
    },
    "速度優先 (1.4秒待機/2スレッド)": {
        "MAX_WORKERS": 2, 
        "DELAY_BETWEEN_REQUESTS": 1.4 
    }
}
IP_API_URL = "http://ip-api.com/json/{ip}?fields=status,country,countryCode,isp,org,query,message"
IPINFO_API_URL = "https://ipinfo.io/{ip}" 
VPNAPI_URL = "https://vpnapi.io/api/{ip}?key={key}"
RDAP_BOOTSTRAP_URL = "https://rdap.apnic.net/ip/{ip}"

RATE_LIMIT_WAIT_SECONDS = 120 

# パブリックDNSサーバーリスト (分散処理用)
PUBLIC_DNS_SERVERS = [
    '8.8.8.8', '8.8.4.4',             # Google
    '1.1.1.1', '1.0.0.1',             # Cloudflare
    '9.9.9.9', '149.112.112.112',     # Quad9
    '208.67.222.222', '208.67.220.220'# OpenDNS
]
PUBLIC_DNS_V6_SERVERS = [
    '2001:4860:4860::8888', '2001:4860:4860::8844', # Google
    '2606:4700:4700::1111', '2606:4700:4700::1001'  # Cloudflare
]
  
RIR_LINKS = {
    'RIPE': 'https://apps.db.ripe.net/db-web-ui/#/query?searchtext={ip}',
    'ARIN': 'https://search.arin.net/rdap/?query={ip}',
    'APNIC': 'https://wq.apnic.net/static/search.html',
    'JPNIC': 'https://www.nic.ad.jp/ja/whois/ja-gateway.html',
    'AFRINIC': 'https://www.afrinic.net/whois',
    'ICANN Whois': 'https://lookup.icann.org/',
}

# リンク集
SECONDARY_TOOL_BASE_LINKS = {
    'VirusTotal': 'https://www.virustotal.com/',
    'Whois.com': 'https://www.whois.com/',
    'Who.is': 'https://who.is/',
    'DomainSearch.jp': 'https://www.domainsearch.jp/',
    'Aguse': 'https://www.aguse.jp/',
    'IP2Proxy': 'https://www.ip2proxy.com/',
    'VPNAPI.io': 'https://vpnapi.io/',
    'DNS Checker': 'https://dnschecker.org/',
    'DNSlytics': 'https://dnslytics.com/',
    'IP Location': 'https://iplocation.io/',
    'CP-WHOIS': 'https://doco.cph.jp/whoisweb.php',
    }

COUNTRY_CODE_TO_RIR = {
    'JP': 'JPNIC', 'CN': 'APNIC', 'AU': 'APNIC', 'KR': 'APNIC', 'IN': 'APNIC',
    'ID': 'APNIC', 'MY': 'APNIC', 'NZ': 'APNIC', 'SG': 'APNIC',
    'TH': 'APNIC', 'VN': 'APNIC', 'PH': 'APNIC', 'PK': 'APNIC', 
    'BD': 'APNIC', 'HK': 'APNIC', 'TW': 'APNIC', 'NP': 'APNIC', 'LK': 'APNIC',
    'MO': 'APNIC', 
    'US': 'ARIN', 'CA': 'ARIN',
    'ZA': 'AFRINIC', 'EG': 'AFRINIC', 'NG': 'AFRINIC',
    'KE': 'AFRINIC', 'DZ': 'AFRINIC', 'MA': 'AFRINIC', 'GH': 'AFRINIC', 
    'CM': 'AFRINIC', 'TN': 'AFRINIC', 'ET': 'AFRINIC', 'TZ': 'AFRINIC',
    'DE': 'RIPE', 'FR': 'RIPE', 'GB': 'RIPE', 'RU': 'RIPE',
    'NL': 'RIPE', 'IT': 'RIPE', 'ES': 'RIPE', 'PL': 'RIPE', 
    'TR': 'RIPE', 'UA': 'RIPE', 'SA': 'RIPE', 'IR': 'RIPE', 
    'CH': 'RIPE', 'SE': 'RIPE', 'NO': 'RIPE', 'DK': 'RIPE', 
    'BE': 'RIPE', 'AT': 'RIPE', 'GR': 'RIPE', 'PT': 'RIPE',
    'IE': 'RIPE', 'FI': 'RIPE', 'CZ': 'RIPE', 'RO': 'RIPE',
    'HU': 'RIPE', 'IL': 'RIPE', 'KZ': 'RIPE', 'BG': 'RIPE',
    'HR': 'RIPE', 'RS': 'RIPE', 'AE': 'RIPE', 'QA': 'RIPE',
}

COUNTRY_CODE_TO_NUMERIC_ISO = {
    'AF': 4, 'AL': 8, 'DZ': 12, 'AS': 16, 'AD': 20, 'AO': 24, 'AI': 660, 'AQ': 10, 'AG': 28, 'AR': 32,
    'AM': 51, 'AW': 533, 'AU': 36, 'AT': 40, 'AZ': 31, 'BS': 44, 'BH': 48, 'BD': 50, 'BB': 52, 'BY': 112,
    'BE': 56, 'BZ': 84, 'BJ': 204, 'BM': 60, 'BT': 64, 'BO': 68, 'BA': 70, 'BW': 72, 'BV': 74, 'BR': 76,
    'VG': 92, 'IO': 86, 'BN': 96, 'BG': 100, 'BF': 854, 'BI': 108, 'KH': 116, 'CM': 120, 'CA': 124, 'CV': 132,
    'KY': 136, 'CF': 140, 'TD': 148, 'CL': 152, 'CN': 156, 'CX': 162, 'CC': 166, 'CO': 170, 'KM': 174, 'CG': 178,
    'CD': 180, 'CK': 184, 'CO': 170, 'CR': 188, 'HR': 191, 'CU': 192, 'CY': 196, 'CZ': 203, 'DK': 208, 'DJ': 262, 'DM': 212,
    'DO': 214, 'EC': 218, 'EG': 818, 'SV': 222, 'GQ': 226, 'ER': 232, 'EE': 233, 'ET': 231, 'FK': 238, 'FO': 234,
    'FJ': 242, 'FI': 246, 'FR': 250, 'GF': 254, 'PF': 258, 'TF': 260, 'GA': 266, 'GM': 270, 'GE': 268, 'DE': 276,
    'GH': 288, 'GI': 292, 'GR': 300, 'GL': 304, 'GD': 308, 'GP': 312, 'GU': 316, 'GT': 320, 'GN': 324, 'GW': 624,
    'GY': 328, 'HT': 332, 'HM': 334, 'VA': 336, 'HN': 340, 'HK': 344, 'HU': 348, 'IS': 352, 'IN': 356, 'ID': 360,
    'IR': 364, 'IQ': 368, 'IE': 372, 'IL': 376, 'IT': 380, 'CI': 384, 'JM': 388, 'JP': 392, 'JO': 400, 'KZ': 398,
    'KE': 404, 'KI': 296, 'KP': 408, 'KR': 410, 'KW': 414, 'KG': 417, 'LA': 418, 'LV': 428, 'LB': 422, 'LS': 426,
    'LR': 430, 'LY': 434, 'LI': 438, 'LT': 440, 'LU': 442, 'MO': 446, 'MK': 807, 'MG': 450, 'MW': 454, 'MY': 458,
    'MV': 462, 'ML': 466, 'MT': 470, 'MH': 584, 'MQ': 474, 'MR': 478, 'MU': 480, 'YT': 175, 'MX': 484, 'FM': 583,
    'MD': 498, 'MC': 492, 'MN': 496, 'MS': 500, 'MA': 504, 'MZ': 508, 'MM': 104, 'NA': 516, 'NR': 520, 'NP': 524,
    'NL': 528, 'AN': 530, 'NC': 540, 'NZ': 554, 'NI': 558, 'NE': 562, 'NG': 566, 'NU': 570, 'NF': 574, 'MP': 580,
    'NO': 578, 'OM': 512, 'PK': 586, 'PW': 585, 'PS': 275, 'PA': 591, 'PG': 598, 'PY': 600, 'PE': 604, 'PH': 608,
    'PN': 612, 'PL': 616, 'PT': 620, 'PR': 630, 'QA': 634, 'RE': 638, 'RO': 642, 'RU': 643, 'RW': 646, 'SH': 654,
    'KN': 659, 'LC': 662, 'PM': 666, 'VC': 670, 'WS': 882, 'SM': 674, 'ST': 678, 'SA': 682, 'SN': 686, 'RS': 688,
    'SC': 690, 'SL': 694, 'SG': 702, 'SK': 703, 'SI': 705, 'SB': 90, 'SO': 706, 'ZA': 710, 'GS': 239, 'ES': 724,
    'LK': 144, 'SD': 736, 'SR': 740, 'SJ': 744, 'SZ': 748, 'SE': 752, 'CH': 756, 'SY': 760, 'TW': 158, 'TJ': 762,
    'TZ': 834, 'TH': 764, 'TL': 626, 'TG': 768, 'TK': 772, 'TO': 776, 'TT': 780, 'TN': 788, 'TR': 792, 'TM': 795,
    'TC': 796, 'TV': 798, 'UG': 800, 'UA': 804, 'AE': 784, 'GB': 826, 'US': 840, 'UM': 581, 'UY': 858, 'UZ': 860,
    'VU': 548, 'VE': 862, 'VN': 704, 'VI': 850, 'WF': 876, 'EH': 732, 'YE': 887, 'ZM': 894, 'ZW': 716
}

# --- COUNTRY_JP_NAME 全体 ---
COUNTRY_JP_NAME = {
    "AF": "アフガニスタン・イスラム首長国","AL": "アルバニア共和国","DZ": "アルジェリア民主人民共和国","AS": "アメリカ領サモア","AD": "アンドラ公国","AO": "アンゴラ共和国",
    "AI": "アンギラ","AQ": "南極","AG": "アンティグア・バーブーダ","AR": "アルゼンチン共和国","AM": "アルメニア共和国","AW": "アルバ","AU": "オーストラリア連邦",
    "AT": "オーストリア共和国","AZ": "アゼルバイジャン共和国","BS": "バハマ国","BH": "バーレーン王国","BD": "バングラデシュ人民共和国","BB": "バルバドス","BY": "ベラルーシ共和国",
    "BE": "ベルギー王国","BZ": "ベリーズ","BJ": "ベナン共和国","BM": "バミューダ","BT": "ブータン王国","BO": "ボリビア多民族国","BA": "ボスニア・ヘルツェゴビナ",
    "BW": "ボツワナ共和国","BR": "ブラジル連邦共和国","BN": "ブルネイ・ダルサラーム国","BG": "ブルガリア共和国","BF": "ブルキナファソ","BI": "ブルンジ共和国","KH": "カンボジア王国","CM": "カメルーン共和国",
    "CA": "カナダ","CV": "カーボベルデ共和国","CF": "中央アフリカ共和国","TD": "チャド共和国","CL": "チリ共和国","CN": "中華人民共和国","CO": "コロンビア共和国","CR": "コスタリカ共和国",
    "HR": "クロアチア共和国","CU": "キューバ共和国","CY": "キプロス共和国","CZ": "チェコ共和国","DK": "デンマーク王国","DJ": "ジブチ共和国","DM": "ドミニカ国","DO": "ドミニカ共和国",
    "EC": "エクアドル共和国","EG": "エジプト・アラブ共和国","SV": "エルサルバドル共和国","EE": "エストニア共和国","ET": "エチオピア連邦民主共和国","FI": "フィンランド共和国","FR": "フランス共和国","DE": "ドイツ連邦共和国",
    "GR": "ギリシャ共和国","GL": "グリーンランド","GT": "グアテマラ共和国","GY": "ガイアナ共和国","HK": "中華人民共和国香港特別行政区","HU": "ハンガリー","IN": "インド共和国","ID": "インドネシア共和国",
    "IR": "イラン・イスラム共和国","IQ": "イラク共和国","IE": "アイルランド","IL": "イスラエル国","IT": "イタリア共和国","JP": "日本","KR": "大韓民国","TW": "台湾","MY": "マレーシア",
    "MX": "メキシコ合衆国","NL": "オランダ王国","NZ": "ニュージーランド","NO": "ノルウェー王国","PK": "パキスタン・イスラム共和国","PA": "パナマ共和国","PE": "ペルー共和国","PH": "フィリピン共和国",
    "PL": "ポーランド共和国","PT": "ポルトガル共和国","QA": "カタール国","RO": "ルーマニア","RU": "ロシア連邦","SA": "サウジアラビア王国","SG": "シンガポール共和国","ZA": "南アフリカ共和国",
    "ES": "スペイン王国","SE": "スウェーデン王国","CH": "スイス連邦","TH": "タイ王国","TR": "トルコ共和国","UA": "ウクライナ","AE": "アラブ首長国連邦","GB": "グレートブリテン及び北アイルランド連合王国",
    "US": "アメリカ合衆国","VN": "ベトナム社会主義共和国","YE": "イエメン共和国","ZM": "ザンビア共和国","ZW": "ジンバブエ共和国"
}

# --- TLD (Top Level Domain) 情報辞書 ---
TLD_INFO = {
    "ru": {"name": "Russian Federation", "jp_name": "ロシア連邦", "url": "https://cctld.ru/en"},
    "cn": {"name": "China", "jp_name": "中華人民共和国", "url": "https://www.cnnic.cn/"},
    "jp": {"name": "Japan", "jp_name": "日本", "url": "https://jprs.jp/"},
    "kr": {"name": "Republic of Korea", "jp_name": "大韓民国", "url": "https://kisa.or.kr/"},
    "kp": {"name": "Democratic People's Republic of Korea", "jp_name": "北朝鮮", "url": "N/A"},
    "tw": {"name": "Taiwan", "jp_name": "台湾", "url": "https://www.twnic.tw/"},
    "hk": {"name": "Hong Kong", "jp_name": "香港", "url": "https://www.hkirc.hk/"},
    "us": {"name": "United States", "jp_name": "アメリカ合衆国", "url": "https://www.about.us/"},
    "uk": {"name": "United Kingdom", "jp_name": "イギリス", "url": "https://www.nominet.uk/"},
    "io": {"name": "British Indian Ocean Territory", "jp_name": "英領インド洋地域 (IT系多用)", "url": "https://www.nic.io/"},
    "co": {"name": "Colombia", "jp_name": "コロンビア (企業多用)", "url": "https://www.cointernet.com.co/"},
    "tv": {"name": "Tuvalu", "jp_name": "ツバル (メディア多用)", "url": "https://www.nic.tv/"},
    "com": {"name": "Commercial", "jp_name": "商用組織 (VeriSign)", "url": "https://www.verisign.com/"},
    "net": {"name": "Network", "jp_name": "ネットワーク組織 (VeriSign)", "url": "https://www.verisign.com/"},
    "org": {"name": "Organization", "jp_name": "非営利組織 (PIR)", "url": "https://pir.org/"},
    "info": {"name": "Information", "jp_name": "情報提供 (Identity Digital)", "url": "https://identity.digital/"},
    "biz": {"name": "Business", "jp_name": "ビジネス (GoDaddy)", "url": "https://www.go.co/"},
    "xyz": {"name": "General", "jp_name": "一般 (XYZ.COM)", "url": "https://gen.xyz/"},
    "top": {"name": "General", "jp_name": "一般 (.TOP Registry)", "url": "https://www.nic.top/"},
}

# --- ISP名称の日本語マッピング (企業名統一版) ---
ISP_JP_NAME = {
    # --- NTT Group ---
    'NTT Communications Corporation': 'NTTドコモビジネス株式会社', 
    'NTT COMMUNICATIONS CORPORATION': 'NTTドコモビジネス株式会社',
    'NTT DOCOMO BUSINESS,Inc.': 'NTTドコモビジネス株式会社',
    'NTT DOCOMO, INC.': '株式会社NTTドコモ',
    'NTT PC Communications, Inc.': 'NTTPCコミュニケーションズ株式会社',
    'NTT Communications Corporation / EDION': 'OCN(NTTドコモビジネス株式会社)',
    'BP-DOUJIMA': 'エヌ・ティ・ティ・ブロードバンドプラットフォーム株式会社',
    
    # --- KDDI Group ---
    'Kddi Corporation': 'KDDI株式会社',
    'KDDI CORPORATION': 'KDDI株式会社',
    'Chubu Telecommunications Co., Inc.': '中部テレコミュニケーション株式会社',
    'Chubu Telecommunications Company, Inc.': '中部テレコミュニケーション株式会社',
    'Hokkaido Telecommunication Network Co., Inc.': '北海道総合通信網株式会社',
    'Energia Communications, Inc.': '株式会社エネルギア・コミュニケーションズ',
    'STNet, Inc.': '株式会社STNet',
    'QTNet, Inc.': '株式会社QTnet',
    'BIGLOBE Inc.': 'ビッグローブ株式会社',
    'Wire and Wireless Co., Ltd.': '株式会社ワイヤ・アンド・ワイヤレス',
    
    # --- SoftBank Group ---
    'SoftBank Corp.': 'ソフトバンク株式会社',
    'Yahoo Japan Corporation': 'LINEヤフー株式会社',
    'LY Corporation': 'LINEヤフー株式会社',
    'LINE Corporation': 'LINEヤフー株式会社',
    
    # --- Rakuten Group ---
    'Rakuten Group, Inc.': '楽天グループ株式会社',
    'Rakuten Mobile, Inc.': '楽天モバイル株式会社',
    'Rakuten Communications Corp.': '楽天コミュニケーションズ株式会社',
    
    # --- Sony Group ---
    'Sony Network Communications Inc.': 'ソニーネットワークコミュニケーションズ株式会社',
    'So-net Entertainment Corporation': 'ソニーネットワークコミュニケーションズ株式会社', 
    'So-net Corporation': 'ソニーネットワークコミュニケーションズ株式会社',
    
    # --- Major ISPs / VNEs ---
    'Internet Initiative Japan Inc.': '株式会社インターネットイニシアティブ',
    'NIFTY Corporation': 'ニフティ株式会社',
    'FreeBit Co., Ltd.': 'フリービット株式会社',
    'TOKAI Communications Corporation': '株式会社TOKAIコミュニケーションズ',
    'DREAM TRAIN INTERNET INC.': '株式会社ドリーム・トレイン・インターネット',
    'ASAHI Net, Inc.': '株式会社朝日ネット',
    'Asahi Net': '株式会社朝日ネット',
    'Optage Inc.': '株式会社オプテージ',
    'Jupiter Telecommunications Co., Ltd.': 'JCOM株式会社', 
    'JCOM Co., Ltd.': 'JCOM株式会社',
    'JCN': 'JCOM株式会社', 
    'SAKURA Internet Inc.': 'さくらインターネット株式会社',
    'GMO Internet, Inc.': 'GMOインターネットグループ株式会社',
    'INTERNET MULTIFEED CO.': 'インターネットマルチフィード株式会社',
    'IDC Frontier Inc.': '株式会社アイディーシーフロンティア',
    
    # --- Others ---
    'ARTERIA Networks Corporation': 'アルテリア・ネットワークス株式会社',
    'UCOM Corporation': 'アルテリア・ネットワークス株式会社',
    'VECTANT Ltd.': 'アルテリア・ネットワークス株式会社',
    'KIBI Cable Television Co., Ltd.': '株式会社吉備ケーブルテレビ',
    'LogicLinks, Inc.': '株式会社LogicLinks',
}

# 強力な名寄せルール (部分一致検索)
ISP_REMAP_RULES = [
    ('jcn', 'JCOM株式会社'), ('jupiter', 'JCOM株式会社'), ('cablenet', 'JCOM株式会社'),
    ('dion', 'KDDI株式会社'), ('au one', 'KDDI株式会社'), ('kddi', 'KDDI株式会社'),
    ('k-opti', 'オプテージ株式会社'), ('ctc', '中部テレコミュニケーションズ株式会社'),
    ('vectant', 'アルテリア・ネットワークス株式会社'), ('arteria', 'アルテリア・ネットワークス株式会社'),('v-vne', 'アルテリア・ネットワークス株式会社'),
    ('softbank', 'ソフトバンク株式会社'), ('bbtec', 'ソフトバンク株式会社'),
    ('ocn', 'OCN(NTTドコモビジネス株式会社)'), ('nifty', 'ニフティ株式会社'), ('asahi net', '株式会社朝日ネット'),
    ('rakuten mobile', '楽天モバイル株式会社'),('rmn', '楽天モバイル株式会社'), ('rakuten communications', '楽天コミュニケーションズ株式会社'),
    ('so-net', 'ソニーネットワークコミュニケーションズ株式会社'), ('nuro', 'ソニー (NURO)'),
    ('biglobe', 'ビッグローブ株式会社'), ('iij', '株式会社インターネットイニシアティブ(IIJ)'),
    ('transix', 'インターネットマルチフィード株式会社 (transix)'),
    ('v6plus', 'JPNE (v6プラス)'),
    ('logiclinks', '株式会社LogicLinks'),('lgls', '株式会社LogicLinks'),
    ('plala', '株式会社NTTドコモ (ぷらら)'),('docomo', '株式会社NTTドコモ'),('maps', '株式会社NTTドコモ'),
    ('wi2', '株式会社ワイヤ・アンド・ワイヤレス'),
]

def normalize_isp_key(text):
    if not text: return ""
    return text.lower().replace(',', '').replace('.', '').strip()

ISP_JP_NAME_NORMALIZED = {normalize_isp_key(k): v for k, v in ISP_JP_NAME.items()}

# --- 匿名化・プロキシ判定用データ ---

@st.cache_data(ttl=86400, show_spinner=False, max_entries=10)
def fetch_tor_exit_nodes():
    try:
        url = "https://check.torproject.org/exit-addresses"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        return set([line.split()[1] for line in response.text.splitlines() if line.startswith("ExitAddress")])
    except requests.exceptions.RequestException as e:
        import logging
        logging.warning(f"Tor出口ノードリストの取得に失敗しました: {e}")
        return set()

@st.cache_data(ttl=86400*3, show_spinner=False)
def fetch_cloud_ip_ranges():
    """ 主要クラウドプロバイダの公式IPレンジ(JSON)を動的に取得し、二分探索用に最適化する """
    cloud_ranges_v4 = []
    cloud_ranges_v6 = []

    def add_range(cidr_str, provider):
        try:
            net = ipaddress.ip_network(cidr_str, strict=False)
            if net.version == 4:
                cloud_ranges_v4.append((int(net.network_address), int(net.broadcast_address), provider))
            else:
                cloud_ranges_v6.append((int(net.network_address), int(net.broadcast_address), provider))
        except Exception:
            pass

    # 1. AWS (Amazon Web Services)
    try:
        r = requests.get("https://ip-ranges.amazonaws.com/ip-ranges.json", timeout=10)
        if r.status_code == 200:
            data = r.json()
            for prefix in data.get("prefixes", []): add_range(prefix.get("ip_prefix"), "AWS")
            for prefix in data.get("ipv6_prefixes", []): add_range(prefix.get("ipv6_prefix"), "AWS")
    except: pass

    # 2. GCP (Google Cloud Platform)
    try:
        r = requests.get("https://www.gstatic.com/ipranges/cloud.json", timeout=10)
        if r.status_code == 200:
            data = r.json()
            for prefix in data.get("prefixes", []):
                if "ipv4Prefix" in prefix: add_range(prefix["ipv4Prefix"], "GCP")
                if "ipv6Prefix" in prefix: add_range(prefix["ipv6Prefix"], "GCP")
    except: pass

    # 3. Azure (Microsoft Download Centerをスクレイピングして動的URLを取得)
    try:
        dl_page = requests.get("https://www.microsoft.com/en-us/download/confirmation.aspx?id=56519", timeout=10)
        match = re.search(r'href="(https://download\.microsoft\.com/download/.*?/ServiceTags_Public_.*?\.json)"', dl_page.text)
        if match:
            azure_url = match.group(1)
            r = requests.get(azure_url, timeout=15)
            if r.status_code == 200:
                data = r.json()
                for val in data.get("values", []):
                    for prefix in val.get("properties", {}).get("addressPrefixes", []):
                        add_range(prefix, "Azure")
    except: pass

    # 4. Cloudflare (Reverse Proxy / WAF)
    try:
        r_v4 = requests.get("https://www.cloudflare.com/ips-v4", timeout=5)
        if r_v4.status_code == 200:
            for line in r_v4.text.splitlines(): add_range(line.strip(), "Cloudflare")
        r_v6 = requests.get("https://www.cloudflare.com/ips-v6", timeout=5)
        if r_v6.status_code == 200:
            for line in r_v6.text.splitlines(): add_range(line.strip(), "Cloudflare")
    except: pass

    # 二分探索(O(log N))できるように開始IPの整数値でソート
    cloud_ranges_v4.sort(key=lambda x: x[0])
    cloud_ranges_v6.sort(key=lambda x: x[0])

    return {"v4": cloud_ranges_v4, "v6": cloud_ranges_v6}

def check_cloud_provider(ip_str, cloud_data):
    """ IPアドレスがクラウド事業者の公式リストに含まれているかを超高速で判定する """
    if not cloud_data: return None
    try:
        ip_obj = ipaddress.ip_address(ip_str)
        ip_int = int(ip_obj)
        target_list = cloud_data["v4"] if ip_obj.version == 4 else cloud_data["v6"]

        # 二分探索で「開始IPが探しているIP以下の最大のインデックス」を見つける
        keys = [r[0] for r in target_list]
        idx = bisect.bisect_right(keys, ip_int) - 1

        if idx >= 0:
            start_ip, end_ip, provider = target_list[idx]
            if start_ip <= ip_int <= end_ip:
                return provider
    except:
        pass
    return None

@st.cache_data(ttl=86400, show_spinner=False, max_entries=10)
def fetch_disposable_domains():
    """ GitHubの有名リポジトリから最新の捨てアドドメイン一覧を取得 (1日1回更新) """
    try:
        url = "https://raw.githubusercontent.com/disposable-email-domains/disposable-email-domains/master/disposable_email_blocklist.conf"
        response = requests.get(url, timeout=10)
        response.raise_for_status()
        if response.status_code == 200:
            # 空行とコメントを除外し、小文字でセット（集合）に格納して高速化
            return set([line.strip().lower() for line in response.text.splitlines() if line.strip() and not line.startswith('//')])
    except requests.exceptions.RequestException as e:
        import logging
        logging.warning(f"使い捨てドメインリストの取得に失敗しました: {e}")
    return set()



# --- 捨てアド (Disposable Email) 検知用グローバル辞書 ---
DISPOSABLE_MX_SERVICES = {
    'sute.jp': '捨てメアド (メルアドぽいぽい)',
    'erinn.biz': '捨てメアド (メルアドぽいぽい)',
    'kuku.lu': '捨てメアド (メルアドぽいぽい)',
    'instaddr.com': '捨てメアド (メルアドぽいぽい)',
    'instaddr.jp': '捨てメアド (メルアドぽいぽい)',
    'm.miril.jp': '捨てメアド (メルアドぽいぽい)',
    '10minutemail': '10 Minute Mail',
    'guerrillamail': 'Guerrilla Mail',
    'temp-mail': 'Temp Mail',
    'nada.email': 'Nada / Tmpmail',
    'maildrop.cc': 'Maildrop',
    'yopmail.com': 'YOPmail',
    'tempmail.plus': 'Temp Mail Plus',
    '1secmail.com': '1SecMail',
    'throwawaymail.com': 'Throwaway Mail',
    'tempmail.org': 'Temp-Mail.org',
    'mail.tm': 'Mail.tm',
    'sharklasers.com': 'Guerrilla Mail',
    'dispostable.com': 'Dispostable',
    'getnada.com': 'Nada.email',
    'mailinator.com': 'Mailinator',
    'moakt.com': 'Moakt',
    'tmails.net': 'T-Mails',
    '33mail.com': '33mail',
    'airmail.cc': 'Airmail',
    'generator.email': 'Generator.email'
}

DISPOSABLE_DOMAIN_SERVICES = {
    'instaddr.com': '捨てメアド (メルアドぽいぽい)',
    'instaddr.jp': '捨てメアド (メルアドぽいぽい)',
    'm.miril.jp': '捨てメアド (メルアドぽいぽい)',
    '10minutemail.com': '10 Minute Mail',
    '10minutemail.net': '10 Minute Mail',
    'guerrillamail.com': 'Guerrilla Mail',
    'mailinator.com': 'Mailinator',
    'yopmail.com': 'YOPmail'
}

def check_disposable_domain(domain, nslookup_raw):
    """ MXレコードやドメイン名から捨てアドサービスを検知し、特定されたサービス名のリストを返す """
    dynamic_disposable_list = fetch_disposable_domains()
    detected_services = []
    raw_lower = nslookup_raw.lower() if nslookup_raw else ""
    query_domain_lower = domain.lower()
    
    import re
    mx_targets = re.findall(r'\bin\s+mx\s+\d+\s+(\S+)', raw_lower)
    targets_to_check = mx_targets + [query_domain_lower]
    
    # 1. 既知の辞書を使った特定
    for target in targets_to_check:
        target = target.strip('.')
        for pattern, service_name in DISPOSABLE_MX_SERVICES.items():
            if pattern in target and service_name not in detected_services:
                detected_services.append(service_name)
        for pattern, service_name in DISPOSABLE_DOMAIN_SERVICES.items():
            if pattern in target and service_name not in detected_services:
                detected_services.append(service_name)
                
    # 2. 外部DB（GitHubリスト）による特定不可ドメインの捕捉
    if not detected_services and dynamic_disposable_list:
        for target in targets_to_check:
            target = target.strip('.')
            parts = target.split('.')
            for i in range(len(parts) - 1): 
                domain_to_check = '.'.join(parts[i:])
                if domain_to_check in dynamic_disposable_list:
                    label = f"外部DB検知 ({domain_to_check} / サービス名特定不可)"
                    if label not in detected_services:
                        detected_services.append(label)
                    break 
    return detected_services   

def get_jp_names(english_isp, country_code):
    jp_country = COUNTRY_JP_NAME.get(country_code, country_code)
    
    if not english_isp:
        return "N/A", jp_country

    normalized_input = normalize_isp_key(english_isp)
    jp_isp = english_isp 

    if english_isp in ISP_JP_NAME:
        jp_isp = ISP_JP_NAME[english_isp]
    elif normalized_input in ISP_JP_NAME_NORMALIZED:
        jp_isp = ISP_JP_NAME_NORMALIZED[normalized_input]
    else:
        for keyword, mapped_name in ISP_REMAP_RULES:
            # 単語の境界(\b)を判定し、edionの中のdion等、意図しない部分文字列へのマッチを排除する
            if re.search(rf'\b{re.escape(keyword)}\b', normalized_input):
                jp_isp = mapped_name
                break
        
    jp_country = COUNTRY_JP_NAME.get(country_code, country_code)
    return jp_isp, jp_country

from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

@st.cache_resource
def get_session():
    session = requests.Session()
    session.headers.update({"User-Agent": "WhoisBatchTool/2.4 (+RDAP)"})
    
    # ネットワーク瞬断に対応するための自動リトライ機能 (3回, バックオフ)
    retries = Retry(total=3, backoff_factor=1, status_forcelist=[500, 502, 503, 504])
    adapter = HTTPAdapter(max_retries=retries)
    session.mount("http://", adapter)
    session.mount("https://", adapter)
    
    return session

session = get_session()

@st.cache_data(max_entries=10)
def get_world_map_data():
    try:
        world_geojson = alt.topo_feature('https://cdn.jsdelivr.net/npm/vega-datasets@v1.29.0/data/world-110m.json', 'countries')
        return world_geojson
    except:
        return None

WORLD_MAP_GEOJSON = get_world_map_data()

# --- ヘルパー関数群 ---

def extract_actual_ip(target):
    """ 'ドメイン (IP)' の形式からIPアドレスだけを抽出する関数 """
    if not isinstance(target, str): return target
    if "(" in target and ")" in target:
        possible_ip = target.split("(")[-1].replace(")", "").strip()
        try:
            ipaddress.ip_address(possible_ip)
            return possible_ip
        except ValueError:
            pass
    return target

def clean_ocr_error_chars(target):
    cleaned_target = target.replace('Ⅱ', '11').replace('I', '1').replace('l', '1').replace('|', '1').replace('O', '0').replace('o', '0').replace(';', '.').replace(',', '.')
    if ':' not in cleaned_target:
        cleaned_target = cleaned_target.replace('S', '5').replace('s', '5')
    return cleaned_target

def is_valid_ip(target):
    try:
        ipaddress.ip_address(extract_actual_ip(target))
        return True
    except ValueError:
        return False

def is_valid_domain(target):
    """ 入力された文字列が有効なFQDN（ドメイン名）の形式を満たしているか判定する """
    if not isinstance(target, str): return False
    if is_valid_ip(target): return False
    if '.' not in target or target.startswith('.') or target.endswith('.'): return False
    if re.search(r'\s', target): return False
    parts = target.split('.')
    if len(parts) < 2 or not parts[-1].isalpha() or len(parts[-1]) < 2: return False
    return True

def is_ipv4(target):
    try:
        ipaddress.IPv4Address(extract_actual_ip(target))
        return True
    except ValueError:
        return False

def ip_to_int(ip):
    actual_ip = extract_actual_ip(ip)
    try:
        if is_ipv4(actual_ip):
            return struct.unpack("!I", socket.inet_aton(actual_ip))[0]
        return 0
    except OSError:
        return 0

def get_cidr_block(ip, netmask_range=(8, 24)):
    actual_ip = extract_actual_ip(ip)
    try:
        ip_obj = ipaddress.ip_address(actual_ip)
        if ip_obj.version == 4:
            netmask = netmask_range[1] 
            network = ipaddress.ip_network(f'{actual_ip}/{netmask}', strict=False)
            return str(network)
        elif ip_obj.version == 6:
            netmask = 48
            network = ipaddress.ip_network(f'{actual_ip}/{netmask}', strict=False)
            return str(network)
        return None
    except ValueError:
        return None

def get_authoritative_rir_link(ip, country_code):
    rir_name = COUNTRY_CODE_TO_RIR.get(country_code)
    # RIR共通のポップアップ説明文
    desc = "IPアドレスを管轄する公式レジストリ。法的な保有組織などの最も正確な情報を確認できます。"
    
    if rir_name and rir_name in RIR_LINKS:
        encoded_ip = quote(ip, safe='')
        if rir_name in ['RIPE', 'ARIN']:
            link_url = RIR_LINKS[rir_name].format(ip=encoded_ip)
            return f"[{rir_name} ]({link_url} \"{desc}\")"
        elif rir_name in ['JPNIC', 'APNIC', 'LACNIC', 'AFRINIC']:
            link_url = RIR_LINKS[rir_name]  
            return f"[{rir_name} (手動検索) ]({link_url} \"{desc}\")"
    return f"[Whois (汎用検索) ]({RIR_LINKS.get('APNIC', 'https://wq.apnic.net/static/search.html')} \"{desc}\")"

def get_copy_target(ip_display):
    if not ip_display: return ""
    return str(ip_display).split(' - ')[0].split(' ')[0]

def create_secondary_links(target):
    actual_ip = extract_actual_ip(target)
    is_composite = (actual_ip != target and "(" in target) # ドメインとIPの複合型か判定
    is_ip = is_valid_ip(target) and not is_composite
    
    # --- ツールごとの解説文（オンマウス時のツールチップ用） ---
    tool_tips = {
        'VirusTotal': '世界中のウイルス対策エンジンで一括スキャン。危険なIPか即座に判別。',
        'Aguse': '日本語表示。ブラックリスト判定や、サーバー証明書情報が見やすい。',
        'Aguse (Domain)': '日本語表示。ブラックリスト判定や、サーバー証明書情報が見やすい。',
        'ipinfo.io': '地図上の位置、ホスティング(クラウド)かどうかの詳細判定に強い。',
        'IP2Proxy': '匿名プロキシやVPNからのアクセスかどうかを専門的に判定。',
        'VPNAPI.io': '匿名プロキシやVPNからのアクセスかどうかを専門的に判定。(本ツールでAPI実装済み)',
        'IP Location': 'IPアドレスの地理的位置をGoogleマップ等で視覚的に表示。',
        'Whois.com': 'ドメインの保有者情報（英語）を確認するのに最適。',
        'DNS Checker': 'IPv6のWhois情報が世界中でどう見えているかを確認。',
        'CP-WHOIS (手動)': '利用者認証が必要な検索ツール。ここでの検索結果はデータとして信頼性が高い。',
        'DNS History (手動)': '過去のDNSレコードの変更履歴を確認。'
    }

    links = {}

    if is_composite:
        # --- ドメイン(IP) 複合型専用 厳選リンク ---
        domain_part = target.split("(")[0].strip()
        encoded_domain = quote(domain_part, safe='')
        encoded_ip = quote(actual_ip, safe='')
        
        links['VirusTotal'] = f'https://www.virustotal.com/gui/search/{encoded_domain}'
        links['Aguse (Domain)'] = f'https://www.aguse.jp/?url={encoded_domain}'
        links['ipinfo.io'] = f'https://ipinfo.io/{encoded_ip}'
        links['IP Location'] = f'https://iplocation.io/ip/{encoded_ip}'
        links['DNS History (手動)'] = 'https://dnshistory.org/'

    elif is_ip:
        encoded_target = quote(actual_ip, safe='')
        if is_ipv4(actual_ip):
            links['VirusTotal'] = f'https://www.virustotal.com/gui/search/{encoded_target}'
            links['Aguse'] = f'https://www.aguse.jp/?url={encoded_target}'
            links['ipinfo.io'] = f'https://ipinfo.io/{encoded_target}'
            links['IP2Proxy'] = f'https://www.ip2proxy.com/{encoded_target}'
            links['VPNAPI.io'] = f'https://vpnapi.io/api/{encoded_target}'
            links['IP Location'] = f'https://iplocation.io/ip/{encoded_target}'
        else:
            links['VirusTotal'] = f'https://www.virustotal.com/gui/search/{encoded_target}'
            links['ipinfo.io'] = f'https://ipinfo.io/{encoded_target}'
            links['IP2Proxy'] = f'https://www.ip2proxy.com/{encoded_target}'
            links['VPNAPI.io'] = f'https://vpnapi.io/api/{encoded_target}'
            links['IP Location'] = f'https://iplocation.io/ip/{encoded_target}'
            links['DNS Checker'] = f'https://dnschecker.org/ipv6-whois-lookup.php?query={encoded_target}'
    else:
        # --- 純粋なドメイン用 厳選リンク (DNS解決失敗時) ---
        encoded_target = quote(target, safe='')
        links['VirusTotal'] = f'https://www.virustotal.com/gui/search/{encoded_target}'
        links['Aguse'] = f'https://www.aguse.jp/?url={encoded_target}'
        links['Whois.com'] = f'https://www.whois.com/whois/{encoded_target}'
        links['DNS History (手動)'] = 'https://dnshistory.org/'

    links['CP-WHOIS (手動)'] = 'https://doco.cph.jp/whoisweb.php'

    link_html = ""
    for name, url in links.items():
        if url: 
            # 辞書から説明文を取得（なければ空文字）
            desc = tool_tips.get(name, "")
            link_html += f"[{name} ]({url} \"{desc}\") | "
    
    return link_html.rstrip(' | ')

# RDAPデータ取得関数 (公式台帳への照会)
def fetch_rdap_data(ip):
    try:
        url = RDAP_BOOTSTRAP_URL.format(ip=ip)
        # 海外レジストリ(AFRINIC等)の遅延を考慮し、タイムアウトを8秒に設定
        response = session.get(url, timeout=8, allow_redirects=True)
        response.raise_for_status()
        if response.status_code == 200:
            data = response.json()
            # 汎用的なRDAPレスポンスから名前を探す (name, handle, remarks)
            network_name = data.get('name', '')
            if not network_name and 'handle' in data:
                network_name = data['handle']
            return {'name': network_name, 'json': data, 'url': url}
    except requests.exceptions.Timeout:
        pass
    except requests.exceptions.RequestException:
        pass
    except ValueError:
        pass
    return None

# ドメイン専用RDAP取得関数
@st.cache_data(ttl=3600, show_spinner=False, max_entries=1000)
def fetch_domain_rdap_data(domain):
    """ ドメイン専用のRDAP情報を取得する関数 (rdap.org リゾルバを利用) """
    try:
        url = f"https://rdap.org/domain/{domain}"
        response = session.get(url, timeout=8, allow_redirects=True)
        response.raise_for_status()
        if response.status_code == 200:
            data = response.json()
            return {'json': data, 'url': response.url}
    except requests.exceptions.Timeout:
        pass
    except requests.exceptions.RequestException:
        pass
    except ValueError:
        pass
    return None

@st.cache_data(ttl=3600, show_spinner=False, max_entries=2000)
def fetch_classic_whois(target):
    """ OS非依存：Port 43を利用した旧式WHOIS取得 (ドメイン・IP両対応) """
    import socket
    import ipaddress
    try:
        is_ip = False
        try:
            ipaddress.ip_address(target)
            is_ip = True
        except ValueError:
            pass
            
        # 1. IANAから権威WHOISサーバーを特定
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(5)
            s.connect(('whois.iana.org', 43))
            if is_ip:
                s.send((target + "\r\n").encode('utf-8'))
            else:
                tld = target.split('.')[-1].lower()
                s.send((tld + "\r\n").encode('utf-8'))
            
            response = b""
            while True:
                data = s.recv(4096)
                if not data: break
                response += data
            
        iana_response = response.decode('utf-8', errors='replace')
        whois_server = None
        for line in iana_response.splitlines():
            line_lower = line.lower()
            if line_lower.startswith('whois:'):
                whois_server = line.split(':', 1)[1].strip()
                break
            elif line_lower.startswith('refer:'):
                whois_server = line.split(':', 1)[1].strip()
                break
        
        # IANAに記載がない場合の汎用推測
        if not whois_server:
            if is_ip:
                whois_server = "whois.arin.net" # IPのフォールバック
            else:
                tld = target.split('.')[-1].lower()
                whois_server = f"{tld}.whois-servers.net" # ドメインのフォールバック
            
        # 2. 権威サーバーに直接クエリを投げる
        with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
            s.settimeout(5)
            s.connect((whois_server, 43))
            
            # JPRS (.jp) の場合、英語出力を強制するために /e を付与する
            if not is_ip and target.endswith('.jp'):
                query_str = f"{target}/e\r\n"
            else:
                query_str = f"{target}\r\n"
                
            s.send(query_str.encode('utf-8'))
            whois_text = b""
            while True:
                data = s.recv(4096)
                if not data: break
                whois_text += data
        
        # エンコーディング対応 (JPRS等のISO-2022-JP対応)
        try:
            decoded_text = whois_text.decode('iso-2022-jp').strip()
        except UnicodeDecodeError:
            decoded_text = whois_text.decode('utf-8', errors='replace').strip()
            
        if not decoded_text:
            return "Error: WHOISサーバーに接続できましたが、データが空でした（応答なし）。\n短時間での連続アクセスによる一時的なブロック（Rate Limit）の可能性が高いです。", whois_server
            
        return decoded_text, whois_server
        
    except socket.timeout:
        error_msg = "Error: WHOISサーバーからの応答がタイムアウトしました。\n短時間での連続アクセスによる一時的な制限（Rate Limit）の可能性が高いです。\nしばらく時間をおいてから再度お試しください。"
        return error_msg, whois_server if 'whois_server' in locals() and whois_server else "不明"
    except ConnectionRefusedError:
        error_msg = "Error: WHOISサーバーへの接続が拒否されました。\n接続制限、または相手方サーバーがダウンしている可能性があります。"
        return error_msg, whois_server if 'whois_server' in locals() and whois_server else "不明"
    except Exception as e:
        error_msg = f"Error: WHOIS情報の取得中にシステムエラーが発生しました ({str(e)})"
        return error_msg, "不明"

# SecurityTrails API取得関数 (過去のAレコード・AAAAレコード履歴)
def get_securitytrails_data(domain, api_key, start_date=None, end_date=None):
    """ SecurityTrails APIを使用してドメインの過去のIP履歴(IPv4/IPv6)を取得し、期間でフィルタリングする """
    if not api_key or not domain:
        return None
    
    headers = {
        "APIKEY": api_key,
        "accept": "application/json"
    }
    
    combined_records = []
    
    import requests
    # Aレコード (IPv4) 取得
    try:
        url_a = f"https://api.securitytrails.com/v1/history/{domain}/dns/a"
        res_a = session.get(url_a, headers=headers, timeout=10)
        
        # HTTPステータスコードが200番台以外なら例外を発生させる
        res_a.raise_for_status() 
        
        data_a = res_a.json()
        if "records" in data_a:
            combined_records.extend(data_a["records"])
            
    except requests.exceptions.Timeout:
        # タイムアウトした場合の処理
        # print(f"[Timeout] SecurityTrails API (A): {domain}")
        pass
    except requests.exceptions.HTTPError as e:
        # 月間制限(50回)等のレートリミット到達時、エラーフラグを返す
        if e.response is not None and e.response.status_code == 429:
            return {"error": "rate_limit"}
        pass
    except requests.exceptions.RequestException as e:
        # その他のネットワークエラー（DNS解決失敗など）
        pass
    except ValueError:
        # JSONのパースエラー（APIが想定外のHTMLなどを返してきた場合）
        pass

    # AAAAレコード (IPv6) 取得
    try:
        url_aaaa = f"https://api.securitytrails.com/v1/history/{domain}/dns/aaaa"
        res_aaaa = session.get(url_aaaa, headers=headers, timeout=10)
        
        # HTTPステータスコードが200番台以外なら例外を発生させる
        res_aaaa.raise_for_status()
        
        data_aaaa = res_aaaa.json()
        if "records" in data_aaaa:
            combined_records.extend(data_aaaa["records"])
            
    except requests.exceptions.Timeout:
        # タイムアウトした場合の処理
        # print(f"[Timeout] SecurityTrails API (AAAA): {domain}")
        pass
    except requests.exceptions.HTTPError as e:
        # AAAAレコード側にもレートリミット到達の検知を追加
        if e.response is not None and e.response.status_code == 429:
            return {"error": "rate_limit"}
        pass
    except requests.exceptions.RequestException as e:
        # その他のネットワークエラー
        pass
    except ValueError:
        # JSONのパースエラー
        pass
    

    if combined_records:
        # まず first_seen (初回観測日) の降順で全体をソート (新しい順)
        combined_records.sort(key=lambda x: str(x.get('first_seen', '1970-01-01')), reverse=True)
        
        filtered_records = []
        is_date_filtered = False
        
        if start_date and end_date:
            is_date_filtered = True
            start_str = start_date.strftime("%Y-%m-%d")
            end_str = end_date.strftime("%Y-%m-%d")
            for rec in combined_records:
                rec_first = str(rec.get('first_seen', '9999-12-31'))
                rec_last = str(rec.get('last_seen', '1970-01-01'))
                
                # レコードの生存期間が指定された期間と重なっているかを判定
                if rec_first <= end_str and rec_last >= start_str:
                    filtered_records.append(rec)
        else:
            # 期間指定がない場合は最新20件のみを抽出
            filtered_records = combined_records[:20]

        if filtered_records:
            result_data = {
                "records": filtered_records,
                "is_date_filtered": is_date_filtered
            }
            # 日付フィルタが有効な場合、戻り値に日付の文字列を追加する
            if is_date_filtered:
                result_data["start_date"] = start_date.strftime("%Y-%m-%d")
                result_data["end_date"] = end_date.strftime("%Y-%m-%d")
            return result_data

    return None

# SecurityTrails API取得関数 (Reverse IP / ドメイン逆検索)
# fetch_allフラグを受け取り、ページネーションループを回す
def get_securitytrails_reverse_ip(ip, api_key, fetch_all=False):
    """ SecurityTrails APIを使用してIPアドレスに紐づくドメイン群を取得する """
    if not api_key or not ip:
        return None
    
    headers = {
        "APIKEY": api_key,
        "accept": "application/json",
        "content-type": "application/json"
    }
    
    ip_key = "ipv4" if is_ipv4(ip) else "ipv6"
    payload = {"filter": {ip_key: ip}}
    
    try:
        url = "https://api.securitytrails.com/v1/domains/list"
        res = session.post(url, headers=headers, json=payload, timeout=10)
        res.raise_for_status()
        data = res.json()
        
        # 全件取得オンかつ複数ページある場合、次ページを取得し続ける
        if fetch_all:
            total_pages = data.get('meta', {}).get('total_pages', 1)
            current_page = 1
            # 念のため暴走防止で最大100ページ（1万件）で頭打ちにする
            while current_page < total_pages and current_page <= 100:
                current_page += 1
                payload['page'] = current_page
                try:
                    time.sleep(1) # API制限回避のウェイト
                    res_next = session.post(url, headers=headers, json=payload, timeout=10)
                    res_next.raise_for_status()
                    data_next = res_next.json()
                    if 'records' in data_next:
                        data['records'].extend(data_next['records'])
                except requests.exceptions.HTTPError as e:
                    if e.response is not None and e.response.status_code == 429:
                        data['error'] = "rate_limit_during_pagination" # 途中で制限に達した専用フラグ
                        break
                    else:
                        break
                except Exception:
                    break
        return data
        
    except requests.exceptions.HTTPError as e:
        # 初回リクエストでのレートリミット到達検知
        if e.response is not None and e.response.status_code == 429:
            return {"error": "rate_limit"}
        return None
    except Exception:
        return None

# Shodan InternetDB API Logic (No API Key Required)
def check_internetdb_risk(ip, max_retries=3):
    """
    Shodan InternetDB APIを使用して、ポートスキャン結果と脆弱性をチェックする。
    タイムアウトによるデータ欠損を防ぐため、リトライ機構とバックオフを実装。
    """
    RISK_PORTS = {
        21: "Vuln:FTP",
        23: "Vuln:Telnet (High Risk)",
        1080: "Proxy:SOCKS",
        3128: "Proxy:Squid",
        5554: "IoT:Android/Emu",
        5555: "IoT:Android/ADB (High Risk)",
        7547: "Vuln:TR-069",
        1900: "Vuln:UPnP",
        8080: "Proxy:HTTP",
    }
    
    for attempt in range(max_retries):
        try:
            url = f"https://internetdb.shodan.io/{ip}"
            # タイムアウトを5秒に延長し、猶予を持たせる
            response = requests.get(url, timeout=5)
            
            if response.status_code == 404:
                return "[データなし]"
            elif response.status_code == 429:
                return "エラー: Shodanのアクセス制限超過"
            elif 500 <= response.status_code < 600:
                return f"エラー: Shodanサーバー側の障害 ({response.status_code})"
            elif response.status_code != 200:
                return f"エラー: Shodan通信障害 ({response.status_code})"
                
            data = response.json()
            found_risks = []
            open_ports = data.get('ports', [])
            vulns = data.get('vulns', [])
            
            for p in open_ports:
                if p in RISK_PORTS:
                    found_risks.append(RISK_PORTS[p])
            
            if vulns:
                found_risks.append(f"CVEs({len(vulns)})")
                
            if found_risks:
                unique_risks = sorted(list(set(found_risks)))
                return " / ".join(unique_risks)
            else:
                if open_ports:
                    return "[No Match (Other Ports)]"
                return "[No Match]"
            
        except requests.exceptions.ConnectionError:
            # 物理的なネットワーク切断時は上位ループへ例外を投げ、15秒待機のサーキットブレーカーを発動させる
            raise
        except requests.exceptions.Timeout:
            # 最終試行でもタイムアウトした場合のみエラーを返す
            if attempt == max_retries - 1:
                return "エラー: Shodan応答タイムアウト (サーバー混雑)"
            time.sleep(1.5) # リトライ前に1.5秒の待機を挟む（バックオフ）
        except requests.exceptions.HTTPError:
            return "エラー: Shodan通信失敗 (HTTPエラー)"
        except requests.exceptions.RequestException:
            return "エラー: ネットワーク接続に失敗しました"
        except ValueError:
            return "エラー: データ解析失敗 (相手から不正なデータが返されました)"
        
# VPNAPI.io 取得関数
def get_vpnapi_data(ip, api_key):
    """
    VPNAPI.io APIを使用してプロキシ・VPN判定の詳細データを取得する。
    """
    if not api_key:
        return None
    try:
        url = VPNAPI_URL.format(ip=ip, key=api_key)
        response = session.get(url, timeout=5)
        
        # レートリミット到達検知
        if response.status_code == 429:
            return {"error": "rate_limit"}
            
        response.raise_for_status()
        if response.status_code == 200:
            data = response.json()
            # vpnapi.io の仕様：security キーが存在するかで判定
            if "security" in data:
                return data
    except requests.exceptions.Timeout:
        pass
    except requests.exceptions.RequestException:
        pass
    except ValueError:
        pass
    return None


# IP逆引き関数 (PTRレコード取得 - dnspython使用/高信頼設定)
def resolve_ip_nslookup(ip):
    """ dnspythonを使用して、外部DNSサーバーを直接指定し、逆引き(PTR)ホスト名を取得する """
    hostnames = []
    raw_output = ""
    try:
        import dns.resolver
        import dns.reversename
        
        rev_name = dns.reversename.from_address(ip)
        
        # システムの不安定なDNS設定を回避し、信頼できる公開DNS（Google/Cloudflare）を明示的に指定
        resolver = dns.resolver.Resolver(configure=False)
        resolver.nameservers = random.sample(PUBLIC_DNS_SERVERS, 2)
        resolver.timeout = 3 # 高速応答を期待し、タイムアウトを3秒に最適化
        resolver.lifetime = 3
        
        # PTRレコードをクエリ
        answers = resolver.resolve(rev_name, 'PTR')       
       
        # 取得したレコードを処理
        raw_lines = []
        for rdata in answers:
            # 末尾のドットを削除してクリーンなホスト名を取得
            host = rdata.target.to_text(omit_final_dot=True)
            if host and host not in hostnames:
                hostnames.append(host)
            raw_lines.append(f"{rev_name} domain name pointer {host}")
            
        raw_output = "\n".join(raw_lines)
        
    except ImportError:
        raw_output = "Error: 'dnspython' ライブラリがインストールされていません。\nターミナルで 'pip install dnspython' を実行してください。"
    except dns.resolver.NXDOMAIN:
        raw_output = f"NXDOMAIN: {ip} に対するPTRレコードが見つかりませんでした。"
    except dns.resolver.NoAnswer:
        raw_output = f"NoAnswer: {ip} に対するPTRレコードの応答がありません。"
    except (dns.resolver.Timeout, dns.exception.Timeout):
        raw_output = "Error: DNSクエリがタイムアウトしました。"
    except Exception as e:
        raw_output = f"Error executing dnspython: {str(e)}"
    
    return hostnames, raw_output

# --- API通信関数 (Main) ---
def get_ip_details_from_api(ip, cidr_cache_snapshot, learned_isps_snapshot, delay_between_requests, rate_limit_wait_seconds, tor_nodes, cloud_ip_data, use_rdap, use_internetdb, use_rdns, use_st_reverse_ip, api_key=None, vpnapi_key=None, st_api_key=None, st_start_date=None, st_end_date=None, use_st_rev_fetchall=False, is_single_target=False):
    actual_ip = extract_actual_ip(ip)
    
    result = {
        'Target_IP': ip, 
        'ISP_API_Raw': 'N/A', 'ISP_JP': 'N/A', 
        'RDAP_Name_Raw': '', 'RDAP_JP': '',    
        'ISP': 'N/A', 
        'Country': 'N/A', 'Country_JP': 'N/A', 'CountryCode': 'N/A', 
        'RIR_Link': 'N/A', 'Secondary_Security_Links': 'N/A', 'Status': 'N/A',
        'RDAP_JSON': None, 'VPNAPI_JSON': None, 'RDAP_URL': '', 'IPINFO_JSON': None, 'IoT_Risk': '',
        'DOMAIN_RDAP_JSON': None, 'DOMAIN_RDAP_URL': '', 'ST_JSON': None, 'RDNS_DATA': None,
        'Proxy_Type': '', 'ST_REVERSE_IP_JSON': None,
        'DOMAIN_WHOIS_TEXT': None, 'DOMAIN_WHOIS_SERVER': None,
        'IP_WHOIS_TEXT': None, 'IP_WHOIS_SERVER': None
    }
    new_cache_entry = None
    new_learned_isp = None
    cidr_block = get_cidr_block(actual_ip)
    
    if cidr_block and cidr_block in cidr_cache_snapshot:
        cached_data = cidr_cache_snapshot[cidr_block]
        # KeyError回避のため .get() を使用 (キーがない場合は0を返し、必ず再取得させる)
        if time.time() - cached_data.get('Timestamp', 0) < 86400:
            result.update(cached_data) 
            result['Target_IP'] = ip  # 本来のリクエストIPを再設定し、キャッシュによる上書きを防ぐ
            result['Status'] = "Success (Cache)" 
            result['Secondary_Security_Links'] = create_secondary_links(ip)
            return result, None, None

    try:
        time.sleep(delay_between_requests) 
        
        # --- API通信セクション ---
        if api_key:
            url = IPINFO_API_URL.format(ip=actual_ip) 
            headers = {"Authorization": f"Bearer {api_key}"}
            response = session.get(url, headers=headers, timeout=10)
            
            if response.status_code == 429:
                result['Status'] = 'エラー: API利用制限 (待機後に自動再試行します)'
                result['Defer_Until'] = time.time() + rate_limit_wait_seconds
                return result, None, None
                
            response.raise_for_status()
            data = response.json()
            result['IPINFO_JSON'] = data 
            
            org_raw = data.get('org', '')
            raw_isp = re.sub(r'^AS\d+\s+', '', org_raw) if org_raw else 'N/A'
            result['ISP_API_Raw'] = raw_isp
            result['CountryCode'] = data.get('country', 'N/A')
            result['Country'] = result['CountryCode']
            
            status_api = 'Success (Pro)'

        else:
            url = IP_API_URL.format(ip=actual_ip)
            response = session.get(url, timeout=45)
            
            if response.status_code == 429:
                result['Status'] = 'エラー: API利用制限 (待機後に自動再試行します)'
                result['Defer_Until'] = time.time() + rate_limit_wait_seconds
                return result, None, None
            
            response.raise_for_status()
            data = response.json()
            
            if data.get('status') == 'success':
                result['CountryCode'] = data.get('countryCode', 'N/A')
                result['Country'] = data.get('country', 'N/A')
                raw_isp_val = data.get('isp', 'N/A')
                raw_org_val = data.get('org', '')
                result['ISP_API_Raw'] = raw_isp_val if raw_org_val == raw_isp_val else f"{raw_isp_val} / {raw_org_val}"
                
                status_api = 'Success (API)'
            else:
                result['Status'] = f"エラー: IP情報取得失敗 ({data.get('message', '原因不明')})"
                return result, None, None

        # --- 匿名通信・クラウドインフラ 高精度判定 ---
        
        # 1. 公式リスト・Torリストに基づく自前判定
        cloud_provider = check_cloud_provider(actual_ip, cloud_ip_data)
        
        if actual_ip in tor_nodes:
            result['Proxy_Type'] = "🧅 TorNode"
        elif cloud_provider:
            result['Proxy_Type'] = f"☁️ Hosting ({cloud_provider})"
        else:
            result['Proxy_Type'] = ""

        # 2. VPNAPI.io による実地検証 (APIキーがある場合のみ上書き・結合)
        if vpnapi_key:
            proxy_data = get_vpnapi_data(actual_ip, vpnapi_key)
            if proxy_data:
                result['VPNAPI_JSON'] = proxy_data
                sec = proxy_data.get('security', {})
                if any(sec.values()):
                    detected = [k.upper() for k, v in sec.items() if v]
                    p_type = "/".join(detected)
                    # クラウドとVPNの両方に合致した場合は情報を結合
                    if cloud_provider:
                        result['Proxy_Type'] = f"[{p_type}] (Confirmed) on {cloud_provider}"
                    else:
                        result['Proxy_Type'] = f"[{p_type}] (Confirmed)"
                else:
                    if cloud_provider:
                        result['Proxy_Type'] = f"☁️ Hosting ({cloud_provider} / API Confirmed)"
                    else:
                        result['Proxy_Type'] = "Standard Connection (API Verified)"
        
        # --- RDAP等の補助データ取得 ---
        if use_rdap:
            rdap_res = fetch_rdap_data(actual_ip) 
            if rdap_res:
                raw_rdap_name = rdap_res['name']
                result['RDAP_Name_Raw'] = raw_rdap_name 
                result['RDAP_JSON'] = rdap_res['json']
                result['RDAP_URL'] = rdap_res['url']
                rdap_jp, _ = get_jp_names(raw_rdap_name, result['CountryCode'])
                result['RDAP_JP'] = rdap_jp

            is_composite = (actual_ip != ip and "(" in ip)

            # 複合ターゲット（ドメインから解決されたIP）の場合は、生WHOISの取得をスキップしてIP-BANを防ぐ
            if not is_composite and is_single_target:
                w_text_ip, w_server_ip = fetch_classic_whois(actual_ip)
                if w_text_ip:
                    result['IP_WHOIS_TEXT'] = w_text_ip
                    result['IP_WHOIS_SERVER'] = w_server_ip

            if is_composite:
                domain_part = ip.split("(")[0].strip()
                res_d = fetch_domain_rdap_data(domain_part)
                if res_d:
                    result['DOMAIN_RDAP_JSON'] = res_d['json']
                    result['DOMAIN_RDAP_URL'] = res_d['url']
                
                # RDAPの成否に関わらず、生のWHOISテキストは証拠として常に取得を試みる
                if is_single_target:
                    w_text, w_server = fetch_classic_whois(domain_part)
                    if w_text:
                        result['DOMAIN_WHOIS_TEXT'] = w_text
                        result['DOMAIN_WHOIS_SERVER'] = w_server

        is_composite = (actual_ip != ip and "(" in ip)
        if is_composite and st_api_key:
            st_res = get_securitytrails_data(ip.split("(")[0].strip(), st_api_key, st_start_date, st_end_date)
            if st_res: result['ST_JSON'] = st_res

        if use_rdns:
            rdns_hosts, rdns_raw = resolve_ip_nslookup(actual_ip)
            if rdns_raw: result['RDNS_DATA'] = {'hosts': rdns_hosts, 'raw': rdns_raw}

        if use_st_reverse_ip and st_api_key:
            st_rev_res = get_securitytrails_reverse_ip(actual_ip, st_api_key, use_st_rev_fetchall)
            if st_rev_res: result['ST_REVERSE_IP_JSON'] = st_rev_res

        if use_internetdb:
            result['IoT_Risk'] = check_internetdb_risk(actual_ip)
        else:
            result['IoT_Risk'] = "[Not Checked]" 

        result['Status'] = status_api
        result['RIR_Link'] = get_authoritative_rir_link(actual_ip, result['CountryCode'])
        result['Secondary_Security_Links'] = create_secondary_links(ip)

        isp_jp, country_jp = get_jp_names(result['ISP_API_Raw'], result['CountryCode'])
        result['ISP_JP'] = isp_jp
        result['Country_JP'] = country_jp
        result['ISP'] = result['ISP_JP'] if result['ISP_JP'] != 'N/A' else result['ISP_API_Raw']

        # キャッシュの鮮度判定用に現在時刻のタイムスタンプを付与
        result['Timestamp'] = time.time()

        if cidr_block:
            new_cache_entry = { cidr_block: result } 

    except requests.exceptions.ConnectionError:
        # 物理的なネットワーク切断（Wi-Fi切れ等）を検知した場合、15秒間保留キューに入れる
        result['Status'] = '待機: ネットワーク切断 (自動再試行します)'
        result['Defer_Until'] = time.time() + 15
        return result, None, None
    except requests.exceptions.Timeout:
        result['Status'] = 'エラー: 応答タイムアウト (相手サーバーの混雑または停止)'
    except requests.exceptions.HTTPError as e:
        status_code = e.response.status_code if e.response is not None else "不明"
        result['Status'] = f'エラー: 通信拒否または存在なし (HTTP {status_code})'
    except requests.exceptions.RequestException as e:
        result['Status'] = f'エラー: ネットワーク接続失敗 ({type(e).__name__})'
    except ValueError:
        result['Status'] = 'エラー: データ形式が不正 (JSON解析失敗)'
    except Exception as e:
        result['Status'] = f'エラー: 予期せぬシステム例外 ({type(e).__name__})'

    return result, new_cache_entry, new_learned_isp

def get_domain_details(domain, nslookup_raw="", st_api_key=None, st_start_date=None, st_end_date=None, is_single_target=False):
    # 捨てアド検知を実行
    detected_disposables = check_disposable_domain(domain, nslookup_raw)
    proxy_type_val = f"⚠️ 捨てアド ({' / '.join(detected_disposables)})" if detected_disposables else "N/A (Domain)"

    # TLD情報辞書から公式レジストリのリンクと日本語名を動的生成
    tld_val = domain.split('.')[-1].lower() if '.' in domain else ""
    tld_jp_name = ""
    if tld_val in TLD_INFO and TLD_INFO[tld_val]["url"] != "N/A":
        reg_name = TLD_INFO[tld_val]["jp_name"]
        tld_jp_name = f"{reg_name} (.{tld_val.upper()})"
        reg_url = TLD_INFO[tld_val]["url"]
        domain_link = f"[{reg_name} (.{tld_val.upper()} Registry)]({reg_url})"
    else:
        tld_jp_name = f"未分類 (.{tld_val.upper()})" if tld_val else "不明"
        domain_link = f"[ICANN Whois (手動検索)]({RIR_LINKS['ICANN Whois']})"
    
    # --- 1. SecurityTrails (日付フィルタ対応) ---
    st_json = None
    if st_api_key:
        st_json = get_securitytrails_data(domain, st_api_key, st_start_date, st_end_date)
    
    # --- 2. ドメインRDAPとWHOISの取得  ---
    domain_rdap_json = None
    domain_rdap_url = ''
    domain_whois_text = None
    domain_whois_server = None
    rdap_name_raw = '' # 初期値を空にする

    try:
        rdap_res = fetch_domain_rdap_data(domain)
        if rdap_res:
            domain_rdap_json = rdap_res['json']
            domain_rdap_url = rdap_res['url']
            
            # RDAPからレジストラ（登録代行業者）の特定を試みる
            entities = domain_rdap_json.get("entities", [])
            for ent in entities:
                roles = ent.get("roles", [])
                vcard_array = ent.get("vcardArray", [])
                if len(vcard_array) > 1:
                    for vcard in vcard_array[1]:
                        if "registrar" in roles and vcard[0] == "fn":
                            rdap_name_raw = vcard[3] # 確定した業者名を入れる
                            break
                if rdap_name_raw: break
        
        # 生のWHOISテキストは常に裏で取得しておく（個別レポート用）
        if is_single_target:
            domain_whois_text, domain_whois_server = fetch_classic_whois(domain)
            
    except Exception:
        pass

    # --- 3. 結果の返却 ---
    return {
        'Target_IP': domain, 
        'ISP_API_Raw': '', # ドメイン時は一覧のWhois列を空にする
        'ISP_JP': '',      # ドメイン時は一覧のWhois列を空にする
        'RDAP_Name_Raw': rdap_name_raw, # RDAPで取れた業者名（なければ空）
        'RDAP_JP': tld_jp_name,        # 唯一の確定情報である国名/TLDを表示
        'ISP': 'Domain/Host', 
        'Country': tld_val.upper() if tld_val else 'N/A',
        'Country_JP': reg_name if 'reg_name' in locals() else 'N/A',
        'CountryCode': tld_val.upper() if tld_val else 'N/A',
        'RIR_Link': domain_link,
        'Secondary_Security_Links': create_secondary_links(domain),
        'Status': 'Success (Domain)',
        
        'RDAP': '', 'RDAP_JSON': None, 'VPNAPI_JSON': None, 'RDAP_URL': '', 'IPINFO_JSON': None, 'IoT_Risk': '',
        'Proxy_Type': proxy_type_val,
        'DISPOSABLE_SERVICES': detected_disposables,
        'DOMAIN_RDAP_JSON': domain_rdap_json,
        'DOMAIN_RDAP_URL': domain_rdap_url,
        'DOMAIN_WHOIS_TEXT': domain_whois_text,
        'DOMAIN_WHOIS_SERVER': domain_whois_server,
        'ST_JSON': st_json, 
        'RDNS_DATA': None,
        'ST_REVERSE_IP_JSON': None,
        'IP_WHOIS_TEXT': None,
        'IP_WHOIS_SERVER': None
    }

def get_simple_mode_details(target):
    if is_valid_ip(target):
        rir_link_content = f"[Whois (汎用検索 - APNIC窓口)]({RIR_LINKS['APNIC']})"
    else:
        tld_val = target.split('.')[-1].lower() if '.' in target else ""
        if tld_val in TLD_INFO and TLD_INFO[tld_val]["url"] != "N/A":
            reg_name = TLD_INFO[tld_val]["jp_name"]
            reg_url = TLD_INFO[tld_val]["url"]
            rir_link_content = f"[{reg_name} (.{tld_val.upper()} Registry)]({reg_url})"
        else:
            rir_link_content = f"[ICANN Whois (手動検索)]({RIR_LINKS['ICANN Whois']})"
        
    return {
        'Target_IP': target, 
        'ISP': 'N/A (簡易モード)',
        'Country': 'N/A (簡易モード)',
        'CountryCode': 'N/A',
        'RIR_Link': rir_link_content,
        'Secondary_Security_Links': create_secondary_links(target),
        'Status': 'Success (簡易モード)',
        'RDAP': '', 'RDAP_JSON': None, 'VPNAPI_JSON': None, 'RDAP_URL': '', 'IPINFO_JSON': None, 'IoT_Risk': '',
        'DOMAIN_RDAP_JSON': None, 'DOMAIN_RDAP_URL': '', 'ST_JSON': None, 'RDNS_DATA': None,
        'DISPOSABLE_SERVICES': [],
        'DOMAIN_WHOIS_TEXT': None, 'DOMAIN_WHOIS_SERVER': None,
        'IP_WHOIS_TEXT': None, 'IP_WHOIS_SERVER': None
    }

# --- ヘルパー関数群 ---

def group_results_by_isp(results):
    grouped = {}
    final_grouped_results = []
    non_aggregated_results = []
    successful_results = [res for res in results if res['Status'].startswith('Success')]

    for res in successful_results:
        is_ip = is_valid_ip(res['Target_IP'])
        if not is_ip or not is_ipv4(res['Target_IP']) or res['ISP'] == 'N/A' or res['Country'] == 'N/A' or res['ISP'] == 'N/A (簡易モード)':
            if res['Status'].startswith('Success (IPv4 CIDR Cache)'):
                non_aggregated_results.append(res)
            else:
                non_aggregated_results.append(res)
            continue
        
        key = (res['ISP'], res['CountryCode']) 
        
        if key not in grouped:
            grouped[key] = {
                'IP_Ints': [], 'IPs_List': [], 'RIR_Link': res['RIR_Link'],
                'Secondary_Security_Links': res['Secondary_Security_Links'],
                'ISP': res['ISP'], 
                'Country': res['Country'], 
                'Status': res['Status'],
                'ISP_JP': res.get('ISP_JP', 'N/A'),
                'Country_JP': res.get('Country_JP', 'N/A')
            }
        ip_int = ip_to_int(res['Target_IP'])
        if ip_int != 0:
            grouped[key]['IP_Ints'].append(ip_int)
            grouped[key]['IPs_List'].append(res['Target_IP'])
        else:
            res['Status'] = 'Error: IPv4 Int Conversion Failed'
            non_aggregated_results.append(res)

    non_aggregated_results.extend([res for res in results if not res['Status'].startswith('Success')])
    
    for key, data in grouped.items():
        if not data['IP_Ints']: 
            continue
            
        sorted_ip_ints = sorted(data['IP_Ints'])
        min_int = sorted_ip_ints[0]
        max_int = sorted_ip_ints[-1]
        count = len(data['IPs_List'])
        try:
            min_ip = str(ipaddress.IPv4Address(min_int))
            max_ip = str(ipaddress.IPv4Address(max_int))
        except ValueError:
            min_ip = data['IPs_List'][0]
            max_ip = data['IPs_List'][-1]
        
        target_ip_display = min_ip if count == 1 else f"{min_ip} - {max_ip} (x{count} IPs)"
        status_display = data['Status'] if count == 1 else f"Aggregated ({count} IPs)"
        
        final_grouped_results.append({
            'Target_IP': target_ip_display, 
            'Country': data['Country'], 
            'Country_JP': data['Country_JP'], 
            'ISP': data['ISP'],
            'ISP_JP': data['ISP_JP'], 
            'RIR_Link': data['RIR_Link'], 
            'Secondary_Security_Links': data['Secondary_Security_Links'],
            'Status': status_display,
            'IoT_Risk': 'Aggr Mode (Skip)' # 集約時はShodan個別判定は省略
        })
    
    final_grouped_results.extend(non_aggregated_results)

    return final_grouped_results

# --- リアルタイム集計関数 ---
def summarize_in_realtime(raw_results):
    isp_counts = {}
    country_counts = {}
    country_code_counts = {}
    proxy_counts = {} 

    target_frequency = st.session_state.get('target_freq_map', {})
    st.session_state['debug_summary'] = {} 

    country_all_df_raw = pd.DataFrame({
        'NumericCode': pd.Series(dtype='int64'), 
        'Count': pd.Series(dtype='int64'),
        'Country': pd.Series(dtype='str')
    })

    success_ipv4 = [
        r for r in raw_results 
        if r['Status'].startswith('Success') and is_ipv4(r['Target_IP'])
    ]

    for r in success_ipv4:
        ip = r.get('Target_IP')
        frequency = target_frequency.get(ip, 1) 

        isp_name = r.get('ISP_JP', r.get('ISP', 'N/A'))
        country_name = r.get('Country_JP', r.get('Country', 'N/A'))
        cc = r.get('CountryCode', 'N/A')
        
        # プロキシ判定の取得
        proxy_val = r.get('Proxy_Type', '')
        if not proxy_val: 
            # API不使用時かつローカル検知(Tor/Cloud)に引っかからなかった場合の客観的な表現
            proxy_val = "非Tor / API未検証"
            
        if isp_name and isp_name not in ['N/A', 'N/A (簡易モード)']:
            isp_counts[isp_name] = isp_counts.get(isp_name, 0) + frequency
        
        if country_name and country_name != 'N/A':
            country_counts[country_name] = country_counts.get(country_name, 0) + frequency
            
        if cc and cc != 'N/A':
            country_code_counts[cc] = country_code_counts.get(cc, 0) + frequency
            
        proxy_counts[proxy_val] = proxy_counts.get(proxy_val, 0) + frequency # ⬅️ NEW: 集計

    # --- ISP集計 ---
    isp_full_df = pd.DataFrame(list(isp_counts.items()), columns=['ISP', 'Count']).sort_values('Count', ascending=False)
    isp_df = isp_full_df.head(10).copy() if not isp_full_df.empty else pd.DataFrame(columns=['ISP', 'Count'])
    if not isp_df.empty: isp_df['ISP'] = isp_df['ISP'].str.wrap(25)

    # --- 国集計 ---
    country_full_df = pd.DataFrame(list(country_counts.items()), columns=['Country', 'Count']).sort_values('Count', ascending=False)
    country_df = country_full_df.head(10).copy() if not country_full_df.empty else pd.DataFrame(columns=['Country', 'Count'])
    if not country_df.empty: country_df['Country'] = country_df['Country'].str.wrap(25)

    # --- プロキシ集計 ---
    proxy_full_df = pd.DataFrame(list(proxy_counts.items()), columns=['Proxy_Type', 'Count']).sort_values('Count', ascending=False)
    proxy_df = proxy_full_df.copy() if not proxy_full_df.empty else pd.DataFrame(columns=['Proxy_Type', 'Count'])

    # ヒートマップ用
    if country_code_counts:
        map_data = []
        for cc, cnt in country_code_counts.items():
            num = COUNTRY_CODE_TO_NUMERIC_ISO.get(cc)
            if num is not None:
                map_data.append({'NumericCode': int(num), 'Count': int(cnt), 'Country': COUNTRY_JP_NAME.get(cc, cc)})
        country_all_df_raw = pd.DataFrame(map_data).astype({'NumericCode': 'int64', 'Count': 'int64'})
        
    st.session_state['debug_summary']['country_code_counts'] = country_code_counts
    st.session_state['debug_summary']['country_all_df'] = country_all_df_raw.to_dict('records')

    # --- ターゲット頻度集計 ---
    freq_map = st.session_state.get('target_freq_map', {})
    finished = st.session_state.get('finished_ips', set())
    freq_list = [{'Target_IP': t, 'Count': c} for t, c in freq_map.items() if t in finished]
    freq_full_df = pd.DataFrame(freq_list).sort_values('Count', ascending=False) if freq_list else pd.DataFrame(columns=['Target_IP', 'Count'])
    freq_df = freq_full_df.head(10).copy() if not freq_full_df.empty else pd.DataFrame(columns=['Target_IP', 'Count'])

    # 戻り値に proxy_df を追加
    return isp_df, country_df, freq_df, country_all_df_raw, isp_full_df, country_full_df, freq_full_df, proxy_df

# --- 集計結果描画ヘルパー関数 (2x2ダッシュボード) ---
def draw_summary_content(isp_summary_df, country_summary_df, target_frequency_df, country_all_df, proxy_df, title):
    st.markdown(f"**{title}**")
    
    # 2x2のグリッドを作成
    row1_col1, row1_col2 = st.columns(2)
    row2_col1, row2_col2 = st.columns(2)
    
    # 【左上】 世界マップ (コンパクト版・南極カットアウト)
    with row1_col1:
        st.markdown("📍 **国別 ヒートマップ**")
        if WORLD_MAP_GEOJSON and not country_all_df.empty:
            base = alt.Chart(WORLD_MAP_GEOJSON).mark_geoshape(
                stroke='black', strokeWidth=0.1, fill="#f0f0f052"
            ).project(
                type='mercator',
                scale=65,             # 大陸を拡大表示
                translate=[220, 150]  # 南極をフレームアウトさせつつ中央に配置
            ) 
            
            heatmap = alt.Chart(WORLD_MAP_GEOJSON).mark_geoshape(
                stroke='black', strokeWidth=0.1
            ).encode(
                color=alt.Color('Count:Q', scale=alt.Scale(type='log', scheme='yelloworangered'), legend=None),
                tooltip=[alt.Tooltip('Country:N', title='国名'), alt.Tooltip('Count:Q', title='件数', format=',')]
            ).transform_lookup(
                lookup='id', from_=alt.LookupData(country_all_df, key='NumericCode', fields=['Count', 'Country'])
            ).project(
                type='mercator',
                scale=65,             
                translate=[220, 150]  
            )
            
            # 高さを250pxに固定してコンパクトに
            chart = alt.layer(base, heatmap).resolve_scale(color='independent').properties(height=250)
            st.altair_chart(chart, width="stretch")
        else:
            st.info("データなし")

    # 【右上】 ISP横棒グラフ
    with row1_col2:
        st.markdown("🏢 **ISP別 件数 (Top 10)**")
        if not isp_summary_df.empty:
            chart = alt.Chart(isp_summary_df).mark_bar().encode(
                x=alt.X('Count:Q', title='件数'),
                y=alt.Y('ISP:N', sort='-x', title=''),
                tooltip=['ISP', 'Count']
            ).properties(height=250)
            st.altair_chart(chart, width="stretch")
        else:
            st.info("データなし")

    # 【左下】 国別 円グラフ
    with row2_col1:
        st.markdown("🌍 **国別 割合 (Pie Chart)**")
        if not country_summary_df.empty:
            chart = alt.Chart(country_summary_df).mark_arc().encode(
                theta=alt.Theta(field="Count", type="quantitative"),
                color=alt.Color(field="Country", type="nominal", legend=alt.Legend(title="国名", orient="right")),
                tooltip=["Country", "Count"]
            ).properties(height=250)
            st.altair_chart(chart, width="stretch")
        else:
            st.info("データなし")

    # 【右下】 プロキシ・リスク ドーナツチャート
    with row2_col2:
        st.markdown("🕵️ **プロキシ・VPN 割合 (Donut)**")
        if not proxy_df.empty:
            chart = alt.Chart(proxy_df).mark_arc(innerRadius=50).encode(
                theta=alt.Theta(field="Count", type="quantitative"),
                color=alt.Color(field="Proxy_Type", type="nominal", legend=alt.Legend(title="判定", orient="right")),
                tooltip=["Proxy_Type", "Count"]
            ).properties(height=250)
            st.altair_chart(chart, width="stretch")
        else:
            st.info("データなし")

# 💡 HTMLレポート生成関数
def generate_full_report_html(isp_full_df, country_full_df, freq_full_df):
    
    def create_chunked_chart_specs(df, x_col, y_col, title_base, chunk_size=50):
        specs = []
        # データ全体での最大値を取得 (ページまたぎのスケール統一のため)
        global_max = df[x_col].max() if not df.empty else 0

        # データフレームを分割
        chunks = [df[i:i + chunk_size] for i in range(0, df.shape[0], chunk_size)]
        
        for i, chunk in enumerate(chunks):
            chart_title = f"{title_base} ({i+1}/{len(chunks)})" if len(chunks) > 1 else title_base
            
            # 数値ラベル付きチャート
            # 💡 x軸のスケールを全体最大値で固定する
            base = alt.Chart(chunk).encode(
                x=alt.X(x_col, title='Count', scale=alt.Scale(domain=[0, global_max])),
                y=alt.Y(y_col, sort='-x', title=y_col),
                tooltip=[y_col, x_col]
            )
            bars = base.mark_bar()
            text = base.mark_text(
                align='left',
                baseline='middle',
                dx=5, 
                fontSize=11,
                fontWeight='bold'
            ).encode(
                text=x_col
            )
            chart = (bars + text).properties(
                title=chart_title,
                width=700,
                height=alt.Step(20) 
            )
            specs.append(chart.to_dict())
        return specs

    # 各カテゴリのチャートスペックを生成
    target_specs = create_chunked_chart_specs(freq_full_df, 'Count', 'Target_IP', 'Target IP Counts (All)')
    isp_specs = create_chunked_chart_specs(isp_full_df, 'Count', 'ISP', 'ISP Counts (All)')
    country_specs = create_chunked_chart_specs(country_full_df, 'Count', 'Country', 'Country Counts (All)')

    # HTMLテンプレート
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <title>Whois Search Full Report</title>
      <script src="https://cdn.jsdelivr.net/npm/vega@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-lite@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-embed@6"></script>
      <style>
        body {{ font-family: "Helvetica Neue", Arial, sans-serif; padding: 40px; background-color: #fff; color: #333; }}
        h1 {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 30px; }}
        h2 {{ 
            color: #1e3a8a; 
            margin-top: 50px; 
            border-left: 5px solid #1e3a8a; 
            padding-left: 15px; 
            page-break-before: always; 
        }}
        h2:first-of-type {{ page-break-before: auto; }} 
        
        .chart-container {{ 
            margin-bottom: 40px; 
            padding: 10px; 
            page-break-inside: avoid; 
        }}
        
        @media print {{
            body {{ padding: 0; background-color: #fff; }}
            .no-print {{ display: none; }}
            h2 {{ margin-top: 20px; }}
        }}
      </style>
    </head>
    <body>
      <h1>Whois検索結果分析レポート</h1>
      <p style="text-align: center; color: #666;">Generated by Whois Search Tool</p>

      <h2>対象IPアドレス カウント (全 {len(freq_full_df)} 件)</h2>
      <div id="target_charts"></div>

      <h2>ISP別 カウント (全 {len(isp_full_df)} 件)</h2>
      <div id="isp_charts"></div>

      <h2>国別 カウント (全 {len(country_full_df)} 件)</h2>
      <div id="country_charts"></div>

      <script type="text/javascript">
        // Embed charts function
        function embedCharts(containerId, specs) {{
            const container = document.getElementById(containerId);
            specs.forEach((spec, index) => {{
                const div = document.createElement('div');
                div.id = containerId + '_' + index;
                div.className = 'chart-container';
                container.appendChild(div);
                vegaEmbed('#' + div.id, spec, {{actions: false}});
            }});
        }}

        // Data from Python (Serialized to JSON)
        const targetSpecs = {json.dumps(target_specs)};
        const ispSpecs = {json.dumps(isp_specs)};
        const countrySpecs = {json.dumps(country_specs)};

        // Render
        if (targetSpecs.length > 0) embedCharts('target_charts', targetSpecs);
        else document.getElementById('target_charts').innerHTML = '<p>データなし</p>';

        if (ispSpecs.length > 0) embedCharts('isp_charts', ispSpecs);
        else document.getElementById('isp_charts').innerHTML = '<p>データなし</p>';

        if (countrySpecs.length > 0) embedCharts('country_charts', countrySpecs);
        else document.getElementById('country_charts').innerHTML = '<p>データなし</p>';
      </script>
    </body>
    </html>
    """
    return html_template

# --- 🛡️ 脅威インテリジェンス (STIX 2.1) 生成関数 ---
def generate_stix2_bundle(results):
    """ 調査結果をSTIX 2.1形式のBundle (JSON) に変換する """
    objects = []
    # STIXの時刻は UTC の ISO8601 形式が必須
    now_str = datetime.datetime.utcnow().strftime('%Y-%m-%dT%H:%M:%S.000Z')

    # 1. ツール自身のIdentity（作成者）オブジェクト
    identity_id = f"identity--{uuid.uuid4()}"
    objects.append({
        "type": "identity",
        "spec_version": "2.1",
        "id": identity_id,
        "created": now_str,
        "modified": now_str,
        "name": "検索大臣 - IP/Domain OSINT Tool",
        "identity_class": "system"
    })

    # 2. 各ターゲットをIndicator（脅威インジケーター）オブジェクトとして変換
    for res in results:
        target = res.get("Target_IP", "")
        if not target or target == "N/A": continue
        
        # 集約モード時の表記 (1.1.1.1 - 1.1.1.5) やドメイン複合 (example.com (1.1.1.1)) をクリーンアップ
        clean_target = str(target).split(' - ')[0].split(' ')[0]
        if "(" in clean_target:
            clean_target = clean_target.split("(")[0].strip()

        # STIXのサイバー観測パターンの構築
        if is_ipv4(clean_target):
            pattern = f"[ipv4-addr:value = '{clean_target}']"
        elif is_valid_ip(clean_target): # IPv6
            pattern = f"[ipv6-addr:value = '{clean_target}']"
        else:
            pattern = f"[domain-name:value = '{clean_target}']"

        isp = res.get("ISP_JP", "N/A")
        country = res.get("Country_JP", "N/A")
        proxy = res.get("Proxy_Type", "")
        risk = res.get("IoT_Risk", "")

        # 脅威の説明文を構築
        desc = f"【OSINT調査結果】\nISP: {isp}\n国: {country}\n匿名化/クラウド: {proxy if proxy else 'なし'}\nIoTリスク: {risk}"
        
        # SOC向けのアラート用ラベル
        labels = ["osint-target"]
        if proxy and proxy != "未検証" and "Standard" not in proxy: 
            labels.append("anonymization-proxy-vpn")
        if risk and risk not in ["[No Match]", "[Not Checked]", ""]: 
            labels.append("vulnerable-iot")

        indicator_id = f"indicator--{uuid.uuid4()}"
        objects.append({
            "type": "indicator",
            "spec_version": "2.1",
            "id": indicator_id,
            "created_by_ref": identity_id,
            "created": now_str,
            "modified": now_str,
            "name": f"OSINT Target: {clean_target}",
            "description": desc,
            "indicator_types": ["malicious-activity" if len(labels) > 1 else "anomalous-activity"],
            "pattern": pattern,
            "pattern_type": "stix",
            "valid_from": now_str,
            "labels": labels
        })

    # 3. すべてを1つのBundleにパッキング
    bundle = {
        "type": "bundle",
        "id": f"bundle--{uuid.uuid4()}",
        "objects": objects
    }
    
    return json.dumps(bundle, indent=4, ensure_ascii=False)

# 📈 クロス分析用HTMLレポート生成関数
def generate_cross_analysis_html(chart_spec, x_col, group_col):
    html_template = f"""
    <!DOCTYPE html>
    <html>
    <head>
      <title>Whois Cross Analysis Report</title>
      <script src="https://cdn.jsdelivr.net/npm/vega@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-lite@5"></script>
      <script src="https://cdn.jsdelivr.net/npm/vega-embed@6"></script>
      <style>
        body {{ font-family: "Helvetica Neue", Arial, sans-serif; padding: 40px; background-color: #fff; color: #333; }}
        h1 {{ text-align: center; border-bottom: 2px solid #333; padding-bottom: 10px; margin-bottom: 30px; }}
        .info {{ text-align: center; color: #666; margin-bottom: 20px; }}
        .chart-container {{ 
            width: 100%; 
            display: flex; 
            justify-content: center; 
            margin-bottom: 40px; 
            padding: 10px; 
        }}
      </style>
    </head>
    <body>
      <h1>クロス分析レポート: {x_col} vs {group_col}</h1>
      <p class="info">Generated by Whois Search Tool</p>

      <div id="chart" class="chart-container"></div>

      <script type="text/javascript">
        const spec = {json.dumps(chart_spec)};
        vegaEmbed('#chart', spec, {{actions: true}});
      </script>
    </body>
    </html>
    """
    return html_template

# --- Excel生成ヘルパー関数 ---
def convert_df_to_excel(df):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_excel_path = tmp.name
    try:
        with pd.ExcelWriter(tmp_excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sheet1')
        with open(tmp_excel_path, "rb") as f:
            data = f.read()
        return data
    finally:
        if os.path.exists(tmp_excel_path):
            os.remove(tmp_excel_path)

# --- Advanced Excel Generator (Pivot & Chart) ---
def create_advanced_excel(df, time_col_name=None):
    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        tmp_excel_path = tmp.name
    
    try:
        # 1. IPアドレスが含まれているかを判定
        target_col = 'IPアドレス' if 'IPアドレス' in df.columns else ('Target_IP' if 'Target_IP' in df.columns else df.columns[0])
        has_ip = False
        if target_col in df.columns:
            has_ip = any(is_valid_ip(str(val)) for val in df[target_col].dropna())
        
        # ==========================================
        # パターンA: すべてドメイン(非IP)の場合
        # ==========================================
        if not has_ip:
            if target_col in df.columns:
                df = df.rename(columns={target_col: 'Target Domain'})
            # 不要な列を削除（日本語名に対応）
            cols_to_drop = ['Whois結果（元データ）', 'Whois結果（日本語名称）', '国名（英語）', '国名', 'プロキシ種別', 'ステータス', 'IoTリスク', 'RDAP結果（元データ）', 'RDAP結果（日本語名称）', 'ISP', 'ISP_JP', 'Country', 'Country_JP']
            df = df.drop(columns=[c for c in cols_to_drop if c in df.columns], errors='ignore')
            with pd.ExcelWriter(tmp_excel_path, engine='openpyxl') as writer:
                df.to_excel(writer, index=False, sheet_name='Domain Results')
            with open(tmp_excel_path, "rb") as f:
                data = f.read()
            return data

        # ==========================================
        # パターンB: IPアドレスが含まれる場合 (分析グラフ付き)
        # ==========================================
        # 必須カラムの補完 (日本語名を基準にする)
        required_cols = {
            'プロキシ種別': '',
            'Whois結果（日本語名称）': 'N/A',
            '国名': 'N/A'
        }
        for col, default_val in required_cols.items():
            if col not in df.columns:
                df[col] = default_val

        # データ前処理：空欄や欠損値を空文字（何もなし）に統一
        df['プロキシ種別'] = df['プロキシ種別'].fillna('')
        
        has_time_analysis = False
        if time_col_name and time_col_name in df.columns:
            try:
                df['Hour'] = pd.to_datetime(df[time_col_name], errors='coerce').dt.hour
                has_time_analysis = True
            except Exception:
                pass

        count_col = df.columns[0]

        # 書き込み開始
        with pd.ExcelWriter(tmp_excel_path, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Raw Data')
            wb = writer.book
            
            def add_chart_sheet(pivot_df, sheet_name, chart_title, x_title, y_title, description, chart_type="col", stacked=False):
                if pivot_df.empty: return
                pivot_df.to_excel(writer, sheet_name=sheet_name, startrow=4)
                ws = wb[sheet_name]
                ws['A1'] = chart_title
                ws['A1'].font = Font(size=14, bold=True, color="1E3A8A")
                ws['A2'] = description
                ws['A2'].font = Font(size=11, color="555555", italic=True)
                ws['A2'].alignment = Alignment(wrap_text=True, vertical="top")
                ws.merge_cells('A2:H3')
                
                chart = BarChart()
                chart.type = chart_type
                chart.style = 10 
                chart.title = chart_title
                chart.height = 15 
                chart.width = 25  
                # 凡例を非表示にする
                chart.legend = None
                
                if stacked:
                    chart.grouping = "stacked"
                    chart.overlap = 100
                else:
                    chart.varyColors = True
                    
                # 環境依存による表示エラーを回避するため、全環境で安定するデータラベル表示に統一
                from openpyxl.chart.label import DataLabelList
                chart.dataLabels = DataLabelList()
                chart.dataLabels.showVal = True
                    
                chart.x_axis.title = x_title
                chart.y_axis.title = y_title
                chart.y_axis.majorGridlines = ChartLines() 
                chart.y_axis.delete = False        
                chart.y_axis.numFmt = '0'          
                chart.y_axis.majorTickMark = 'out' 
                chart.y_axis.tickLblPos = 'nextTo' 
                chart.layout = Layout(manualLayout=ManualLayout(x=0.03, y=0.05, h=0.75, w=0.85))
                
                data_start_row = 5 
                data_end_row = data_start_row + len(pivot_df)
                data = Reference(ws, min_col=2, min_row=data_start_row, max_row=data_end_row, max_col=len(pivot_df.columns)+1)
                cats = Reference(ws, min_col=1, min_row=data_start_row+1, max_row=data_end_row)
                chart.add_data(data, titles_from_data=True)
                chart.set_categories(cats)
                ws.add_chart(chart, "E5")

            # プロキシデータが有効か（全て空文字ではないか）を判定するフラグ
            proxy_col = 'プロキシ種別'
            has_valid_risk_data = False
            if proxy_col in df.columns:
                if not (df[proxy_col] == '').all():
                    has_valid_risk_data = True

            # ---------------------------------------------------------
            # 1. Report_Whois_Volume: ISP別アクセス数
            # ---------------------------------------------------------
            isp_col = 'Whois結果（日本語名称）'
            if isp_col in df.columns and not (df[isp_col] == 'N/A').all():
                top_isps = df[isp_col].value_counts().head(20).index
                df_isp = df[df[isp_col].isin(top_isps)]
                pivot_isp_vol = df_isp.pivot_table(index=isp_col, values=count_col, aggfunc='count')
                
                if not pivot_isp_vol.empty:
                    pivot_isp_vol = pivot_isp_vol.sort_values(count_col, ascending=False)
                    desc_isp_vol = "どの組織・プロバイダからのアクセスが最も多いかを可視化しています。特定のISPからのアクセス集中は、そのサービスの利用者層または特定のキャンペーンの影響を示唆します。"
                    add_chart_sheet(pivot_isp_vol, 'Report_Whois_Volume', 'Whois Access Volume Ranking (Top 20)', 'ISP Name', 'Count', desc_isp_vol)

                # ---------------------------------------------------------
                # 2. Report_Whois_Risk: ISP別リスク分析
                # ---------------------------------------------------------
                if has_valid_risk_data:
                    pivot_isp_risk = df_isp.pivot_table(index=isp_col, columns=proxy_col, values=count_col, aggfunc='count', fill_value=0)
                    if not pivot_isp_risk.empty:
                        desc_isp_risk = "そのISPが安全な一般回線か、注意が必要なサーバー/VPN経由かを判定しています。「未検証」はAPIによる検証なし、「Standard Connection (API Verified)」はAPI検証済みの一般回線を示します。"
                        add_chart_sheet(pivot_isp_risk, 'Report_Whois_Risk', 'Risk Analysis by Whois (Top 20)', 'ISP Name', 'Count', desc_isp_risk, stacked=True)

            # ---------------------------------------------------------
            # 3. Report_RDAP_Volume: RDAP別アクセス数
            # ---------------------------------------------------------
            rdap_col = 'RDAP結果（日本語名称）'
            if rdap_col in df.columns:
                # 空文字もN/Aとして扱うための前処理
                df[rdap_col] = df[rdap_col].replace('', 'N/A')
                
                # RDAPデータが有効な場合のみ実行
                if not (df[rdap_col] == 'N/A').all():
                    top_rdaps = df[rdap_col].value_counts().head(20).index
                    df_rdap = df[df[rdap_col].isin(top_rdaps)]
                    pivot_rdap_vol = df_rdap.pivot_table(index=rdap_col, values=count_col, aggfunc='count')
                    
                    if not pivot_rdap_vol.empty:
                        pivot_rdap_vol = pivot_rdap_vol.sort_values(count_col, ascending=False)
                        desc_rdap_vol = "公式レジストリ（RDAP）に登録されている法的な保有組織ごとのアクセス数です。Whois（運用者）とは異なる、IPアドレスブロックの真の所有者傾向を可視化します。"
                        add_chart_sheet(pivot_rdap_vol, 'Report_RDAP_Volume', 'RDAP Access Volume Ranking (Top 20)', 'RDAP Name', 'Count', desc_rdap_vol)

                    # ---------------------------------------------------------
                    # 4. Report_RDAP_Risk: RDAP別リスク分析
                    # ---------------------------------------------------------
                    if has_valid_risk_data:
                        pivot_rdap_risk = df_rdap.pivot_table(index=rdap_col, columns=proxy_col, values=count_col, aggfunc='count', fill_value=0)
                        if not pivot_rdap_risk.empty:
                            desc_rdap_risk = "法的保有組織（RDAP）ごとの接続環境を分析しています。特定の組織が保有するIP帯域が、プロキシやVPNインフラとして集中的に悪用されていないかを確認できます。"
                            add_chart_sheet(pivot_rdap_risk, 'Report_RDAP_Risk', 'Risk Analysis by RDAP (Top 20)', 'RDAP Name', 'Count', desc_rdap_risk, stacked=True)
            
            # ---------------------------------------------------------
            # 5. Report_Country: 国別アクセス数
            # ---------------------------------------------------------
            if '国名' in df.columns and not (df['国名'] == 'N/A').all():
                pivot_country = df.pivot_table(index='国名', values=count_col, aggfunc='count')
                if not pivot_country.empty:
                    pivot_country = pivot_country.sort_values(count_col, ascending=False).head(15)
                    desc_country = "国ごとのアクセス数をランキング化しています。サービス提供エリア外からの予期せぬアクセス検知や、海外からの攻撃予兆の発見に役立ちます。"
                    add_chart_sheet(pivot_country, 'Report_Country', 'Country Access Volume (Top 15)', 'Country Name', 'Count', desc_country)

            # ---------------------------------------------------------
            # 6. Report_Time: 時間帯分析
            # ---------------------------------------------------------
            if has_time_analysis:
                pivot_time_vol = df.pivot_table(index='Hour', values=count_col, aggfunc='count', fill_value=0).reindex(range(24), fill_value=0)
                desc_time_vol = "何時にアクセスが集中しているかを可視化しています。一般的なユーザーは活動時間帯に、Botなどは深夜早朝や24時間一定のアクセスを行う傾向があります。"
                add_chart_sheet(pivot_time_vol, 'Report_Time_Volume', 'Hourly Access Trend', 'Hour (0-23h)', 'Count', desc_time_vol)

                if has_valid_risk_data:
                    pivot_time_risk = df.pivot_table(index='Hour', columns=proxy_col, values=count_col, aggfunc='count', fill_value=0).reindex(range(24), fill_value=0)
                    desc_time_risk = "深夜帯などに怪しいアクセス（Hosting/VPN等）が増えていないかを確認できます。夜間にHosting判定が増加する場合、Botによる自動巡回の可能性があります。"
                    add_chart_sheet(pivot_time_risk, 'Report_Time_Risk', 'Hourly Risk Trend', 'Hour (0-23h)', 'Count', desc_time_risk, stacked=True)
                
        with open(tmp_excel_path, "rb") as f:
            data = f.read()
        return data
    finally:
        if os.path.exists(tmp_excel_path):
            os.remove(tmp_excel_path)

def generate_individual_html_report(res, clean_ip, report_opts=None):
    """ 
    個別IPの詳細HTMLレポートを生成する独立関数
    """
    import json
    import html
    import re
    import datetime
    from urllib.parse import urlparse
    
    # オプションが指定されていない場合はすべてTrue（全出力）とする
    if report_opts is None:
        report_opts = {"tld": True, "dns": True, "subnet": True, "rdap": True, "whois": True, "ipinfo": True, "vpnapi": True, "st": True, "rdns": True}
    import html
    import re
    import datetime
    from urllib.parse import urlparse
    
    # --- 1. データの展開 ---
    target_ip = res.get('Target_IP', 'N/A')
    rdap_url = res.get('RDAP_URL')
    rdap_json = res.get('RDAP_JSON')
    ipinfo_json = res.get('IPINFO_JSON')
    vpnapi_json = res.get('VPNAPI_JSON')
    domain_rdap_json = res.get('DOMAIN_RDAP_JSON')
    domain_rdap_url = res.get('DOMAIN_RDAP_URL')
    domain_whois_text = res.get('DOMAIN_WHOIS_TEXT')
    domain_whois_server = res.get('DOMAIN_WHOIS_SERVER')
    ip_whois_text = res.get('IP_WHOIS_TEXT') 
    ip_whois_server = res.get('IP_WHOIS_SERVER')  
    st_json = res.get('ST_JSON')
    
    rdns_data = res.get('RDNS_DATA', {})
    rdns_raw = rdns_data.get('raw', '') if isinstance(rdns_data, dict) else ""
    rdns_hosts = rdns_data.get('hosts', []) if isinstance(rdns_data, dict) else []
    
    nslookup_data = {}
    domain_name_for_nslookup = ""
    
    if "(" in target_ip and ")" in target_ip:
        domain_name_for_nslookup = target_ip.split("(")[0].strip()
        nslookup_data = st.session_state.get('resolved_dns_map', {}).get(domain_name_for_nslookup, {})
    elif not is_valid_ip(target_ip): 
        domain_name_for_nslookup = target_ip
        nslookup_data = st.session_state.get('resolved_dns_map', {}).get(domain_name_for_nslookup, {})
    
    nslookup_raw = nslookup_data.get('raw', '') if isinstance(nslookup_data, dict) else ""
    nslookup_ips = nslookup_data.get('ips', []) if isinstance(nslookup_data, dict) else []
    
    if isinstance(nslookup_data, str):
        nslookup_raw = nslookup_data
    
    if not ((rdap_url and rdap_json) or ipinfo_json or domain_rdap_json or domain_whois_text or nslookup_raw or st_json or rdns_raw or ip_whois_text):
        return None # レポート生成に必要なデータがない場合は None を返す

    jst_timezone = datetime.timezone(datetime.timedelta(hours=9))
    now_jst = datetime.datetime.now(jst_timezone)
    current_time_str = now_jst.strftime("%Y年%m月%d日 %H時%M分")

    tabs_html = ""
    contents_html = ""
    first_tab_id = None
    
    # --- 2. サブネット・ネットワーク情報 (IPv4 / IPv6 両対応) ---
    try:
        ip_obj = ipaddress.ip_address(clean_ip)
        if report_opts.get("subnet", True):
            tab_id = "tab-subnet"
            if not first_tab_id: first_tab_id = tab_id
            tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">サブネット情報</button>\n'
            
            # ==========================================
            # IPv4 の場合の処理
            # ==========================================
            if ip_obj.version == 4:
                target_net = ipaddress.IPv4Network(f"{clean_ip}/24", strict=False)
                calc_source = "デフォルト推測 (/24 基準)"
                
                # RDAPデータに基づく正確なCIDR算出
                if rdap_json and "startAddress" in rdap_json and "endAddress" in rdap_json:
                    try:
                        s_ip = ipaddress.IPv4Address(rdap_json["startAddress"])
                        e_ip = ipaddress.IPv4Address(rdap_json["endAddress"])
                        cidrs = list(ipaddress.summarize_address_range(s_ip, e_ip))
                        for cidr in cidrs:
                            if ip_obj in cidr:
                                target_net = cidr
                                calc_source = "RDAP公式割当範囲"
                                break
                    except Exception:
                        pass
                
                first_octet = int(str(ip_obj).split('.')[0])
                if 1 <= first_octet <= 126: ip_class = "クラスA"
                elif 128 <= first_octet <= 191: ip_class = "クラスB"
                elif 192 <= first_octet <= 223: ip_class = "クラスC"
                elif 224 <= first_octet <= 239: ip_class = "クラスD マルチキャスト"
                elif 240 <= first_octet <= 255: ip_class = "クラスE 実験用"
                else: ip_class = "不明"

                host_min = str(target_net[1]) if target_net.num_addresses > 2 else "なし"
                host_max = str(target_net[-2]) if target_net.num_addresses > 2 else "なし"
                num_hosts = max(0, target_net.num_addresses - 2)

                subnet_content = f"""
                <div id="{tab_id}" class="tab-content">
                    <h1 class="theme-ipinfo" style="color: #00695c; border-color: #00695c;">サブネット・ネットワーク範囲計算 (IPv4)</h1>
                    <div class="description" style="background-color: #e0f2f1; border-color: #b2dfdb;">
                        <strong>論理ネットワーク範囲：</strong><br>
                        入力されたIPアドレスが属するネットワーク境界を算出する。RDAPから公式のIPアドレス割当範囲が取得できた場合はその範囲に基づく正確なCIDRを適用し、情報がない場合は一般的なCクラス相当（/24）を基準として計算する。
                    </div>
                    <h2>IPアドレスクラス及び基本情報</h2>
                    <table>
                        <tr><th>対象IPアドレス</th><td><strong>{clean_ip}</strong></td></tr>
                        <tr><th>IPアドレスクラス</th><td><strong>{ip_class}</strong></td></tr>
                        <tr><th>算出基準</th><td><strong>{calc_source}</strong></td></tr>
                    </table>
                    <h2>ネットワーク範囲の計算結果</h2>
                    <table>
                        <tr><th>サブネットマスク</th><td><strong>/{target_net.prefixlen} ({target_net.netmask})</strong></td></tr>
                        <tr><th>ネットワークアドレス<br>(開始IP)</th><td><strong>{target_net.network_address}</strong></td></tr>
                        <tr><th>ホストアドレス範囲<br>(使用可能IP)</th><td><strong>{host_min} ～ {host_max}</strong></td></tr>
                        <tr><th>ブロードキャストアドレス<br>(終了IP)</th><td><strong>{target_net.broadcast_address}</strong></td></tr>
                        <tr><th>アドレス数</th><td><strong>IPアドレス総数: {target_net.num_addresses:,} (ホストアドレス数: {num_hosts:,})</strong></td></tr>
                    </table>
                </div>
                """
                contents_html += subnet_content

            # ==========================================
            # IPv6 の場合の処理 (新規追加)
            # ==========================================
            elif ip_obj.version == 6:
                # 一般的なLAN・VLANの境界である /64 をデフォルトとする
                target_net = ipaddress.IPv6Network(f"{clean_ip}/64", strict=False)
                calc_source = "デフォルト推測 (/64 標準サブネット基準)"
                
                # RDAPデータに基づく正確なCIDR算出
                if rdap_json and "startAddress" in rdap_json and "endAddress" in rdap_json:
                    try:
                        s_ip = ipaddress.IPv6Address(rdap_json["startAddress"])
                        e_ip = ipaddress.IPv6Address(rdap_json["endAddress"])
                        cidrs = list(ipaddress.summarize_address_range(s_ip, e_ip))
                        for cidr in cidrs:
                            if ip_obj in cidr:
                                target_net = cidr
                                calc_source = "RDAP公式割当範囲"
                                break
                    except Exception:
                        pass
                
                # IPv6のプレフィックス長に基づく割り当て規模の推定
                prefix = target_net.prefixlen
                if prefix <= 32:
                    scope_desc = "LIR / 大規模ISP割当 (非常に広大なインフラ空間)"
                elif prefix <= 48:
                    scope_desc = "企業・大規模拠点割当 (一般的なエンタープライズ境界)"
                elif prefix <= 56:
                    scope_desc = "一般家庭・小規模拠点割当 (コンシューマ向けルーター等)"
                elif prefix == 64:
                    scope_desc = "単一セグメント (標準的な1つのLAN・VLAN)"
                else:
                    scope_desc = "デバイス・ホスト固有割当"

                # IPv6はアドレス数が膨大なため、指数表記を併用して視認性を高める
                total_ips = target_net.num_addresses
                if total_ips > 10**6:
                    ips_display = f"2^{128 - prefix} 個 (約 {total_ips:.2e})"
                else:
                    ips_display = f"{total_ips:,} 個"

                # IPv6ではブロードキャストアドレスの概念がないため、すべてのアドレスを表記
                network_addr = str(target_net.network_address)
                # target_net.broadcast_address は ipaddressモジュールの仕様上、ネットワークの最終IPを返す
                last_addr = str(target_net.broadcast_address) 

                subnet_content = f"""
                <div id="{tab_id}" class="tab-content">
                    <h1 class="theme-ipinfo" style="color: #4a148c; border-color: #4a148c;">サブネット・ネットワーク範囲計算 (IPv6)</h1>
                    <div class="description" style="background-color: #f3e5f5; border-color: #e1bee7;">
                        <strong>IPv6 論理ネットワーク範囲：</strong><br>
                        IPv6アドレスの階層構造に基づき、属するネットワーク境界を算出する。RDAPから公式のIPアドレス割当範囲が取得できた場合はその範囲に基づく正確なCIDRを適用し、情報がない場合は標準的な単一セグメント（/64）を基準として計算する。
                    </div>
                    <h2>IPv6 アドレス属性</h2>
                    <table>
                        <tr><th>対象IPアドレス</th><td><strong>{clean_ip}</strong></td></tr>
                        <tr><th>算出基準</th><td><strong>{calc_source}</strong></td></tr>
                        <tr><th>ネットワーク規模推定</th><td><strong>{scope_desc}</strong></td></tr>
                    </table>
                    <h2>ネットワーク範囲の計算結果</h2>
                    <table>
                        <tr><th>プレフィックス長</th><td><strong>/{prefix}</strong></td></tr>
                        <tr><th>ネットワークアドレス<br>(開始IP)</th><td><strong>{network_addr}</strong></td></tr>
                        <tr><th>アドレス範囲<br>(割当可能範囲)</th><td><strong>{network_addr} ～<br>{last_addr}</strong></td></tr>
                        <tr><th>アドレス総数</th><td><strong>{ips_display}</strong></td></tr>
                    </table>
                </div>
                """
                contents_html += subnet_content

    except ValueError:
        pass

    # --- 3. nslookup (DNS正引き) ---
    if nslookup_raw and report_opts.get("dns", True):
        tab_id = "tab-nslookup"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">DNS正引き</button>\n'
        
        escaped_nslookup = html.escape(nslookup_raw)
        for ip_str in nslookup_ips:
            escaped_ip = html.escape(ip_str)
            escaped_nslookup = escaped_nslookup.replace(escaped_ip, f'<span class="json-hl">{escaped_ip}</span>')
            
        cmd_str = f"resolver = dns.resolver.Resolver(); resolver.nameservers=['8.8.8.8']; resolver.resolve('{domain_name_for_nslookup}', 'A/AAAA/MX')"
        ip_list_str = "<br>".join([html.escape(ip) for ip in nslookup_ips]) if nslookup_ips else "取得なし"
        
        # --- 捨てアド (Disposable Email) 検知ロジック ---
        detected_services = res.get('DISPOSABLE_SERVICES', [])
        
        table_alert_row = ""
        mx_alert_html = ""
        if detected_services:
            services_str = html.escape("、".join(detected_services))
            mx_alert_html = f"""
            <div style="background-color: #ffebee; border-left: 5px solid #f44336; padding: 15px; margin-bottom: 20px; border-radius: 4px;">
                <strong style="color: #c62828; font-size: 1.1em;">使い捨てメールサービス (Disposable Email) を検知</strong><br>
                <span style="color: #b71c1c; font-size: 0.9em;">対象のドメイン、またはそのMX（メール交換）レコードが、使い捨てメールサービスのインフラを示している。</span>
            </div>
            """
            table_alert_row = f"""
                <tr><th>使い捨てメール判定<br>(Disposable Service)</th><td><strong style="color: #c62828;">該当あり ({services_str})</strong></td></tr>
            """

        nslookup_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-rdap" style="color: #424242; border-color: #424242;">DNS正引き・MX解決結果 (dnspython)</h1>
            {mx_alert_html}
            <div class="description" style="background-color: #eceff1; border-color: #cfd8dc;">
                <strong>DNS (Domain Name System) 正引き解決記録：</strong><br>
                入力されたドメイン名に対し、IPアドレス（A/AAAA/MXレコード）の特定を行った結果を示す。<br>
                一般的な <code>nslookup</code> コマンドは実行環境のDNS設定に依存するが、本ツールではPythonの専門ライブラリを使用して、信頼性の高いパブリックDNS（Google/Cloudflare）へ<strong>直接的に問い合わせ</strong>を行っている。<br>
            </div>
            <h2>対象ドメイン及び取得結果</h2>
            <table>
                <tr><th>対象ドメイン<br>(Target Domain)</th><td><strong>{html.escape(domain_name_for_nslookup)}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
                <tr><th>取得IPアドレス<br>(Resolved IPs)</th><td><strong>{ip_list_str}</strong></td></tr>
                {table_alert_row}
            </table>
            <h2>内部実行クエリ (Python)</h2>
            <div class="raw-data" style="background-color: #263238; color: #eceff1; font-weight: bold; font-family: Consolas, monospace;">>>> {cmd_str}</div>
            <h2>実行結果 (ライブラリ出力)</h2>
            <div class="raw-data" style="font-family: Consolas, monospace;">{escaped_nslookup}</div>
        </div>
        """
        contents_html += nslookup_content

    # --- 4. ドメイン情報 (TLD) ---
    if domain_name_for_nslookup and report_opts.get("tld", True):
        tab_id = "tab-domain-info"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">ドメイン情報</button>\n'
        
        tld_val = domain_name_for_nslookup.split('.')[-1].lower() if '.' in domain_name_for_nslookup else ""
        tld_data = TLD_INFO.get(tld_val, {"name": "その他の国・地域 / gTLD", "jp_name": "未分類", "url": "N/A"})
        iana_url = f"https://www.iana.org/domains/root/db/{tld_val}.html" if tld_val else ""
        
        tld_url_html = f'<a href="{tld_data["url"]}" target="_blank" style="color: #0066cc; font-weight: bold;">{tld_data["url"]}</a>' if tld_data["url"] != "N/A" else "情報なし"
        iana_url_html = f'<a href="{iana_url}" target="_blank" style="color: #0066cc; font-weight: bold;">{iana_url}</a>' if iana_url else "情報なし"
        
        domain_info_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-rdap" style="color: #2e7d32; border-color: #2e7d32;">ドメイン統括情報 (TLD)</h1>
            <div class="description" style="background-color: #e8f5e9; border-color: #a5d6a7;">
                <strong>Domain Information：</strong><br>
                対象ドメインを管轄するトップレベルドメイン(TLD)情報を示す。
            </div>
            <h2>対象ドメイン</h2>
            <table>
                <tr><th>対象ドメイン<br>(Target Domain)</th><td><strong style="font-size: 1.2em;">{html.escape(domain_name_for_nslookup)}</strong></td></tr>
            </table>
            <h2>トップレベルドメイン (TLD) 管理情報</h2>
            <table>
                <tr><th>TLD</th><td><strong>.{html.escape(tld_val)}</strong></td></tr>
                <tr><th>管轄国 / 種別 (日本語)</th><td><strong>{html.escape(tld_data['jp_name'])}</strong> ({html.escape(tld_data['name'])})</td></tr>
                <tr><th>管理元 (レジストリ) URL</th><td>{tld_url_html}</td></tr>
                <tr><th>IANA 公式データベース</th><td>{iana_url_html}</td></tr>
            </table>
        </div>
        """
        contents_html += domain_info_content

    # --- 4.1 RDAP (Domain) ---
    if domain_rdap_json and domain_rdap_url and report_opts.get("rdap", True):
        tab_id = "tab-domain-rdap"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">RDAP(Domain)</button>\n'
        
        parsed_url_d = urlparse(domain_rdap_url)
        registry_name_d = parsed_url_d.netloc if parsed_url_d.netloc else "不明"

        events = domain_rdap_json.get("events", [])
        reg_date = "情報なし"
        exp_date = "情報なし"
        for ev in events:
            if ev.get("eventAction") == "registration":
                reg_date = ev.get("eventDate", "情報なし")
            elif ev.get("eventAction") == "expiration":
                exp_date = ev.get("eventDate", "情報なし")

        entities = domain_rdap_json.get("entities", [])
        registrar_name = "情報なし"
        registrant_org = "情報なし"
        for ent in entities:
            roles = ent.get("roles", [])
            vcard_array = ent.get("vcardArray", [])
            if len(vcard_array) > 1:
                for vcard in vcard_array[1]:
                    if "registrar" in roles and vcard[0] == "fn":
                        registrar_name = vcard[3]
                    if "registrant" in roles and vcard[0] in ["org", "fn"]:
                        registrant_org = vcard[3]

        nameservers_data = domain_rdap_json.get("nameservers", [])
        ns_list = [ns.get("ldhName") for ns in nameservers_data if ns.get("ldhName")]
        ns_html = "<br>".join(ns_list) if ns_list else "情報なし"
        
        raw_json_str_d = json.dumps(domain_rdap_json, indent=4, ensure_ascii=False)
        escaped_json_d = html.escape(raw_json_str_d)
        
        # ダブルクォーテーションがエスケープされた &quot; に対応したハイライト
        highlight_keys_d = ['registrar', 'registrant', 'registration', 'expiration', 'nameservers']
        for hk in highlight_keys_d:
            escaped_json_d = escaped_json_d.replace(f'&quot;{hk}&quot;', f'<span class="json-hl">&quot;{hk}&quot;</span>')
        
        for val in [registrar_name, registrant_org, reg_date, exp_date]:
            if val and val != "情報なし":
                esc_val = html.escape(val)
                escaped_json_d = escaped_json_d.replace(f'&quot;{esc_val}&quot;', f'<span class="json-hl">&quot;{esc_val}&quot;</span>')

        domain_rdap_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-rdap" style="color: #2e7d32; border-color: #2e7d32;">RDAP取得結果（ドメイン属性）</h1>
            <div class="description" style="background-color: #e8f5e9; border-color: #a5d6a7;">
                <strong>Domain RDAP Information：</strong><br>
                公式レジストリから直接取得した法的登録情報（RDAP）を示す。
            </div>
            <h2>対象ドメイン及び回答元レジストリ情報等</h2>
            <table>
                <tr><th>対象ドメイン<br>(Target Domain)</th><td><strong style="font-size: 1.2em;">{html.escape(domain_name_for_nslookup)}</strong></td></tr>
                <tr><th>回答元レジストリ<br>(Registry)</th><td><strong>{registry_name_d}</strong></td></tr>
                <tr><th>参照元URL<br>(Source)</th><td><a href="{domain_rdap_url}" target="_blank" style="color: #0066cc; word-break: break-all; font-weight: bold;">{domain_rdap_url}</a></td></tr>
            </table>
            <h2>RDAP取得結果</h2>
            <table>
                <tr><th>登録組織名<br>(Registrant Org)</th><td><strong>{registrant_org}</strong></td></tr>
                <tr><th>管理レジストラ<br>(Registrar)</th><td><strong>{registrar_name}</strong></td></tr>
                <tr><th>ネームサーバー<br>(Nameservers)</th><td><strong>{ns_html}</strong></td></tr>
                <tr><th>登録日時<br>(Registration Date)</th><td><strong>{reg_date}</strong></td></tr>
                <tr><th>有効期限<br>(Expiration Date)</th><td><strong>{exp_date}</strong></td></tr>
            </table>
            <h2>参照元データ (JSON形式)</h2>
            <div class="raw-data">{escaped_json_d}</div>
        </div>
        """
        contents_html += domain_rdap_content

    # --- 4.2 WHOIS (Domain) ---
    if domain_whois_text and report_opts.get("whois", True):
        tab_id = "tab-domain-whois"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">WHOIS(Domain)</button>\n'
        
        escaped_whois = html.escape(str(domain_whois_text))
        escaped_server = html.escape(str(domain_whois_server))
        escaped_domain = html.escape(domain_name_for_nslookup)
        
        equiv_cmd = f"$ whois -h {escaped_server} {escaped_domain}"
        
        domain_whois_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-rdap" style="color: #607d8b; border-color: #607d8b;">WHOIS取得結果（Port 43）</h1>
            <div class="description" style="background-color: #eceff1; border-color: #cfd8dc;">
                <strong>Domain WHOIS Information：</strong><br>
                Whoisプロトコル（ポート43経由）で権威サーバーから取得した生のWHOISテキストデータを示す。
            </div>
            <h2>対象ドメイン及び回答元サーバー</h2>
            <table>
                <tr><th>対象ドメイン<br>(Target Domain)</th><td><strong style="font-size: 1.2em;">{escaped_domain}</strong></td></tr>
                <tr><th>回答元サーバー<br>(WHOIS Server)</th><td><strong>{escaped_server}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
            </table>
            <h2>実行コマンド</h2>
            <div class="raw-data" style="background-color: #263238; color: #eceff1; font-weight: bold; font-family: Consolas, monospace; margin-bottom: 20px;">{equiv_cmd}</div>
            <h2>WHOISデータ</h2>
            <div class="raw-data" style="font-family: Consolas, monospace; white-space: pre-wrap;">{escaped_whois}</div>
        </div>
        """
        contents_html += domain_whois_content

    # --- 5. RDAP (IP) ---
    if rdap_json and rdap_url and report_opts.get("rdap", True):
        tab_id = "tab-rdap"
        if not first_tab_id: first_tab_id = tab_id
        
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">RDAP(IP)</button>\n'
        
        actual_rdap_url = rdap_url
        for link in rdap_json.get("links", []):
            if link.get("rel") == "self":
                actual_rdap_url = link.get("href", actual_rdap_url)
                break
        
        name_val = rdap_json.get("name", "情報なし")
        country_val = rdap_json.get("country", "情報なし")
        start_ip = rdap_json.get("startAddress", "情報なし")
        end_ip = rdap_json.get("endAddress", "情報なし")
        
        remarks_list = rdap_json.get("remarks", [])
        descriptions = []
        for remark in remarks_list:
            desc = remark.get("description", [])
            if isinstance(desc, list):
                descriptions.extend(desc)
            elif isinstance(desc, str):
                descriptions.append(desc)
        
        remarks_html = ""
        if descriptions:
            remarks_text = "<br>".join(descriptions)
            remarks_html = f"""
                <tr>
                    <th>備考・プロジェクト情報<br>(Remarks / Description)</th>
                    <td><strong>{remarks_text}</strong><span class="help-text">RDAPデータの備考欄に記載されている付加情報であり、保有者と運用者が異なる理由等が記載されている場合がある。</span></td>
                </tr>
            """

        country_display = country_val
        if country_val == "JP": country_display = "JP (Japan)"
        elif country_val == "US": country_display = "US (United States)"
        
        parsed_url = urlparse(actual_rdap_url)
        registry_name = parsed_url.netloc if parsed_url.netloc else "RDAP"

        raw_json_str = json.dumps(rdap_json, indent=4, ensure_ascii=False)
        escaped_json = html.escape(raw_json_str)
        
        # &quot; に対応した正規表現でのハイライト処理
        highlight_keys = ['name', 'country', 'startAddress', 'endAddress']
        for hk in highlight_keys:
            simple_pattern = r'(&quot;' + hk + r'&quot;:\s*&quot;.*?&quot;)'
            escaped_json = re.sub(simple_pattern, r'<span class="json-hl">\1</span>', escaped_json)
        
        if descriptions:
            escaped_json = re.sub(r'(&quot;(remarks|description)&quot;\s*:)', r'<span class="json-hl">\1</span>', escaped_json)
            for desc in descriptions:
                esc_desc = html.escape(desc)
                target_str = f'&quot;{esc_desc}&quot;'
                replacement = f'<span class="json-hl">{target_str}</span>'
                escaped_json = escaped_json.replace(target_str, replacement)

        rdap_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-rdap" style="color: #1e3a8a; border-color: #1e3a8a;">RDAP取得結果（IPアドレス属性）</h1>
            <div class="description" style="background-color: #e3f2fd; border-color: #90caf9;">
                <strong>IP RDAP Information：</strong><br>
                インターネット資源（IPアドレス等）の登録主体（組織又は個人）を法的に特定し得る登録情報を取得した結果を示す。
            </div>
            <h2>対象IPアドレス及び回答元レジストリ情報等</h2>
            <table>
                <tr><th>対象IPアドレス<br>(Target IP)</th><td><strong>{clean_ip}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
                <tr><th>回答元レジストリ<br>(Registry)</th><td><strong>{registry_name}</strong></td></tr>
                <tr><th>参照元URL<br>(Source)</th><td><a href="{actual_rdap_url}" target="_blank" style="color: #0066cc; word-break: break-all; font-weight: bold;">{actual_rdap_url}</a></td></tr>
            </table>
            <h2>RDAP取得結果</h2>
            <table>
                <tr><th>法的保有者<br>(Key: name)</th><td><strong>{name_val}</strong><span class="help-text">対象のIPアドレスブロックを公式に管理・保有している組織名（レジストリ登録情報）を示す。</span></td></tr>
                {remarks_html}
                <tr><th>登録国コード<br>(Key: country)</th><td><strong>{country_display}</strong><span class="help-text">当該IPアドレス資源が法的に割り当てられている管轄国を示す。</span></td></tr>
                <tr><th>IPアドレス割当範囲<br>(Key: startAddress, endAddress)</th><td><strong>{start_ip} ～ {end_ip}</strong><span class="help-text">対象のIPアドレスを包含する、レジストリから当該組織に対して運用および管理権限が委譲されたIPアドレス帯域を示す。</span></td></tr>
            </table>
            <h2>参照元データ (JSON形式)</h2>
            <div class="raw-data">{escaped_json}</div>
        </div>
        """
        contents_html += rdap_content

    # --- 5.1 WHOIS (IP) ---
    if ip_whois_text and report_opts.get("whois", True):
        tab_id = "tab-ip-whois"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">WHOIS(IP)</button>\n'
        
        escaped_whois = html.escape(str(ip_whois_text))
        escaped_server = html.escape(str(ip_whois_server))
        escaped_ip = html.escape(clean_ip)
        
        equiv_cmd = f"$ whois -h {escaped_server} {escaped_ip}"
        
        ip_whois_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-rdap" style="color: #607d8b; border-color: #607d8b;">WHOIS取得結果（Port 43）</h1>
            <div class="description" style="background-color: #eceff1; border-color: #cfd8dc;">
                <strong>IP WHOIS Information：</strong><br>
                WHOISプロトコル（ポート43経由）で取得したWHOISテキストデータを示す。
            </div>
            <h2>対象IPアドレス及び回答元サーバー</h2>
            <table>
                <tr><th>対象IPアドレス<br>(Target IP)</th><td><strong>{escaped_ip}</strong></td></tr>
                <tr><th>回答元サーバー<br>(WHOIS Server)</th><td><strong>{escaped_server}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
            </table>
            <h2>実行コマンド</h2>
            <div class="raw-data" style="background-color: #263238; color: #eceff1; font-weight: bold; font-family: Consolas, monospace; margin-bottom: 20px;">{equiv_cmd}</div>
            <h2>WHOISデータ</h2>
            <div class="raw-data" style="font-family: Consolas, monospace; white-space: pre-wrap;">{escaped_whois}</div>
        </div>
        """
        contents_html += ip_whois_content

    # --- 6. IPinfo ---
    if ipinfo_json and report_opts.get("ipinfo", True):
        tab_id = "tab-ipinfo"
        if not first_tab_id: first_tab_id = tab_id
        
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">IPinfo</button>\n'
        
        raw_json_str = json.dumps(ipinfo_json, indent=4, ensure_ascii=False)
        escaped_json = html.escape(raw_json_str)
        
        # &quot; 対応
        highlight_keys = ['ip', 'hostname', 'city', 'region', 'country', 'loc', 'org']
        for hk in highlight_keys:
            simple_pattern = r'(&quot;' + hk + r'&quot;:\s*&quot;.*?&quot;)'
            escaped_json = re.sub(simple_pattern, r'<span class="json-hl">\1</span>', escaped_json)

        ip_val = ipinfo_json.get("ip", "情報なし")
        hostname_val = ipinfo_json.get("hostname", "情報なし")
        city_val = ipinfo_json.get("city", "情報なし")
        region_val = ipinfo_json.get("region", "情報なし")
        country_val = ipinfo_json.get("country", "情報なし")
        loc_val = ipinfo_json.get("loc", "情報なし")
        org_val = ipinfo_json.get("org", "情報なし")
        
        req_ipinfo_url = f"https://ipinfo.io/{ip_val if ip_val != '情報なし' else clean_ip}"

        privacy_html = ""
        geo_heading = "<h2>地理的情報</h2>"

        if "privacy" in ipinfo_json:
            privacy = ipinfo_json.get("privacy", {})
            privacy_flags = []
            if privacy.get("vpn"): privacy_flags.append("VPN")
            if privacy.get("proxy"): privacy_flags.append("Proxy")
            if privacy.get("tor"): privacy_flags.append("Tor")
            if privacy.get("relay"): privacy_flags.append("Relay")
            if privacy.get("hosting"): privacy_flags.append("Hosting")
            
            if privacy_flags:
                geo_heading = "<h2>地理的情報・匿名化判定</h2>"
                privacy_val = ", ".join(privacy_flags)
                privacy_html = f"""
                    <tr>
                        <th>プライバシー・リスク判定<br>(Privacy Status)</th>
                        <td><strong>{privacy_val}</strong><span class="help-text">VPN、Proxy、Tor、Hosting等として利用されているかを判定した結果。</span></td>
                    </tr>
                """

        map_html = ""
        if loc_val != "情報なし" and "," in loc_val:
            map_url = f"https://maps.google.com/maps?q={loc_val}&hl=ja&z=14&output=embed"
            map_html = f"""
            <h2>位置情報マップ</h2>
            <div class="map-container" style="width: 100%; height: 400px; margin-bottom: 20px; border: 1px solid #ccc; border-radius: 5px; overflow: hidden;">
                <iframe width="100%" height="100%" frameborder="0" scrolling="no" marginheight="0" marginwidth="0" src="{map_url}"></iframe>
            </div>
            """

        ipinfo_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-ipinfo">Whois属性及び地理位置情報取得結果</h1>
            <div class="description">
                <strong>IPinfo（IP Geolocation Data）：</strong><br>
                IPinfoとは、IPアドレスに基づき、当該アドレスの推定地理的位置（国、都市、地域、郵便番号、緯度経度等）、所属組織（ASN、ISP名、ドメイン等）、ネットワーク特性を判定するサービスである。
            </div>
            <h2>基本情報</h2>
            <table>
                <tr><th>対象IPアドレス<br>(Key: ip)</th><td><strong>{ip_val}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
                <tr><th>リクエストURL<br>(Request URL)</th><td><a href="{req_ipinfo_url}" target="_blank" style="color: #00897b; word-break: break-all;">{req_ipinfo_url}</a></td></tr>
            </table>
            <h2>IP情報取得結果</h2>
            <table>
                <tr><th>ホストネーム<br>(Key: hostname)</th><td><strong>{hostname_val}</strong></td></tr>
                <tr><th>組織/ISP<br>(Key: org)</th><td><strong>{org_val}</strong></td></tr>
            </table>
            {geo_heading}
            <table>
                <tr><th>地域<br>(Key: city, region, country)</th><td><strong>{country_val}, {region_val}, {city_val}</strong></td></tr>
                <tr><th>推定座標<br>(Key: loc)</th><td><strong>{loc_val}</strong></td></tr>
                {privacy_html}
            </table>
            {map_html}
            <h2>参照元データ (JSON形式)</h2>
            <div class="raw-data">{escaped_json}</div>
        </div>
        """
        contents_html += ipinfo_content

    # --- 7. VPNAPI.io ---
    if vpnapi_json and report_opts.get("vpnapi", True):
        tab_id = "tab-vpnapi"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">VPNAPI.io</button>\n'
        
        sec = vpnapi_json.get('security', {})
        if any(sec.values()):
            proxy_status_text = "該当あり (匿名通信検知)"
            status_color = "red"
            detected_types = [k.upper() for k, v in sec.items() if v]
            p_type_val = " / ".join(detected_types)
        else:
            proxy_status_text = "該当なし"
            status_color = "green"
            p_type_val = "情報なし"

        net = vpnapi_json.get('network', {})
        org_val = net.get('autonomous_system_organization', '情報なし')

        loc = vpnapi_json.get('location', {})
        c_name_val = loc.get('country', '情報なし')

        raw_json_str = json.dumps(vpnapi_json, indent=4, ensure_ascii=False)
        escaped_json = html.escape(raw_json_str)

        # " 対応
        highlight_keys_vpn = ['vpn', 'proxy', 'tor', 'relay', 'country', 'ip', 'autonomous_system_organization']
        for hk in highlight_keys_vpn:
            simple_pattern = r'("' + hk + r'":\s*.*?,?\n)'
            escaped_json = re.sub(simple_pattern, r'<span class="json-hl">\1</span>', escaped_json)

        vpn_req_ip = vpnapi_json.get('ip', clean_ip)
        req_vpnapi_url = f"https://vpnapi.io/api/{vpn_req_ip}?key=********"

        vpnapi_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-ip2proxy">匿名通信判定結果 (VPNAPI.io)</h1>
            <div class="description" style="background-color: #f3e5f5; border-color: #ce93d8;">
                <strong>VPNAPI.io:</strong><br>
                VPNAPI.ioは、対象IPアドレスがVPN、プロキシ、Torノード、またはリレーネットワークとして利用されているかを検知するための高精度データベースである。
            </div>
            <h2>基本情報</h2>
            <table>
                <tr><th>対象IPアドレス<br>(Key: ip)</th><td><strong>{vpn_req_ip}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
                <tr><th>リクエストURL<br>(Request URL)</th><td><a href="{req_vpnapi_url}" target="_blank" style="color: #6a1b9a; word-break: break-all;">{req_vpnapi_url}</a></td></tr>
            </table>
            <h2>VPNAPI.io 取得結果</h2>
            <table>
                <tr><th>プロキシ判定<br>(Security)</th><td><strong style="color:{status_color};">{proxy_status_text}</strong></td></tr>
                <tr><th>検知種別<br>(Detected Types)</th><td><strong>{p_type_val}</strong></td></tr>
                <tr><th>運用組織名<br>(Key: autonomous_system_organization)</th><td><strong>{org_val}</strong></td></tr>
                <tr><th>判定国名<br>(Key: country)</th><td><strong>{c_name_val}</strong></td></tr>
            </table>
            <h2>解析用生データ (JSON形式)</h2>
            <div class="raw-data">{escaped_json}</div>
        </div>
        """
        contents_html += vpnapi_content

    # --- 8. SecurityTrails ---
    if st_json and report_opts.get("st", True):
        tab_id = "tab-st"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">SecurityTrails</button>\n'
        
        records = st_json.get("records", [])
        is_date_filtered = st_json.get("is_date_filtered", False)
        # 日付文字列を取得
        st_start_date_str = st_json.get("start_date", "")
        st_end_date_str = st_json.get("end_date", "")
        
        st_html_rows = ""
        unique_ips_ordered = []
        seen_ips = set()

        # ▼ 変更: エラーフラグ時は警告メッセージを強制挿入
        if st_json.get("error") == "rate_limit":
            st_html_rows = "<tr><td colspan='4' style='text-align:center; color:#c62828;'><b>🚨 月間の無料APIリクエスト枠（50回）に到達したため、履歴の取得がブロックされました。</b></td></tr>"
            unique_ips_rows = "<tr><td style='text-align:center;'>-</td></tr>"
        else:
            for rec in records: 
                values = rec.get("values", [])
                ips_in_rec = []
                for v in values:
                    ip_val = v.get("ip", "")
                    if ip_val:
                        ips_in_rec.append(html.escape(str(ip_val)))
                        if ip_val not in seen_ips:
                            seen_ips.add(ip_val)
                            unique_ips_ordered.append(ip_val)

                ips = "<br>".join(ips_in_rec)
                first_seen = html.escape(str(rec.get("first_seen", "情報なし")))
                last_seen = html.escape(str(rec.get("last_seen", "情報なし")))
                orgs = rec.get("organizations", [])
                org = html.escape(str(orgs[0])) if orgs and orgs[0] else "情報なし"
                st_html_rows += f"<tr><td>{ips}</td><td>{first_seen}</td><td>{last_seen}</td><td>{org}</td></tr>"
                
            if not st_html_rows:
                st_html_rows = "<tr><td colspan='4' style='text-align:center;'>A/AAAAレコードの履歴データが見つかりませんでした。</td></tr>"
                
            unique_ips_rows = ""
            for ip in unique_ips_ordered:
                unique_ips_rows += f"<tr><td><strong>{html.escape(str(ip))}</strong></td></tr>"
            if not unique_ips_rows:
                unique_ips_rows = "<tr><td style='text-align:center;'>取得されたIPアドレスはありません。</td></tr>"

        raw_json_str_st = json.dumps(st_json, indent=4, ensure_ascii=False)
        escaped_json_st = html.escape(raw_json_str_st)
        
        # ハイライト修正 (&quot; に対応)
        highlight_keys_st = ['ip']
        for hk in highlight_keys_st:
            simple_pattern = r'((?:&quot;|")' + hk + r'(?:&quot;|")\s*:\s*[^\n\r]*)'
            escaped_json_st = re.sub(simple_pattern, r'<span class="json-hl">\1</span>', escaped_json_st)
            
        table_heading = "レコード履歴 (抽出結果全件)" if is_date_filtered else "レコード履歴 (最新20件)"
        target_domain_esc = html.escape(str(domain_name_for_nslookup))
        url_a = f"https://api.securitytrails.com/v1/history/{target_domain_esc}/dns/a"
        url_aaaa = f"https://api.securitytrails.com/v1/history/{target_domain_esc}/dns/aaaa"
        
        # 期間指定がある場合のみ挿入するHTMLを生成
        date_filter_html = ""
        if is_date_filtered and st_start_date_str and st_end_date_str:
            date_filter_html = f"""
                <tr><th>抽出期間 (開始日)<br>(Start Date)</th><td><strong>{html.escape(str(st_start_date_str))}</strong></td></tr>
                <tr><th>抽出期間 (終了日)<br>(End Date)</th><td><strong>{html.escape(str(st_end_date_str))}</strong></td></tr>
            """
        
        st_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-ip2proxy" style="color: #e65100; border-color: #e65100;">レコード履歴 (SecurityTrails)</h1>
            <div class="description" style="background-color: #fff3e0; border-color: #ffcc80;">
                <strong>SecurityTrails Historical DNS Data：</strong><br>
                SecurityTrailsのAPIを利用して、対象ドメインに過去紐付いていたIPアドレス（Aレコード/AAAAレコード）の変遷を取得した結果を示す。
            </div>
            <h2>対象ドメイン及び取得情報</h2>
            <table>
                <tr><th>対象ドメイン<br>(Target Domain)</th><td><strong>{target_domain_esc}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
                {date_filter_html}
                <tr><th>リクエストURL<br>(Request URL)</th><td>
                    <a href="{url_a}" target="_blank" style="color: #0066cc; word-break: break-all;">{url_a}</a><br>
                    <a href="{url_aaaa}" target="_blank" style="color: #0066cc; word-break: break-all;">{url_aaaa}</a>
                </td></tr>
            </table>
            <h2>判明したIPアドレス一覧 (重複排除)</h2>
            <table>
                <tr><th>抽出されたIPアドレス (IPv4/IPv6)</th></tr>
                {unique_ips_rows}
            </table>
            <h2>{table_heading}</h2>
            <table>
                <tr><th>IPアドレス (IPv4/IPv6)</th><th>初回観測日 (First Seen)</th><th>最終観測日 (Last Seen)</th><th>組織 (Organization)</th></tr>
                {st_html_rows}
            </table>
            <h2>参照元データ (JSON形式)</h2>
            <div class="raw-data">{escaped_json_st}</div>
        </div>
        """
        contents_html += st_content

    # --- 8.5 SecurityTrails (Reverse IP) ---
    st_rev_json = res.get('ST_REVERSE_IP_JSON')
    if st_rev_json and report_opts.get("st", True):
        tab_id = "tab-st-revip"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">Reverse IP</button>\n'
        
        records = st_rev_json.get("records", [])
        total_records = st_rev_json.get("meta", {}).get("total_records", len(records))
        total_pages = st_rev_json.get("meta", {}).get("total_pages", 1)
        
        # 取得件数表示の動的変更
        if total_pages > 1:
            if len(records) >= total_records:
                display_count_text = f"<strong>{total_records} 件</strong> (全件取得済)"
            else:
                display_count_text = f"<strong>{total_records} 件</strong> (※最初の {len(records)} 件のみ表示)"
        else:
            display_count_text = f"<strong>{total_records} 件</strong>"
            
        rev_html_rows = ""
        
        if st_rev_json.get("error") == "rate_limit":
            rev_html_rows = "<tr><td style='text-align:center; color:#c62828;'><b>🚨 月間の無料APIリクエスト枠（50回）に到達したため、取得がブロックされました。</b></td></tr>"
            display_count_text = "エラー (上限到達)"
        elif st_rev_json.get("error") == "rate_limit_during_pagination":
            # ページめくり中に制限に達した場合の特別警告
            rev_html_rows = f"<tr><td style='text-align:center; background-color:#fff3e0; color:#e65100;'><b>⚠️ 全件取得の途中でAPI上限に到達しました。取得できた {len(records)} 件までを表示します。</b></td></tr>"
            for rec in records:
                hostname = rec.get("hostname", "")
                if hostname:
                    rev_html_rows += f"<tr><td><strong>{html.escape(str(hostname))}</strong></td></tr>"
        else:
            for rec in records:
                hostname = rec.get("hostname", "")
                if hostname:
                    rev_html_rows += f"<tr><td><strong>{html.escape(str(hostname))}</strong></td></tr>"
                    
            if not rev_html_rows:
                rev_html_rows = "<tr><td style='text-align:center;'>紐づくドメインは見つかりませんでした。</td></tr>"
            
        raw_json_str_rev = json.dumps(st_rev_json, indent=4, ensure_ascii=False)
        escaped_json_rev = html.escape(raw_json_str_rev)
        
        # JSONハイライト機能の追加："hostname" の行を対象にする
        import re
        simple_pattern_rev = r'((?:&quot;|")hostname(?:&quot;|"):\s*(?:&quot;|").*?(?:&quot;|"))'
        escaped_json_rev = re.sub(simple_pattern_rev, r'<span class="json-hl">\1</span>', escaped_json_rev)
        
        rev_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-ip2proxy" style="color: #0288d1; border-color: #0288d1;">Reverse IP 検索結果 (SecurityTrails)</h1>
            <div class="description" style="background-color: #e1f5fe; border-color: #81d4fa;">
                <strong>Reverse IP Lookup：</strong><br>
                対象のIPアドレスがA/AAAAレコードとして設定されているドメイン群を、SecurityTrailsのPassive DNSデータベースから逆引き検索した結果を示す。同一インフラに同居している他のサービス群の特定に利用する。
            </div>
            <h2>対象IPアドレス及び取得結果</h2>
            <table>
                <tr><th>対象IPアドレス<br>(Target IP)</th><td><strong>{clean_ip}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
                <tr><th>ヒット総数<br>(Total Records)</th><td>{display_count_text}</td></tr>
            </table>
            <h2>紐づくドメイン一覧</h2>
            <table>
                <tr><th>ドメイン名 (Hostname)</th></tr>
                {rev_html_rows}
            </table>
            <h2>参照元データ (JSON形式)</h2>
            <div class="raw-data">{escaped_json_rev}</div>
        </div>
        """
        contents_html += rev_content

    # --- 9. rDNS ---
    if rdns_raw and report_opts.get("rdns", True):
        tab_id = "tab-rdns"
        if not first_tab_id: first_tab_id = tab_id
        tabs_html += f'<button class="tab-button" onclick="openTab(event, \'{tab_id}\')" id="btn-{tab_id}">逆引き(rDNS)</button>\n'
        
        escaped_rdns = html.escape(rdns_raw)
        for h_str in rdns_hosts:
            escaped_h = html.escape(h_str)
            escaped_rdns = escaped_rdns.replace(escaped_h, f'<span class="json-hl">{escaped_h}</span>')
            
        cmd_str = f"resolver = dns.resolver.Resolver(configure=False); resolver.nameservers = ['8.8.8.8']; resolver.resolve(dns.reversename.from_address('{clean_ip}'), 'PTR')"
        host_list_str = "<br>".join([html.escape(h) for h in rdns_hosts]) if rdns_hosts else "取得なし"
        
        rdns_content = f"""
        <div id="{tab_id}" class="tab-content">
            <h1 class="theme-rdap" style="color: #424242; border-color: #424242;">DNS逆引き解決結果 (dnspython)</h1>
            <div class="description" style="background-color: #eceff1; border-color: #cfd8dc;">
                <strong>DNS逆引き(Reverse DNS) 解決記録：</strong><br>
                対象のIPアドレスに対してPython公式の国際標準ライブラリである <code>dnspython</code> を実行し、紐づくホスト名（PTRレコード）を取得した結果を示す。
            </div>
            <h2>対象IPアドレス及び取得結果</h2>
            <table>
                <tr><th>対象IPアドレス<br>(Target IP)</th><td><strong>{clean_ip}</strong></td></tr>
                <tr><th>取得日時<br>(Timestamp)</th><td><strong>{current_time_str}</strong></td></tr>
                <tr><th>取得ホスト名<br>(Resolved Hostnames)</th><td><strong>{host_list_str}</strong></td></tr>
            </table>
            <h2>内部実行クエリ (Python)</h2>
            <div class="raw-data" style="background-color: #263238; color: #eceff1; font-weight: bold; font-family: Consolas, monospace;">>>> {cmd_str}</div>
            <h2>実行結果 (ライブラリ出力)</h2>
            <div class="raw-data" style="font-family: Consolas, monospace;">{escaped_rdns}</div>
        </div>
        """
        contents_html += rdns_content

    # --- 10. 統合HTML構築 ---
    full_html = f"""
    <!DOCTYPE html>
    <html lang="ja">
    <head>
        <meta charset="UTF-8">
        <title>IP-OSINT - {clean_ip}</title>
        <style>
            body {{ font-family: 'Helvetica Neue', Arial, sans-serif; padding: 30px; color: #333; line-height: 1.6; max-width: 800px; margin: 0 auto; }}
            .tab-container {{ margin-bottom: 20px; border-bottom: 2px solid #ccc; display: flex; }}
            .tab-button {{ background-color: #f8f9fa; border: 1px solid #ccc; border-bottom: none; outline: none; cursor: pointer; padding: 10px 20px; font-size: 16px; font-weight: bold; color: #555; border-radius: 5px 5px 0 0; margin-right: 5px; transition: 0.3s; }}
            .tab-button:hover {{ background-color: #e9ecef; }}
            .tab-button.active {{ background-color: #1e3a8a; color: white; border-color: #1e3a8a; }}
            .tab-content {{ display: none; animation: fadeEffect 0.4s; }}
            @keyframes fadeEffect {{ from {{opacity: 0;}} to {{opacity: 1;}} }}
            h1 {{ font-size: 24px; border-bottom: 2px solid; padding-bottom: 5px; text-align: center; margin-bottom: 30px;}}
            h1.theme-rdap {{ color: #1e3a8a; border-color: #1e3a8a; }}
            h1.theme-ipinfo {{ color: #00897b; border-color: #00897b; }}
            h1.theme-ip2proxy {{ color: #6a1b9a; border-color: #6a1b9a; }}
            h2 {{ font-size: 18px; margin-top: 30px; border-left: 4px solid #666; padding-left: 10px; }}
            .description {{ background-color: #f8f9fa; padding: 15px; border-radius: 5px; border: 1px solid #e9ecef; margin-bottom: 20px; font-size: 14px; text-align: justify; }}
            table {{ width: 100%; border-collapse: collapse; margin-bottom: 20px; }}
            th, td {{ border: 1px solid #ddd; padding: 12px; text-align: left; font-size: 14px; vertical-align: top; }}
            th {{ background-color: #f2f2f2; width: 30%; }}
            .help-text {{ font-size: 12px; color: #666; display: block; margin-top: 4px; line-height: 1.4; }}
            .raw-data {{ font-family: monospace; background-color: #f4f4f4; padding: 15px; border-radius: 5px; white-space: pre-wrap; font-size: 12px; border: 1px solid #ccc; word-break: break-all; }}
            .json-hl {{ background-color: #fff59d; color: #c62828; font-weight: bold; border-radius: 2px; padding: 1px 3px; transition: 0.3s; -webkit-print-color-adjust: exact; print-color-adjust: exact; }}
            body.hide-hl .json-hl {{ background-color: transparent; color: inherit; font-weight: normal; padding: 0; }}
            body.hide-desc .description, body.hide-desc .help-text {{ display: none; }}
            body.compress-json .raw-data {{ white-space: normal; word-break: break-all; }}
            .controls {{ margin-bottom: 20px; text-align: right; background: #e3f2fd; padding: 10px; border-radius: 5px; border: 1px solid #bbdefb; }}
            .controls label {{ font-size: 14px; cursor: pointer; font-weight: bold; color: #1565c0; margin-right: 15px; display: inline-block; margin-bottom: 5px; }}
            .controls button {{ padding: 8px 16px; font-size: 14px; cursor: pointer; background-color: #1e3a8a; color: white; border: none; border-radius: 3px; transition: background 0.3s; margin-top: 5px; }}
            .controls button:hover {{ background-color: #1565c0; }}
            @media print {{
                body {{ padding: 0; max-width: 100%; }}
                .no-print, .tab-container {{ display: none !important; }}
                .tab-content {{ display: block !important; page-break-after: always; }}
                .tab-content:last-child {{ page-break-after: auto; }}
                .raw-data {{ page-break-inside: auto; }}
                .map-container iframe {{ width: 100% !important; }}
            }}
        </style>
    </head>
    <body>
        <div class="controls no-print">
            <div>
                <label><input type="checkbox" checked onchange="document.body.classList.toggle('hide-desc', !this.checked)"> 解説・ヘルプテキストを表示</label>
                <label><input type="checkbox" checked onchange="document.body.classList.toggle('hide-hl', !this.checked)"> JSONのハイライトを有効化</label>
                <label><input type="checkbox" onchange="document.body.classList.toggle('compress-json', this.checked)"> 生データ(JSON)を圧縮表示</label>
            </div>
            <button onclick="window.print()">🖨️ すべての情報を一括印刷</button>
        </div>
        
        <div class="tab-container no-print">
            {tabs_html}
        </div>
        
        {contents_html}

        <script>
            function openTab(evt, tabId) {{
                let i, tabcontent, tablinks;
                tabcontent = document.getElementsByClassName("tab-content");
                for (i = 0; i < tabcontent.length; i++) {{
                    tabcontent[i].style.display = "none";
                }}
                tablinks = document.getElementsByClassName("tab-button");
                for (i = 0; i < tablinks.length; i++) {{
                    tablinks[i].className = tablinks[i].className.replace(" active", "");
                }}
                document.getElementById(tabId).style.display = "block";
                if(evt) {{
                    evt.currentTarget.className += " active";
                }} else {{
                    document.getElementById("btn-" + tabId).className += " active";
                }}
            }}
            if('{first_tab_id}' !== 'None') {{ openTab(null, '{first_tab_id}'); }}
            window.onbeforeprint = function() {{
                let tabcontents = document.getElementsByClassName("tab-content");
                for (let j = 0; j < tabcontents.length; j++) {{ tabcontents[j].style.display = "block"; }}
            }};
            window.onafterprint = function() {{
                let activeTabId = "";
                let tablinks = document.getElementsByClassName("tab-button");
                for (let k = 0; k < tablinks.length; k++) {{
                    if (tablinks[k].className.indexOf("active") > -1) {{ activeTabId = tablinks[k].id.replace("btn-", ""); break; }}
                }}
                if (activeTabId) {{ openTab(null, activeTabId); }}
            }};
        </script>
    </body>
    </html>
    """
    return full_html

def generate_combined_html_report(target_results, report_opts=None):
    """複数の個別レポートを1つのHTMLファイル（目次付き）に統合する関数"""
    combined_body = ""
    toc_links = []
    common_style = ""
    
    for i, res in enumerate(target_results):
        target_ip = res.get('Target_IP', 'N/A')
        clean_ip = get_copy_target(target_ip)
        html = generate_individual_html_report(res, clean_ip, report_opts)
        if not html: continue
        
        # 1つ目のHTMLから共通スタイルを抽出
        if not common_style:
            style_match = re.search(r'<style>(.*?)</style>', html, re.DOTALL)
            if style_match: common_style = style_match.group(1)
            
        body_match = re.search(r'<body>(.*?)</body>', html, re.DOTALL)
        if body_match:
            bc = body_match.group(1)
            uid = f"_{i}"
            
            # 個別レポート用のコントロールパネルを削除
            bc = re.sub(r'<div class="controls no-print">.*?</button>\s*</div>', '', bc, flags=re.DOTALL)
            
            # 2件目以降は、長文になる解説（description）と補足（help-text）をHTMLから完全に削除して圧縮
            if i > 0:
                bc = re.sub(r'<div class="description".*?</div>', '', bc, flags=re.DOTALL)
                bc = re.sub(r'<span class="help-text">.*?</span>', '', bc, flags=re.DOTALL)
            
            # タブIDとJS関数名の置換（複数IP間でタブ切り替えが干渉しないようにUIDを付与）
            bc = bc.replace('id="tab-', f'id="tab{uid}-')
            bc = bc.replace('id="btn-tab-', f'id="btn-tab{uid}-')
            bc = bc.replace("openTab(event, 'tab-", f"openTab{uid}(event, 'tab{uid}-")
            bc = bc.replace('function openTab(evt, tabId)', f'function openTab{uid}(evt, tabId)')
            bc = bc.replace('openTab(null, ', f'openTab{uid}(null, ')
            bc = re.sub(r"openTab\{uid\}\(null,\s*'tab-(.*?)'\);", rf"openTab{uid}(null, 'tab{uid}-\1');", bc)
            
            # DOM取得スコープの限定
            bc = bc.replace('document.getElementsByClassName("tab-content")', f'document.getElementById("report-wrapper{uid}").getElementsByClassName("tab-content")')
            bc = bc.replace('document.getElementsByClassName("tab-button")', f'document.getElementById("report-wrapper{uid}").getElementsByClassName("tab-button")')
            
            # 個別印刷イベントの削除
            bc = re.sub(r'window\.onbeforeprint\s*=\s*function\(\)\s*\{.*?\};', '', bc, flags=re.DOTALL)
            bc = re.sub(r'window\.onafterprint\s*=\s*function\(\)\s*\{.*?\};', '', bc, flags=re.DOTALL)
            
            toc_links.append(f"<li><a href='#target{uid}' style='text-decoration: none; color: #1e3a8a; font-weight: bold;'>{clean_ip}</a></li>")
            
            # 重複するH2見出しを削除し、目次からのアンカー用divのみを配置
            combined_body += f"""
            <div id='target{uid}' style='margin-bottom: 60px; padding-bottom: 20px; border-bottom: 3px dashed #ccc; page-break-inside: avoid;'>
                <div id='report-wrapper{uid}'>{bc}</div>
            </div>
            """

    if not combined_body: return None
    
    # 統合用の追加CSSと印刷制御JS
    common_style += """
        body { background-color: #f0f2f5; }
        .main-container { background-color: white; padding: 40px; border-radius: 10px; box-shadow: 0 4px 6px rgba(0,0,0,0.1); max-width: 900px; margin: 0 auto; }
        .toc { background-color: #e3f2fd; padding: 20px; border-radius: 5px; margin-bottom: 40px; border: 1px solid #bbdefb; }
        .toc h2 { margin-top: 0; color: #1565c0; font-size: 22px; border-bottom: 2px solid #1565c0; padding-bottom: 5px; }
        .toc ul { column-count: 2; list-style-type: none; padding-left: 0; }
        .toc li { margin-bottom: 8px; padding: 5px; background: white; border-radius: 3px; border: 1px solid #ddd; text-align: center; }
        .toc li:hover { background: #bbdefb; }
        .combined-controls { background: #e3f2fd; padding: 15px; border-radius: 5px; border: 1px solid #bbdefb; margin-bottom: 30px; display: flex; justify-content: space-between; align-items: center; }
        .combined-controls label { font-size: 14px; cursor: pointer; font-weight: bold; color: #1565c0; margin-right: 15px; display: inline-block; }
        .combined-controls button { padding: 10px 20px; font-size: 14px; font-weight: bold; background-color: #1e3a8a; color: white; border: none; border-radius: 5px; cursor: pointer; transition: 0.3s; }
        .combined-controls button:hover { background-color: #1565c0; }
        @media print {
            body { background-color: white; }
            .main-container { box-shadow: none; padding: 0; max-width: 100%; }
            .toc { page-break-after: always; }
        }
    """
    
    combined_js = """
    <script>
        window.onbeforeprint = function() {
            let tabcontents = document.getElementsByClassName("tab-content");
            for (let j = 0; j < tabcontents.length; j++) { tabcontents[j].style.display = "block"; }
        };
    </script>
    """

    full_combined_html = f"""
    <!DOCTYPE html>
    <html lang="ja">
    <head>
        <meta charset="UTF-8">
        <title>Combined IP-OSINT Report</title>
        <style>{common_style}</style>
    </head>
    <body>
        <div class="main-container">
            <h1 style="text-align: center; color: #333; font-size: 28px; margin-bottom: 30px;">🔎 統合 IP-OSINT 調査レポート</h1>
            
            <div class="combined-controls no-print">
                <div>
                    <label><input type="checkbox" checked onchange="document.body.classList.toggle('hide-desc', !this.checked)"> 解説・ヘルプテキストを表示</label>
                    <label><input type="checkbox" checked onchange="document.body.classList.toggle('hide-hl', !this.checked)"> JSONのハイライトを有効化</label>
                    <label><input type="checkbox" onchange="document.body.classList.toggle('compress-json', this.checked)"> 生データ(JSON)を圧縮表示</label>
                </div>
                <button onclick="window.print()">🖨️ この統合レポートを一括印刷する</button>
            </div>

            <div class="toc no-print">
                <h2>📑 目次 (Target List)</h2>
                <ul>{''.join(toc_links)}</ul>
            </div>
            {combined_body}
        </div>
        {combined_js}
    </body>
    </html>
    """
    return full_combined_html

def display_results(results, current_mode_full_text, display_mode, use_rdap_option, pro_api_key, vpnapi_key, st_api_key, use_rdns_option):
    st.markdown("### 📝 検索結果")

    # SecurityTrails API制限到達時のグローバル警告
    st_limit_hit = False
    for r in results:
        st_j = r.get('ST_JSON')
        st_r = r.get('ST_REVERSE_IP_JSON')
        if isinstance(st_j, dict) and st_j.get('error') == 'rate_limit':
            st_limit_hit = True
            break
        if isinstance(st_r, dict) and st_r.get('error') == 'rate_limit':
            st_limit_hit = True
            break
            
    if st_limit_hit:
        st.error("🚨 **SecurityTrails API 利用制限の警告**: 月間の無料リクエスト枠（50回）に到達しました。一部のターゲットにおいて過去の履歴やReverse IP情報が取得できていません。")

    # --- 2. 判定アイコンと表示ルールの解説 ---
    with st.expander("⚠️ 判定アイコンと表示ルールについて"):
        st.info("""
        ### 🔍 判定ロジックの概要
        本ツールは、起動時に取得する**最新のTorノードリスト**および、**VPNAPI.io 専門データベース**とのAPI連携により、通信主体の属性を判定しています。
        
        インターネット上の通信は、その用途に応じて「個人宅・法人拠点からの直接接続」と「非対面的な中継・ホスティング経由の接続」に大別されます。本機能は後者を検知し、調査の優先順位判断を支援することを目的としています。
        
        ---
        
        ### 📌 判定種別の定義と技術的背景
        - **🧅 [Tor Node]**
            - **定義**: Tor（The Onion Router）ネットワークにおける「Exit Node（出口ノード）」を指します。
            - **背景**: 起動時にTor Project公式サイトより最新のノードリストを取得し、照合を行っています。高い匿名性を維持した通信であるため、セキュリティリスクの検討が必要です。

        - **💀 [IoT Risk]** (Shodan InternetDB連携時のみ)
            - **定義**: 外部からアクセス可能な危険なポートが開放されています。
            - **背景**: Shodanのポートスキャン履歴と照合し、ファイアウォールを通過して露出している以下の「踏み台リスク」を警告します。
                - **Telnet (23) / FTP (21)**: 暗号化されていない危険な旧式プロトコル
                - **ADB (5555/5554)**: 認証なしで操作可能なAndroid/FireTV端末・エミュレータ
                - **TR-069 (7547)**: 乗っ取りリスクのあるルーター管理機能
                - **Proxy (1080/3128/8080)**: 攻撃中継点として悪用されるプロキシ
                - **UPnP (1900)**: 外部からLAN内機器を探査される恐れのある機能
            
        - **🍏 [iCloud Private Relay]**
            - **定義**: Appleデバイス（iPhone/Mac）の標準プライバシー保護機能による通信です。
            - **背景**: Appleの提携パートナー（Cloudflare, Akamai等）が提供する出口IPを使用します。ISP名称に含まれる特定のタグ（例: "iCloud Private Relay"）に基づき判別します。基本的には一般ユーザーですが、真のIPは隠蔽されています。
            
        - **☁️ [VPN/Proxy]**
            - **定義**: 商用VPNサービス、公開プロキシ、またはプライバシー保護を目的とした中継団体に属するIPです。
            - **背景**: ISP名称に含まれる特定のキーワード（VPN, Proxy等）および既知の匿名化サービス運営組織名に基づき判別します。
            
        - **☁️ [Hosting/Infra]**
            - **定義**: クラウドサービス（AWS, Azure, GCP等）や、データセンター、ホスティング事業者のインフラストラクチャです。
            - **背景**: 一般的なコンシューマ回線とは異なり、サーバー間通信やBot、クローラー、あるいは攻撃用インフラとして利用されるケースが多いノードです。
            
        ---
        
        ※ VPNAPI.ioの設定がされていない場合は、Tor判定のみを行います。
        """)

    if not results:
        st.info("検索結果がここに表示されます。")
        return

    # --- 1. データフレーム構築 ---
    df_list = []
    orig_data_map = {}
    original_cols = []
    
    # オリジナルデータが存在する場合、IPアドレスをキーとして元データをマップする
    if st.session_state.get('original_df') is not None and st.session_state.get('ip_column_name'):
        orig_df = st.session_state['original_df']
        ip_col = st.session_state['ip_column_name']
        original_cols = [c for c in orig_df.columns if c != ip_col]
        
        for _, row in orig_df.iterrows():
            raw_val = str(row[ip_col]).strip()
            ip_val = clean_ocr_error_chars(raw_val)
            extracted_ip = extract_actual_ip(ip_val)
            # 同じIPが複数ある場合はリストとしてすべての行データを保持する
            if extracted_ip:
                if extracted_ip not in orig_data_map:
                    orig_data_map[extracted_ip] = []
                orig_data_map[extracted_ip].append(row.to_dict())

    for idx, res in enumerate(results):
        c_code = res.get('CountryCode', 'N/A')
        c_jp = res.get('Country_JP', 'N/A')
        country_display = f"{c_jp} ({c_code})"
        target_ip = res.get('Target_IP', 'N/A')
        actual_ip = extract_actual_ip(target_ip)
        
        row_data = {"No.": idx + 1}
        
        # ユーザー要望: 元データの列を一覧ビューの左側に反映
        if original_cols:
            if actual_ip in orig_data_map:
                rows_list = orig_data_map[actual_ip]
                for col in original_cols:
                    # 複数の値がある場合は重複を排除して「 / 」で結合し一覧表示する
                    vals = []
                    for r in rows_list:
                        val = str(r.get(col, '')).strip()
                        if val and val not in vals:
                            vals.append(val)
                    row_data[col] = " / ".join(vals) if vals else ""
            else:
                for col in original_cols:
                    row_data[col] = ''
        
        row_data.update({
            "IPアドレス": target_ip,
            "国名": country_display,
            "Whois(元データ)": res.get('ISP_API_Raw', 'N/A'),
            "Whois(日本語名)": res.get('ISP_JP', 'N/A'),
            "RDAP(元データ)": res.get('RDAP_Name_Raw', ''),
            "RDAP(日本語名)": res.get('RDAP_JP', ''),
            "Proxy種別": res.get('Proxy_Type', ''),
            "IoTリスク": res.get('IoT_Risk', ''),
            "ステータス": res.get('Status', 'N/A')
        })
        df_list.append(row_data)
    
    df = pd.DataFrame(df_list)

    # 💡 UIの一覧ビューからも不要なカラムを動的に消去する
    ui_cols_to_drop = []
    if not use_rdap_option:
        ui_cols_to_drop.extend(["RDAP(元データ)", "RDAP(日本語名)"])
        
    # IoTリスクが取得されていない（オフ または 集約モード）場合はカラムごと消す
    if all(r.get('IoT_Risk', '') in ['[Not Checked]', 'Aggr Mode (Skip)', '', 'N/A'] for r in results):
        ui_cols_to_drop.append("IoTリスク")
        
    if ui_cols_to_drop:
        df = df.drop(columns=[c for c in ui_cols_to_drop if c in df.columns], errors='ignore')

    # --- 2. マスタービュー (on_select有効化) ---
    st.markdown("#### 📊 一覧ビュー (行クリックで選択)")
    
    # カラム設定を動的に生成
    col_config = {
        "No.": st.column_config.NumberColumn(width="small"),
        "IPアドレス": st.column_config.TextColumn(width="medium"),
        "Whois(元データ)": st.column_config.TextColumn(width="medium"),
        "Whois(日本語名)": st.column_config.TextColumn(width="medium"),
    }
    if "RDAP(元データ)" not in ui_cols_to_drop:
        col_config["RDAP(元データ)"] = st.column_config.TextColumn(width="medium")
        col_config["RDAP(日本語名)"] = st.column_config.TextColumn(width="medium")
    if "IoTリスク" not in ui_cols_to_drop:
        col_config["IoTリスク"] = st.column_config.TextColumn(width="medium")

    # オリジナルカラムも設定に追加
    for col in original_cols:
        col_config[col] = st.column_config.TextColumn(width="medium")

    selection_state = st.dataframe(
        df,
        hide_index=True,
        width="stretch",
        height=450,
        on_select="rerun", 
        selection_mode="multi-row",
        column_config=col_config
    )

    st.markdown("---")

    # --- 3. ディテールビュー (一括表示対応) ---
    if "集約" in current_mode_full_text:
        st.info("💡 集約モードでは個別レポート出力はできません。")
        return

    st.markdown("#### 🔍 個別調査 ＆ レポート出力 (Detail View)")
    
    # 選択されたターゲットのリストを作成
    selected_indices = selection_state.selection.rows
    target_results_dict = {} # 重複排除のため辞書を使用 (Key: Target_IP)

    # A. 行が選択されている場合 (手動選択の取得)
    if selected_indices:
        for idx in selected_indices:
            if idx < len(results):
                res = results[idx]
                target_results_dict[res.get('Target_IP')] = res

    # B. フィルタリングUIを常に表示し、条件指定の取得を行う
    st.info("👆 一覧の行クリック選択と、以下の条件指定は同時に併用可能です。")
    with st.expander("🔎 条件でターゲットを一括指定する", expanded=True):
        col_f1, col_f2 = st.columns(2)
        with col_f1:
            all_countries = sorted(list(set([r.get('Country_JP', 'N/A') for r in results])))
            sel_countries = st.multiselect("国名で選択:", all_countries)
        with col_f2:
            all_isps = sorted(list(set([r.get('ISP_JP', 'N/A') for r in results])))
            sel_isps = st.multiselect("Whois(日本語名)で選択:", all_isps)
            
        # --- 元データの属性フィルタUI ---
        orig_filters = {}
        if original_cols and not df.empty:
            st.markdown("---")
            st.markdown("**📁 元データ (アップロードファイル) の属性で絞り込む**")
            
            filter_cols = st.columns(min(len(original_cols), 3) or 1)
            col_idx = 0
            
            for col_name in original_cols:
                with filter_cols[col_idx % 3]:
                    # UIフィルターは、結合された表示用dfではなく、元の生データから正確な選択肢を生成する
                    raw_series = st.session_state['original_df'][col_name] if st.session_state.get('original_df') is not None else df[col_name]
                    
                    # 列が日時として解釈できるか判定
                    is_datetime = False
                    if any(k in col_name.lower() for k in ['date', 'time', '日時', '時間', '時刻']):
                        try:
                            parsed_dates = pd.to_datetime(raw_series, errors='coerce').dropna()
                            if not parsed_dates.empty:
                                is_datetime = True
                                min_date = parsed_dates.min().date()
                                max_date = parsed_dates.max().date()
                        except:
                            pass
                    
                    if is_datetime:
                        date_range = st.date_input(f"📅 {col_name} (範囲)", value=[min_date, max_date], key=f"filter_{col_name}")
                        orig_filters[col_name] = {'type': 'date', 'value': date_range}
                    else:
                        # 結合文字列ではなく、個別のユニークな値を抽出
                        unique_vals = [str(v) for v in raw_series.dropna().unique() if str(v).strip() != '']
                        if 0 < len(unique_vals) <= 50:
                            selected_vals = st.multiselect(f"🏷️ {col_name}", sorted(unique_vals), key=f"filter_{col_name}")
                            orig_filters[col_name] = {'type': 'multiselect', 'value': selected_vals}
                        else:
                            text_search = st.text_input(f"🔍 {col_name} (部分一致)", key=f"filter_{col_name}")
                            orig_filters[col_name] = {'type': 'text', 'value': text_search}
                col_idx += 1
        
        # --- フィルタリング実行 ---
        # 何らかのフィルタ指定が存在するかチェック
        has_filter_input = bool(sel_countries or sel_isps)
        for f in orig_filters.values():
            if f['type'] == 'date' and len(f['value']) == 2:
                has_filter_input = True
            elif f['type'] in ('multiselect', 'text') and f['value']:
                has_filter_input = True

        if has_filter_input:
            for res in results:
                target_ip = res.get('Target_IP')
                actual_ip = extract_actual_ip(target_ip)
                
                c_match = res.get('Country_JP', 'N/A') in sel_countries if sel_countries else True
                i_match = res.get('ISP_JP', 'N/A') in sel_isps if sel_isps else True
                
                orig_match = True
                if orig_filters:
                    rows_list = orig_data_map.get(actual_ip, [])
                    if rows_list:
                        any_row_match = False
                        # IPに紐づく複数の履歴(行)のうち、いずれか1行でも全てのフィルタ条件を満たせば抽出対象とする
                        for row_data in rows_list:
                            row_match = True
                            for col_name, filter_info in orig_filters.items():
                                val = str(row_data.get(col_name, ''))
                                f_type = filter_info['type']
                                f_val = filter_info['value']
                                
                                if f_type == 'date' and len(f_val) == 2:
                                    if val:
                                        try:
                                            row_date = pd.to_datetime(val).date()
                                            if not (f_val[0] <= row_date <= f_val[1]):
                                                row_match = False
                                                break
                                        except:
                                            row_match = False
                                            break
                                    else:
                                        row_match = False
                                        break
                                elif f_type == 'multiselect' and f_val:
                                    if val not in f_val:
                                        row_match = False
                                        break
                                elif f_type == 'text' and f_val:
                                    if f_val.lower() not in val.lower():
                                        row_match = False
                                        break
                            
                            if row_match:
                                any_row_match = True
                                break
                                
                        if not any_row_match:
                            orig_match = False
                    else:
                        orig_match = False # 元データが存在しないIPはフィルタ除外
                
                if c_match and i_match and orig_match:
                    target_results_dict[target_ip] = res

    # 辞書から最終的なリストを生成 (重複は自動的に上書き・排除される)
    target_results = list(target_results_dict.values())
    
    if target_results:
        st.success(f"✅ 合計 **{len(target_results)}** 件が選択されています（手動選択: {len(selected_indices)}件 / フィルタ条件と結合済）。")

    # --- 4. 選択された全ターゲットに対してレポートを表示 ---
    if target_results:
        total_selected = len(target_results)

        st.markdown("##### ⚙️ レポート出力項目の選択")
        st.caption("※ APIキーが未入力の項目や、検索設定でオフになっていた機能はグレーアウト（無効化）されます。")
        
        # 選択されたターゲットにドメインが含まれているか、IPが含まれているかを判定
        has_domain_in_selection = any(not is_valid_ip(r.get('Target_IP', '')) or "(" in r.get('Target_IP', '') for r in target_results)
        has_ip_in_selection = any(is_valid_ip(extract_actual_ip(r.get('Target_IP', ''))) for r in target_results)

        # WHOISデータが実際に取得されているかを判定（複数入力時はIP-BAN回避のためスキップされている）
        has_whois_in_selection = False
        for r in target_results:
            target_ip = r.get('Target_IP', 'N/A')
            detailed = st.session_state.get('detailed_data', {}).get(target_ip, {})
            if detailed.get('IP_WHOIS_TEXT') or detailed.get('DOMAIN_WHOIS_TEXT'):
                has_whois_in_selection = True
                break

        col_opt1, col_opt2, col_opt3, col_opt4, col_opt5, col_opt6, col_opt7, col_opt8, col_opt9 = st.columns(9)
        # ドメイン関連項目は、選択肢にドメイン（または複合型）が含まれない場合はグレーアウト
        with col_opt1: opt_tld = st.checkbox("ドメイン情報", value=has_domain_in_selection, disabled=not has_domain_in_selection)
        with col_opt2: opt_dns = st.checkbox("正引き", value=has_domain_in_selection, disabled=not has_domain_in_selection)
        
        # IP関連項目（サブネット等）は、選択肢に純粋なIPが含まれない場合はグレーアウト
        with col_opt3: opt_subnet = st.checkbox("サブネット", value=has_ip_in_selection, disabled=not has_ip_in_selection)
        
        # RDAPとWHOISを独立して選択可能にする
        with col_opt4: opt_rdap = st.checkbox("RDAP", value=use_rdap_option, disabled=not use_rdap_option)
        
        # WHOISデータが存在する場合のみオン・操作可能にする
        with col_opt5: opt_whois = st.checkbox("WHOIS", value=has_whois_in_selection, disabled=not has_whois_in_selection)
        
        # 既存の制限にもターゲット属性の条件を追加
        with col_opt6: opt_ipinfo = st.checkbox("IPinfo", value=bool(pro_api_key) and has_ip_in_selection, disabled=not bool(pro_api_key) or not has_ip_in_selection)
        with col_opt7: opt_vpnapi = st.checkbox("VPNAPI.io", value=bool(vpnapi_key) and has_ip_in_selection, disabled=not bool(vpnapi_key) or not has_ip_in_selection)
        with col_opt8: opt_st = st.checkbox("SecTrails", value=bool(st_api_key) and has_domain_in_selection, disabled=not bool(st_api_key) or not has_domain_in_selection)
        with col_opt9: opt_rdns = st.checkbox("逆引き", value=use_rdns_option and has_ip_in_selection, disabled=not use_rdns_option or not has_ip_in_selection)
        
        current_report_opts = {
            "tld": opt_tld,
            "dns": opt_dns,
            "subnet": opt_subnet,
            "rdap": opt_rdap,
            "whois": opt_whois,
            "ipinfo": opt_ipinfo,
            "vpnapi": opt_vpnapi,
            "st": opt_st,
            "rdns": opt_rdns
        }

        # 📦 一括ダウンロードボタン (複数選択時のみ表示)
        if total_selected > 1:
            st.markdown("##### 📦 複数レポート一括ダウンロード")
            col_btn1, col_btn2 = st.columns(2)
            
            # 重い処理の前にスピナーを割り込ませ、フリーズではなく「処理中」であることを明示する
            with st.spinner(f"⏳ {total_selected} 件のレポートデータを構築中... (しばらくお待ちください)"):
                
                # ✅ 修正: メモリ節約のため分離されていた詳細データ(detailed_data)を統合した完全なリストを構築
                full_target_results = []
                for res in target_results:
                    clean_ip = get_copy_target(res.get('Target_IP', 'N/A'))
                    full_res = {**res}
                    if clean_ip in st.session_state.get('detailed_data', {}):
                        full_res.update(st.session_state['detailed_data'][clean_ip])
                    full_target_results.append(full_res)

                # 1. 統合レポート(HTML)の生成 (完全なデータリストを渡す)
                combined_html = generate_combined_html_report(full_target_results, current_report_opts)
                
                # 2. ZIPファイルの生成 (tempfileを利用してディスクに書き出し)
                with tempfile.NamedTemporaryFile(delete=False, suffix=".zip") as tmp:
                    tmp_zip_path = tmp.name
                    
                try:
                    with zipfile.ZipFile(tmp_zip_path, "w", zipfile.ZIP_DEFLATED, False) as zip_file:
                        valid_reports_count = 0
                        # 構築済みの full_target_results をループ処理する
                        for full_res in full_target_results:
                            target_ip = full_res.get('Target_IP', 'N/A')
                            clean_ip = get_copy_target(target_ip)
                            
                            html_report = generate_individual_html_report(full_res, clean_ip, current_report_opts)
                            if html_report:
                                safe_filename = re.sub(r'[\\/*?:"<>|]', "_", clean_ip)
                                zip_file.writestr(f"Report_{safe_filename}.html", html_report.encode('utf-8'))
                                valid_reports_count += 1
                
                    if valid_reports_count > 0:
                        current_time = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                        html_filename = f"Combined_Report_{current_time}.html"
                        zip_filename = f"Whois_Reports_Batch_{current_time}.zip"
                        
                        with col_btn1:
                            if combined_html: # ✅ None書き込みエラーを防ぐフェイルセーフ
                                if IS_PUBLIC_MODE:
                                    st.download_button(
                                        label=f"📜 {valid_reports_count} 件を1つの統合レポート(HTML)で保存",
                                        data=combined_html.encode('utf-8'),
                                        file_name=html_filename,
                                        mime="text/html",
                                        type="primary",
                                        width="stretch",
                                        help="選択した全件のレポートが1つのWebページに目次付きでまとまります。"
                                    )
                                else:
                                    render_local_save_ui(
                                        f"💾 {valid_reports_count} 件の統合レポートをローカル保存", 
                                        html_filename, combined_html, "batch_html", "primary"
                                    )
                            else:
                                st.error("統合レポートのデータ生成に失敗しました。")
                                    
                        with col_btn2:
                            with open(tmp_zip_path, "rb") as f:
                                zip_data = f.read()
                            
                            if IS_PUBLIC_MODE:
                                st.download_button(
                                    label=f"🗜️ {valid_reports_count} 件の個別レポートをZIPで保存",
                                    data=zip_data,
                                    file_name=zip_filename,
                                    mime="application/zip",
                                    type="secondary",
                                    width="stretch",
                                    help="各IPごとに独立したHTMLファイルを作成し、ZIPに圧縮してダウンロードします。"
                                )
                            else:
                                render_local_save_ui(
                                    f"💾 {valid_reports_count} 件のZIPをローカル保存 (無制限)", 
                                    zip_filename, zip_data, "batch_zip", "secondary"
                                )
                finally:
                    if os.path.exists(tmp_zip_path):
                        os.remove(tmp_zip_path) # 送信後、またはエラー発生時に確実に削除する
            st.divider()
        
        # Rendering Overload（UI崩壊）を防ぐためのハードリミット設定
        DISPLAY_LIMIT = 50 
        if total_selected > DISPLAY_LIMIT:
            st.warning(f"⚠️ **ブラウザ保護制限**: 選択された件数（{total_selected}件）が上限を超えています。UIのフリーズを防ぐため、画面上での個別プレビューは最初の {DISPLAY_LIMIT} 件のみ表示しています。")
            display_targets = target_results[:DISPLAY_LIMIT]
        else:
            display_targets = target_results

        for i, res in enumerate(display_targets):
            target_ip = res.get('Target_IP', 'N/A')
            clean_ip = get_copy_target(target_ip)
            
            with st.container():
                # チェックボックス用のユニークなキーを生成
                done_key = f"done_target_{clean_ip}_{i}"
                
                # タイトルとチェックボックスを横に並べる
                col_title, col_chk = st.columns([4, 1])
                with col_chk:
                    is_done = st.checkbox("✅ 調査完了", key=done_key)

                # チェック状態に応じて表示UIを動的に切り替える
                if is_done:
                    # 完了時：グレーアウト ＆ 取り消し線
                    with col_title:
                        st.markdown(f"<h5 style='color: #9e9e9e; text-decoration: line-through;'>🎯 [{i+1}/{total_selected}] Target: {target_ip}</h5>", unsafe_allow_html=True)
                    
                    # 詳細情報を Expander に格納し、デフォルトで閉じておく（画面のスペースを空ける）
                    container_context = st.expander("📁 完了済みの詳細データを再確認する", expanded=False)
                else:
                    # 未完了時：通常表示
                    with col_title:
                        st.markdown(f"##### 🎯 [{i+1}/{total_selected}] Target: `{target_ip}`")
                    
                    # 詳細情報を Container に格納し、そのまま展開して表示する
                    container_context = st.container()

                # 詳細情報の描画（完了・未完了問わず中身は同じ）
                with container_context:
                    c1, c2 = st.columns([2, 1])
                    with c1:
                        # リンク集
                        st.markdown(f"**🛡️ 外部調査リンク:**")
                        st.markdown(f"{res.get('Secondary_Security_Links', '-')}")
                        
                        # RIRリンク
                        st.markdown(f"**📚 RIR / Whois 窓口:** {res.get('RIR_Link', '-')}")
                        
                        # コピー枠の横幅を絞り、マウス移動の負担を極小化する
                        code_col, _ = st.columns([1, 2])
                        with code_col:
                            st.code(clean_ip, language=None)
                        
                        # 補足情報
                        st.caption(f"ISP: {res.get('ISP_JP', '-')} / RDAP: {res.get('RDAP_JP', '-')}")    
                    with c2:
                        # HTMLレポート生成
                        full_res = {**res}
                        if clean_ip in st.session_state.get('detailed_data', {}):
                            full_res.update(st.session_state['detailed_data'][clean_ip])
                            
                        html_report = generate_individual_html_report(full_res, clean_ip, current_report_opts)
                        if html_report:
                            st.download_button(
                                label=f"⬇️ レポートDL ({clean_ip})",
                                data=html_report,
                                file_name=f"Report_{clean_ip}.html",
                                mime="text/html",
                                key=f"dl_btn_multi_{clean_ip}_{i}", # if分岐で片方しか実行されないため同じキーでOK
                                width="stretch"
                            )
                        else:
                            st.button("データなし", disabled=True, key=f"no_dl_{i}")
                
                st.divider()
    else:
        st.caption("詳細を表示するには、一覧の行をクリックするか、条件を指定してください。")

# --- リンク分析エンジン ---
def render_spider_web_analysis(df):
    """
    ノードベースの相関グラフ表示機能。Graphvizを使用して描画する。
    """
    st.info("IPアドレス、ISP、国、およびリスクの繋がりを視覚化します。共通のISPやリスクを持つIPが中心に集まり、攻撃インフラの『ハブ』を特定できます。")

    if df.empty:
        st.warning("データがありません。")
        return

    # GraphvizのDOT言語でグラフ構造を定義
    dot_lines = [
        'graph {',
        '  layout=neato;', # ノードを物理的な反発力で自動配置するエンジン
        '  overlap=false;',
        '  splines=true;',
        '  node [fontname="Helvetica", fontsize=10];'
    ]
    
    nodes = set()
    edges = set()
    
    # 描画負荷を考慮し、上位50件程度でプロット
    plot_df = df.head(50).fillna("N/A")

    for _, row in plot_df.iterrows():
        # カラム名の取得を日本語版に変更
        ip = row.get('IPアドレス', row.get('Target_IP', 'Unknown'))
        
        # 優先度: 日本語カラム > 英語カラム > N/A
        isp = row.get('Whois結果（日本語名称）', row.get('ISP_JP', row.get('ISP', 'N/A')))
        country = row.get('国名', row.get('Country_JP', row.get('Country', 'N/A')))
        risk = row.get('IoTリスク', row.get('IoT_Risk', ''))
        proxy = row.get('プロキシ種別', row.get('Proxy Type', ''))

        # 1. IPノード (水色の丸)
        nodes.add(f'"{ip}" [shape=circle, style=filled, fillcolor="#E0F2F1", width=0.8];')

        # 2. ISPノード (オレンジの四角) - IPと線を結ぶ
        if isp != "N/A":
            nodes.add(f'"{isp}" [shape=box, style=filled, fillcolor="#FFF3E0", color="#FF9800", penwidth=2];')
            edges.add(f'"{ip}" -- "{isp}" [color="#FF9800", alpha=0.5];')

        # 3. 国ノード (緑の楕円)
        if country != "N/A":
            nodes.add(f'"{country}" [shape=ellipse, style=filled, fillcolor="#F1F8E9", color="#8BC34A"];')
            edges.add(f'"{ip}" -- "{country}" [style=dotted, color="#8BC34A"];')

        # 4. リスクノード (赤の二重丸) - 複数リスクは分割して線を結ぶ
        if risk and risk not in ["[No Match]", "[Not Checked]", "[No Data]", "N/A", ""]:
            for r in risk.split(" / "):
                nodes.add(f'"{r}" [shape=doublecircle, style=filled, fillcolor="#FFEBEE", color="#F44336", fontcolor="#B71C1C", penwidth=3];')
                edges.add(f'"{ip}" -- "{r}" [color="#F44336", penwidth=2];')

        # 5. プロキシノード (紫の六角形)
        if proxy and proxy != "Standard Connection":
            nodes.add(f'"{proxy}" [shape=hexagon, style=filled, fillcolor="#F3E5F5", color="#9C27B0"];')
            edges.add(f'"{ip}" -- "{proxy}" [color="#9C27B0"];')

    dot_lines.extend(list(nodes))
    dot_lines.extend(list(edges))
    dot_lines.append('}')
    
    dot_string = "\n".join(dot_lines)
    
    # Streamlit標準のGraphviz描画機能を使用
    st.graphviz_chart(dot_string)
    
    with st.expander("💡 読み解きのヒント"):
        st.write("""
        - **大きな塊（ハブ）**: 複数のIPから線が集まっているノード（ISPやリスク）は、今回の調査対象に共通するインフラです。
        - **赤い二重丸**: 危険なポートが露出している共通のリスク要因です。攻撃者の踏み台リストの可能性があります。
        - **独立したノード**: 他と繋がりのないIPは、今回のグループとは別の背景を持つ可能性があります。
        """)

# 📊 元データ結合分析機能 (タブ化対応)
def render_merged_analysis(df_merged):
    st.markdown("### 📈 分析センター")
    
    tab_cross, tab_spider = st.tabs(["📊 クロス分析 (マクロ視点)", "🕸️ リンク分析 (ミクロ視点)"])
    
    with tab_cross:
        st.info("アップロードされたファイルの元の列と、検索で得られたWhois情報を組み合わせて可視化します。")
        # 除外するカラムを日本語名に変更
        exclude_cols = ['Whois結果（元データ）', 'Whois結果（日本語名称）', '国名（英語）', '国名', 'プロキシ種別', 'ステータス', 'IoTリスク', 'RDAP結果（元データ）', 'RDAP結果（日本語名称）', 'ISP', 'ISP_JP', 'Country', 'Country_JP']
        original_cols = [c for c in df_merged.columns if c not in exclude_cols]
        
        # 分析に使用するWhois系カラムを日本語名に変更
        whois_cols = ['国名', 'Whois結果（日本語名称）', 'プロキシ種別', 'IoTリスク', 'ステータス']
        
        col_x, col_grp, col_chart_type = st.columns(3)
        with col_x:
            x_col = st.selectbox("X軸 (カテゴリ/元の列)", original_cols + whois_cols, index=0)
        with col_grp:
            group_col = st.selectbox("積み上げ/色分け (Whois情報など)", ['(なし)'] + whois_cols + original_cols, index=1)
        with col_chart_type:
            chart_type = st.radio("グラフタイプ", ["バーチャート (集計)", "ヒートマップ"], horizontal=True)

        if not df_merged.empty:
            chart = None
            chart_df = df_merged.fillna("N/A").astype(str)
            
            # 5000行以上の場合はサンプリングしてブラウザのクラッシュを防ぐ
            if len(chart_df) > 5000:
                st.warning(f"⚠️ **データ量警告**: データが {len(chart_df)} 件あります。ブラウザのクラッシュを防ぐため、ランダムに抽出した 5000 件のデータでグラフを描画しています。")
                chart_df = chart_df.sample(n=5000, random_state=42)

            if chart_type == "バーチャート (集計)":
                if group_col != '(なし)':
                    chart = alt.Chart(chart_df).mark_bar().encode(
                        x=alt.X(x_col, title=x_col),
                        y=alt.Y('count()', title='件数'),
                        color=alt.Color(group_col, title=group_col),
                        tooltip=[x_col, group_col, 'count()']
                    ).properties(height=400)
                else:
                    chart = alt.Chart(chart_df).mark_bar().encode(
                        x=alt.X(x_col, title=x_col),
                        y=alt.Y('count()', title='件数'),
                        tooltip=[x_col, 'count()']
                    ).properties(height=400)
            elif chart_type == "ヒートマップ":
                 if group_col != '(なし)':
                    chart = alt.Chart(chart_df).mark_rect().encode(
                        x=alt.X(x_col, title=x_col),
                        y=alt.Y(group_col, title=group_col),
                        color=alt.Color('count()', title='件数', scale=alt.Scale(scheme='viridis')),
                        tooltip=[x_col, group_col, 'count()']
                    ).properties(height=400)
                 else:
                     st.warning("ヒートマップには「積み上げ/色分け」項目の選択が必要です。")

            if chart:
                st.altair_chart(chart, width="stretch")
                chart_json = chart.to_dict()
                html_content = generate_cross_analysis_html(chart_json, x_col, group_col if group_col != '(なし)' else 'Count')
                base_fname = st.session_state.get('base_filename', 'WhoisSearchResult')
                ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
                st.download_button(
                    label="⬇️ クロス分析レポート(HTML)をダウンロード",
                    data=html_content,
                    file_name=f"{base_fname}_CrossAnalysis_{x_col}_vs_{group_col}_{ts}.html",
                    mime="text/html"
                )

    with tab_spider:
        # リンク分析関数を呼び出す
        render_spider_web_analysis(df_merged)

# ==========================================
# 状態管理（Session State）用ヘルパー関数
# ==========================================
def init_session_state():
    """ アプリケーション起動時・リセット時に必要なSession Stateを初期化する """
    default_states = {
        'cancel_search': False,
        'raw_results': [],
        'targets_cache': [],
        'is_searching': False,
        'deferred_ips': {},
        'finished_ips': set(),
        'search_start_time': 0.0,
        'target_freq_map': {},
        'cidr_cache': {},
        'debug_summary': {},
        'detailed_data': {},
        'learned_proxy_isps': {}
    }
    
    for key, default_value in default_states.items():
        if key not in st.session_state:
            st.session_state[key] = default_value

def reset_search_state():
    """ 新規検索を開始する際に、前回の巨大なデータを明示的にメモリから解放する """
    # 巨大なリストや辞書を削除してガベージコレクションを促す
    if 'detailed_data' in st.session_state:
        st.session_state['detailed_data'].clear()
    if 'raw_results' in st.session_state:
        st.session_state['raw_results'].clear()
        
    st.session_state.is_searching = True
    st.session_state.cancel_search = False
    st.session_state.deferred_ips = {}
    st.session_state.finished_ips = set()
    st.session_state.search_start_time = time.time()
    clear_recovery_data()


# --- メイン処理 ---
def main():
    # 状態管理の初期化関数を呼び出し
    init_session_state()

    # リカバリUI
    if not IS_PUBLIC_MODE and os.path.exists(BACKUP_FILE) and not st.session_state.is_searching and not st.session_state.raw_results:
        st.warning("⚠️ 前回中断された検索セッションが残っています。")
        col_rec1, col_rec2 = st.columns(2)
        with col_rec1:
            if st.button("🔄 検索を途中から再開する", type="primary"):
                if load_recovery_data():
                    st.rerun()
        with col_rec2:
            if st.button("🗑️ バックアップを破棄する"):
                clear_recovery_data()
                st.rerun()

    tor_nodes = fetch_tor_exit_nodes()
    disposable_domains = fetch_disposable_domains()
    cloud_ip_data = fetch_cloud_ip_ranges()

    # 外部データベース取得エラー時の警告表示
    if not tor_nodes or not disposable_domains:
        st.warning("⚠️ **外部データベース取得エラー**: ネットワークの切断等により、Torノードリストまたは使い捨てメアドリストの取得に失敗しました。該当する検知機能が一時的に停止しています。")
    
    with st.sidebar:
        st.markdown("### 🛠️ Menu")
        selected_menu = option_menu(
            menu_title=None,
            options=["Whois検索", "仕様・解説"],
            icons=["search", "book"],
            default_index=0,
            styles={
                "nav-link": {"font-size": "16px", "text-align": "left", "margin": "5px", "--hover-color": "#eee"},
                "nav-link-selected": {"background-color": "#1e3a8a"},
            }
        )
        st.markdown("---")
        
        # --- Local Mode専用: 出力先設定 ---
        if not IS_PUBLIC_MODE:
            st.markdown("##### 📂 ローカル保存先設定")
            default_export_dir = os.path.join(os.getcwd(), "exports")
            st.session_state['local_export_dir'] = st.text_input("保存先フォルダの絶対パス", value=default_export_dir, help="ファイルの直接保存先を指定します。存在しない場合は自動作成されます。")
            st.markdown("---")
        
        # Proモード設定 (APIキー入力)
        with st.expander("🔑 APIキー設定 (Pro Mode)", expanded=False):
            st.caption("高精度な分析を行うためのAPIキーを設定します。")
        
            # 1. IPinfo (Pro Mode) の設定
            pro_api_key = ""
            if HARDCODED_IPINFO_KEY:
                use_hc_ipinfo = st.checkbox("埋め込みキー (IPinfo) を使用", value=True, help="オフにすると、埋め込まれたAPIキーを無効化し、空欄または手動入力モードに切り替えます。")
                if use_hc_ipinfo:
                    pro_api_key = HARDCODED_IPINFO_KEY
                    st.success(f"✅ IPinfo Key Loaded: {pro_api_key[:4]}***")
                else:
                    pro_api_key = st.text_input("ipinfo.io API Key", type="password", key="input_ipinfo", help="入力するとipinfo.ioの高精度データベースを使用します。空欄の場合はip-api.comを使用します。").strip()
            else:
                pro_api_key = st.text_input("ipinfo.io API Key", type="password", key="input_ipinfo", help="入力するとipinfo.ioの高精度データベースを使用します。空欄の場合はip-api.comを使用します。").strip()

            # 2. VPNAPI.io の設定
            vpnapi_key = ""
            if HARDCODED_VPNAPI_KEY:
                use_hc_vpnapi = st.checkbox("埋め込みキー (VPNAPI.io) を使用", value=True, help="オフにすると、埋め込まれたAPIキーを無効化し、空欄または手動入力モードに切り替えます。")
                if use_hc_vpnapi:
                    vpnapi_key = HARDCODED_VPNAPI_KEY
                    st.success(f"✅ VPNAPI Key Loaded: {vpnapi_key[:4]}***")
                else:
                    vpnapi_key = st.text_input("VPNAPI.io API Key", type="password", key="input_vpnapi", help="VPNAPI.ioのAPIキーを入力することで、IPアドレスの匿名通信判定を取得します。").strip()    
            else:
                vpnapi_key = st.text_input("VPNAPI.io API Key", type="password", key="input_vpnapi", help="VPNAPI.ioのAPIキーを入力することで、IPアドレスの匿名通信判定を取得します。").strip()

            # 3. SecurityTrails の設定
            st_api_key = ""
            if HARDCODED_SECURITYTRAILS_KEY:
                use_hc_st = st.checkbox("埋め込みキー (SecurityTrails) を使用", value=True, help="オフにすると、埋め込まれたAPIキーを無効化し、空欄または手動入力モードに切り替えます。")
                if use_hc_st:
                    st_api_key = HARDCODED_SECURITYTRAILS_KEY
                    st.success(f"✅ SecurityTrails Key Loaded: {st_api_key[:4]}***")
                else:
                    st_api_key = st.text_input("SecurityTrails API Key", type="password", key="input_st", help="FQDN（ドメイン）が入力された際、過去のA/AAAAレコードの履歴を取得するために使用します。").strip()
            else:
                st_api_key = st.text_input("SecurityTrails API Key", type="password", key="input_st", help="FQDN（ドメイン）が入力された際、過去のA/AAAAレコードの履歴を取得するために使用します。").strip()

            st_start_date = None
            st_end_date = None
            if st_api_key:
                st.markdown("##### 📅 履歴取得期間 (SecurityTrails)")
                use_st_date_filter = st.checkbox("期間を指定して全件抽出する", value=False, help="チェックを入れると指定期間の履歴を制限なく抽出します。チェックがない場合は最新20件のみを取得します。")
            
                if use_st_date_filter:
                    # 本日の日付と、約3ヶ月前（90日前）の日付を動的に計算
                    today_date = datetime.date.today()
                    three_months_ago = today_date - datetime.timedelta(days=90)
                    
                    col_dt1, col_dt2 = st.columns(2)
                    with col_dt1:
                        st_start_date = st.date_input("開始日", three_months_ago, help="この日以降に観測された履歴のみを抽出します。期間が長い場合はレポート生成がスキップされる可能性があります。")
                    with col_dt2:
                        st_end_date = st.date_input("終了日", today_date, help="この日以前に観測された履歴のみを抽出します。")
                
                # Reverse IPの設定
                st.markdown("##### ⚙️ Reverse IP 追加設定")
                use_st_rev_fetchall = st.checkbox("Reverse IP 全件取得 (API消費大)", value=False, help="オンにすると、同一IPに紐づくドメインが100件を超える場合、APIを複数回消費して全件取得を試みます。CDNのIPなどを対象にするとクレジットが枯渇する恐れがあります。")
            else:
                use_st_rev_fetchall = False

        st.markdown("---")
        if st.button("🔄 システム/キャッシュを完全リセット", help="キャッシュが古くなった場合やメモリを解放したい場合にクリック"):
            # セッションステートを完全に削除してガベージコレクションを促す
            keys_to_delete = ['cidr_cache', 'detailed_data', 'raw_results', 'resolved_dns_map', 'original_df', 'original_input_list', 'targets_cache']
            for key in keys_to_delete:
                if key in st.session_state:
                    del st.session_state[key]
            
            st.cache_data.clear()
            st.cache_resource.clear()
            init_session_state() # 必要なキーを再構築
            
            st.info("IP/CIDRキャッシュ、検索履歴、およびメモリを完全にクリアしました。")
            time.sleep(1)
            st.rerun()

    if selected_menu == "仕様・解説":
        st.title("📖 マニュアル & ガイド")
        
        # タブで情報を整理して見やすくする
        tab1, tab2, tab3 = st.tabs(["🔰 使い方・モード選択", "⚙️ 仕様・技術詳細", "❓ FAQ"])

        with tab1:
            st.markdown("### 🚀 クイックスタート")
            
            if IS_PUBLIC_MODE:
                st.markdown("""
                1. **入力**: 左側の**テキストエリアにIPアドレスを貼り付ける**か、`.txt` ファイルをアップロードします。
                   > ⚠️ **注意**: 公開サーバー環境のため、Excel/CSVファイルのアップロードは制限されています。
                """)
            else:
                st.markdown("""
                1. **入力**: 左側のテキストエリアに貼り付けるか、**テキスト、CSV、Excelファイル**をアップロードします。
                   > ✅ **Local Mode**: ローカル環境で動作しているため、機密情報を含むファイルの処理も可能です。
                """)

            st.markdown("""
            2. **設定**: 基本的にはそのままでOKです。大量のデータを処理する場合や、より詳細な情報が必要な場合は、下部の設定を変更してください。
            3. **実行**: 「🚀 検索開始」ボタンを押します。
            """)
            
            st.info("💡 **ヒント**: 結果が出たあと、画面下のボタンからExcelファイルをダウンロードすると、自動でグラフ化された分析レポートが見れます。")

            st.markdown("---")
            st.markdown("### ⚙️ 設定項目の解説")
            
            st.markdown("#### 1. 表示モード (Display Mode)")
            st.markdown("検索結果をどのようにリストアップするかを選択します。")
            
            display_mode_df = pd.DataFrame({
                "モード名": ["標準モード", "集約モード", "簡易モード"],
                "API通信": ["あり (消費)", "あり (消費)", "なし (節約)"],
                "説明とメリット": [
                    "入力されたIPを1行ずつ表示します。個別の判定結果を詳しく確認したい場合に最適です。",
                    "同じISP・国で、連続するIPアドレスを1行にまとめます。（例: `1.1.1.1 - 1.1.1.5 (x5)`）。大量のログから「どこの会社からのアクセスが多いか」を概観するのに便利です。",
                    "API通信を行わず、調査用リンクの生成のみ行います。API制限にかかった場合や、外部へIPを送信したくない場合に利用します。"
                ]
            })
            st.table(display_mode_df.set_index("モード名"))

            st.markdown("#### 2. API処理モード (Processing Speed)")
            st.markdown("検索スピードと安定性のバランスを調整します。")
            
            api_mode_df = pd.DataFrame({
                "モード名": ["安定性重視", "速度優先"],
                "動作イメージ": ["🐢 ゆっくり・確実", "🚀 素早く・並列"],
                "説明": [
                    "待機時間を長め(2.5秒)に取り、1件ずつ処理します。APIのレートリミット（制限）にかかりにくく、エラーが出にくい安全運転設定です。",
                    "待機時間を短く(1.4秒)し、2つの処理を同時に走らせます。大量のリストを早く処理したい場合に推奨されますが、回線状況によっては制限にかかりやすくなります。"
                ]
            })
            st.table(api_mode_df.set_index("モード名"))

            st.markdown("#### 3. 詳細オプション")
            st.markdown("""
            - **🔍 公式レジストリ情報 (RDAP)**
                - `ip-api.com` (通常版） の情報に加え、各地域の**公式レジストリ(RDAP)** にも問い合わせを行います。
                - **メリット**: 「運用者(ISP)」だけでなく「法的な保有組織(Org)」まで特定できる確率が上がります。
            
            - **🔑 高精度判定 (ipinfo Key)**
                - **メリット**: VPN/Proxy/Hostingの判定精度が劇的に向上し、企業名の特定精度も高まります。
                - **注意**: データプランの種類（無料プラン、有料プラン）やAPIの利用状況に応じて、提供される情報の項目が異なり、無料版は、地理的位置情報やISP情報などの基本的なデータに限定されます。
                        
            - **🕵️ 匿名通信判定 (VPNAPI.io Key)**
                - **メリット**: VPN、Proxy、Tor等の利用が疑われる不審なIPに対し、VPNAPI.ioの専門データベースから「匿名通信該当結果」を自動取得します。

            - **📜 過去のDNS履歴取得 (SecurityTrails Key)**
                - **メリット**: ドメイン（FQDN）を入力した際、WAF（Cloudflare等）で秘匿される前の過去の生IP（オリジンサーバー）や、紐づいていたIPアドレスの変遷を取得できます。
                - **注意**: 月間50回までの無料枠が存在します。IPアドレス単体の検索では消費されません。
                        
            - **🔄 IP逆引き (Reverse DNS)**
                - **メリット**: IPアドレスに紐づくホスト名（PTRレコード）を取得します。プロバイダの特定や、サーバー用途の推測に役立ちます。
                - **動作仕様**: 精度と網羅性を優先するため、本オプション有効時は自動的に「シングルスレッド・待機延長モード」へ切り替わります。

            - **🔎 IoT Risk Check (InternetDB)**
                - **メリット**: ポート5555(ADB/FireStick)や1080(Proxy)等の露出を検知し、踏み台リスクを警告します（APIキー不要）。
            """)

            st.markdown("---")
            st.markdown("### 💻 動作モードとローカル版の導入")
            
            st.info("""
            このアプリは、実行環境（クラウドかローカルか）によって機能とセキュリティポリシーが変化します。
            機密性の高いデータ（顧客ログ等）を扱う場合や、大量のCSV/Excelを処理したい場合は、**Local版** の利用を強く推奨します。
            """)

            # モード比較表
            mode_compare_df = pd.DataFrame({
                "機能 / 特徴": ["Excel/CSV アップロード", "機密情報の取扱", "実行環境", "主な用途"],
                "☁️ Public Cloud版": ["❌ 不可 (.txtのみ)", "△ 推奨しない (共有サーバー)", "Streamlit Community Cloud", "手軽な単発検索・デモ利用"],
                "🏠 Local Private版": ["✅ 可能 ", "◎ 安全 (自PC内で完結)", "ローカルPC / 社内サーバー", "実務・ログ解析・大量処理"]
            })
            st.table(mode_compare_df.set_index("機能 / 特徴"))

            st.markdown("#### 📥 ローカル版 (Local Private Edition) の導入方法")
            st.markdown("Python環境があれば、どなたでも制限なしのローカル版を使用できます。ソースコードはGitHubで公開されています。")
            
            st.markdown("""
            **1. ソースコードの取得**
            以下のリポジトリからコードをダウンロード（Clone）してください。
            - 🔗 **GitHub Repository**: [github.com/x04z/WhoisApp](https://github.com/x04z/WhoisApp)
            
            **2. 必要なライブラリのインストール**
            ```bash
            pip install streamlit pandas requests streamlit-option-menu altair openpyxl dnspython
            ```
            
            **3. アプリの起動**
            コマンドプロンプトまたはターミナルで以下を実行します。
            ```bash
            streamlit run WhoisAppxxxx.py
            ```
            """)
 
        with tab2:
            st.markdown("""
            #### 1. データソース
            - **IP Geolocation / ISP 情報**: 
                - 通常版: `ip-api.com` (毎分45リクエスト制限)
                - 高精度版: `ipinfo.io` (APIキーに基づく制限)
            - **匿名通信判定 (Proxy/VPN)**: `VPNAPI.io` 
            - **過去のDNS履歴 (Historical DNS)**: `SecurityTrails` (ドメイン入力時のみ実行)
            - **DNS解析 (Forward/Reverse)**: OS標準 `dnspython` (正引き) / `dnspython` ライブラリ (逆引き)
            - **Whois (RDAP)**: APNIC等の各地域レジストリ公式サーバー
            - **IoT Risk Intelligence**: Shodan InternetDB (ポートスキャン履歴/キャッシュ)
            - **Tor出口ノード**: Tor Project公式サイト

            #### 2. 多角的解析の仕組み (API・RDAP・ProxyEvidence・DNS History)
            - **運用者判定 (ip-api/ipinfo)**: 
                - **役割**: 「今、誰がそのIPを運用しているか？」(Service Provider) を答えます。
                - **特徴**: 高速。ISPやクラウド事業者名（Cloudflare, Amazon等）を特定します。
            - **法的保有者判定 (RDAP公式台帳)**: 
                - **役割**: 「そのIPアドレス(土地)の法的な持ち主は誰か？」(Registry Owner) を答えます。
                - **特徴**: 厳密。各地域のレジストリに登録された組織名を特定します。
            - **インフラ紐付け (rDNS/PTR)**: 
                - **役割**: 「そのIPにはどんなホスト名が付いているか？」を特定。Windowsの `nslookup` の制限を回避するため、専用のリゾルバを用いて全レコードを抽出します。
            - **匿名性判定 (VPNAPI.io)**: 
                - **役割**: 「そのIPは意図的に隠蔽（VPN/Proxy等）されているか？」を答えます。
                - **特徴**: 不審な判定時に専門DBから詳細なJSONを取得します。
            - **レコード履歴特定 (SecurityTrails)**: 
                - **役割**: 「そのドメインはどのIPアドレスがレコードに設定されていたか？」を答えます。
                - **特徴**: 履歴追跡。WAF等で現在のIPが隠蔽されていても、過去のIPアドレスを特定できる可能性があります。
            - **メリット**: これらを統合することで、単なる「場所の特定」を超え、「通信主体の隠蔽意図」や「インフラの変遷」までを浮き彫りにします。

            #### 3. 技術的仕様
            - **並列処理**: マルチスレッドによる高速検索
            - **動的負荷調整**: 逆引きオプション有効時は、クエリの衝突とタイムアウトを回避するため、**自動でシングルスレッド・待機延長モード**へ移行し、調査の確実性を担保します。
            - **CIDRキャッシュ**: 同一ネットワーク帯域への重複リクエスト回避
            """)
            st.markdown("#### 4. 判定ステータスの意味")
            
            st.error("🧅 **Tor Node**")
            st.markdown("Tor（The Onion Router）匿名化ネットワークの出口ノードです。発信元の完全な隠蔽を目的としており、攻撃の前兆や違法取引に関連する通信である可能性が高いです。")

            st.error("⚠️ **IoT露出 / 高リスクポート検知**")
            st.markdown("""
            Shodan InternetDBにより、以下の危険なポート開放が確認されたIPです。
            
            - **Telnet (23)**: 暗号化されていない古いプロトコル。**「開いているだけで高リスク」**とみなされます。
            - **ADB (5555/5554)**: Android端末（FireTVなど）のデバッグ機能が認証なしで公開されています。
            - **TR-069 (7547)**: ルーター管理用プロトコル。脆弱性がある場合、ルーターごと乗っ取られる恐れがあります。
            - **Proxy (1080/3128)**: 踏み台として悪用されるプロキシサーバー（SOCKS/Squid）が稼働しています。
            - **UPnP (1900)**: ネットワーク内の機器探索用プロトコルが外部に漏れています。
            """)

            st.warning("🍏 **iCloud Private Relay**")
            st.markdown("Appleデバイス（iPhone/Mac）のプライバシー保護機能による通信です。IPアドレスはAppleの提携パートナー（Cloudflare/Akamai等）のものに置き換わっており、真の発信元は隠蔽されていますが、基本的には一般ユーザーによるアクセスです。")

            st.warning("☁️ **Hosting/VPN/Proxy**")
            st.markdown("データセンター、商用VPN、プロキシ経由の通信です。一般家庭からのアクセスではなく、ボットや匿名化ツールを使用している可能性があります。")
            

        with tab3:
            # --- モード別案内: FAQ ---
            if IS_PUBLIC_MODE:
                st.markdown("""
                **Q. ファイルをアップロードしても大丈夫ですか？**\n
                A. 現在は **Public (Cloud) Mode** で動作しています。サーバーは共有環境のため、**機密情報を含むファイルのアップロードは推奨されません**。テキストエリアへのIP貼り付けを利用するか、個人情報を含まないデータのみを使用してください。
                """)
            else:
                st.markdown("""
                **Q. ファイルをアップロードしても大丈夫ですか？**\n
                A. はい。現在は **Local Mode** で動作しています。データはあなたのPC（またはプライベートサーバー）内で処理され、外部の開発者等に送信されることはありません。安心して機密データを取り扱えます。
                """)

            st.markdown("""
            **Q. 検索が途中で止まりました。**\n
            A. APIの制限（レートリミット）にかかった可能性があります。ツールは自動的に待機して再開しますが、大量（数千件）の検索を行う場合は時間がかかります。「待機中」の表示が出ている場合はそのままお待ちください。なお、通常版API（ip-api）は流量制限が厳しく、数十件程度のバーストで保留（Deferred）状態になることがあります。スムーズな解析が必要な場合は「Local版」の利用、または「Pro Mode (IPinfo)」の適用を検討してください。\n
                        
            **Q. 各種APIキーはどこで手に入りますか？**\n
            A. 本ツールで利用可能な高度判定用APIキーは、以下の公式サイトから無料で登録・取得できます（いずれも無料枠が存在します）。
            * **高精度判定 (ipinfo)**: [ipinfo.io サインアップ](https://ipinfo.io/signup)
            * **匿名通信判定 (VPNAPI.io)**: [VPNAPI.io サインアップ](https://vpnapi.io/signup)
            * **過去のDNS履歴取得 (SecurityTrails)**: [SecurityTrails サインアップ](https://securitytrails.com/app/signup)

            **Q. ISP名と [RDAP: 〇〇] の名前が違うのですが？**\n
            A. **それは「運用者」と「持ち主」の違いです。** 例えば `1.1.1.1` というIPアドレスの場合：
            * **ISP (API)**: `Cloudflare, Inc.` (DNSサービスを提供している運用者)
            * **RDAP (台帳)**: `APNIC-LABS` (IPアドレスブロックを保有している研究組織)
            このように表示されるのはバグではなく、公式レジストリ情報 (RDAP)が、**IPアドレスの「表の運用者」と「裏の保有者」の両方を正しく表している証拠**です。
            
            **Q. ISP名とRDAPの名前が異なる場合、発信者情報開示をどちらに請求すればいいでしょうか？**\n
            A. 個人（契約者）の情報を持っているのは**表の運用者である「ISP / プロバイダ」**の方です。RDAPの情報はあくまで「そのIPアドレスブロックを管理している組織」の情報であり、実際の利用者情報は持っていないことが多いです。発信者情報開示請求を行う場合は、**ISP名を使って手続きを行ってください**。

            **Q. IoT Risk判定が出ましたが、これは確定ですか？**\n
            A. いいえ。まず、本機能はリアルタイムのスキャンではなく、**「Shodanが過去に実施したポートスキャン結果（履歴）」**を参照しています。そのため、現在すでにポートが塞がれている可能性（または新たに開いている可能性）が常に存在します。また、一般回線の場合、そのIPを共有している**多数人の中の1人**が脆弱性を露出させているだけで、無関係な利用者の通信も同じIPとして判定されます。絶対的な証拠ではなく、あくまで「過去にリスクが確認されたノードである」という調査優先度の指標として扱ってください。
            
            **Q. 検知されるポートのリスク詳細を教えてください**\n
            A. 本ツールでは、以下のポート開放状況を監視しています。
                        
            * **⚠️ 23 (Telnet) / 21 (FTP)**
                * **判定**: **極めて危険な古いプロトコル** です。通信が暗号化されないため、パスワード等が盗聴されるリスクがあります。現代のインターネットで意図的に公開する正当な理由はほぼありません。
            
            * **🔥 1080 (SOCKS) / 3128 (Squid) / 8080 (HTTP)**
                * **判定**: **プロキシ (Proxy)** として悪用される典型的なポートです。一般家庭の回線でこれが開いている場合、意図しないプロキシ機能が植え付けられ、踏み台化している可能性が極めて高いです。
            
            * **💀 7547 (CWMP)**
                * **判定**: **ルーター乗っ取りの兆候** です。ISPが管理するためのポートですが、脆弱性がある場合、ルーターそのものがボット化され、「ネットワークの出口」全体が支配されている深刻な状態を示唆します。
            
            * **🤖 5555 / 5554 (ADB/Emu)**
                * **判定**: **Androidデバイスの露出** です。Fire TV StickやAndroid TV、開発用エミュレータなどが、認証なしで外部操作可能な状態で放置されています。
            
            * **📡 1900 (UPnP)**
                * **判定**: **ネットワーク機器の偵察拠点** です。これらが露出していると、攻撃者がネットワーク内の他のデバイスを探査するための入り口として利用されるリスクがあります。
            """)
        return

    # --- メインコンテンツ：Whois検索タブ ---   
    if IS_PUBLIC_MODE:
        mode_title = "☁️ Public Cloud Edition (機能制限あり)"
        mode_color = "gray"
    else:
        mode_title = "🏠 Local Private Edition (フル機能版)"
        mode_color = "green"

    st.title("🔎 検索大臣 - IP/Domain OSINT -")
    st.markdown(f"**Current Mode:** <span style='color:{mode_color}; font-weight:bold;'>{mode_title}</span>", unsafe_allow_html=True)
    # --- アップデート通知エリア  ---
    with st.expander("🌸アップデート情報 (令和８年３月２９日) - 大量ログ解析時の安定化 🌸", expanded=False):
        st.markdown("""
        **Update:**\n
       **⚡ 処理速度の最適化**:
        * 大量処理時のスピードを最優先するため、時間のかかる「RDAP (公式台帳情報)」の取得をデフォルトで【オフ】に変更しました。\n
        **🛠️ UIの最適化とクラッシュ対策**:
        * 取得されていないデータ（複数検索時にスキップされるWHOIS等）の出力チェックボックスが自動的に無効化されるよう改善しました。
        * 巨大な一括ダウンロード生成時にシステムがクラッシュする不具合を修正しました。\n
        """)
    # ------------------------------------------------
    # タブを使って入力モードを切り替え、画面を広く使う
    input_tab1, input_tab2, input_tab3 = st.tabs(["📋 テキスト貼り付け", "📂 ファイル読み込み", "🔍 単一検索 (WHOIS)"])

    with input_tab1:
        manual_input = st.text_area(
            "検索対象を入力 (複数行可: IPアドレス または ドメイン)",
            height=200, 
            placeholder="8.8.8.8\nexample.com\n2404:6800:...",
            help="1行に1つのターゲットを入力してください。"
        )

    with input_tab2:
        # --- モードによるアップロード制限の切り替え ---
        if IS_PUBLIC_MODE:
            # 公開モード (StreamlitCloud版の挙動): txtのみ許可、警告あり
            allowed_types = ['txt']
            label_text = "IPリストをアップロード (.txtのみ)"
            help_text = "※ 1行に1つのターゲットを記載"
        else:
            # ローカルモード (ローカル版の挙動): csv/excel許可
            allowed_types = ['txt', 'csv', 'xlsx', 'xls']
            label_text = "リストをアップロード (txt/csv/xlsx)"
            help_text = "※ 1行に1つのターゲットを記載、またはCSV/ExcelのIP列を自動検出します"

        uploaded_file = st.file_uploader(label_text, type=allowed_types)
        st.caption(help_text)
        
    with input_tab3:
        single_input = st.text_input(
            "単一の検索対象を入力 (IPアドレス または ドメイン)",
            placeholder="8.8.8.8 または example.com",
            help="1つのターゲットだけを素早く検索してレポートを生成します。"
        )

    raw_targets = []
    df_orig = None

    # 元のファイル名をセッションに保存（ダウンロード時のプレフィックス用）
    if uploaded_file:
        st.session_state['base_filename'] = os.path.splitext(uploaded_file.name)[0]
    else:
        st.session_state['base_filename'] = "WhoisSearchResult"

    if manual_input:
        raw_targets.extend(manual_input.splitlines())
        
    if single_input:
        raw_targets.append(single_input.strip())
    
    if uploaded_file:
        # --- 公開モードの場合の読み込み処理 (StreamlitCloud版ロジック) ---
        if IS_PUBLIC_MODE:
             try:
                # シンプルにテキストとして読み込む
                string_data = uploaded_file.read().decode("utf-8")
                raw_targets.extend(string_data.splitlines())
                
                # 元データフレーム機能は無効化
                st.session_state['original_df'] = None
                st.session_state['ip_column_name'] = None
                
                st.info(f"📄 テキスト読み込み完了: {len(raw_targets)} 行")

             except Exception as e:
                st.error(f"ファイル読み込みエラー: {e}")
        
        # --- ローカルモードの場合の読み込み処理  ---
        else:
            ip_col = None
            try:
                if uploaded_file.name.endswith('.csv'):
                    df_orig = pd.read_csv(uploaded_file)
                elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                    df_orig = pd.read_excel(uploaded_file)
                else:
                    # TXTファイル
                    raw_targets.extend(uploaded_file.read().decode("utf-8").splitlines())
                    st.session_state['original_df'] = None
                    st.session_state['ip_column_name'] = None
                
            except Exception as e:
                st.error(f"ファイル読み込みエラー: {e}")
            
            # --- ローカルモードの場合の読み込み処理  ---
            else:
                ip_col = None
                try:
                    if uploaded_file.name.endswith('.csv'):
                        df_orig = pd.read_csv(uploaded_file)
                    elif uploaded_file.name.endswith(('.xlsx', '.xls')):
                        df_orig = pd.read_excel(uploaded_file)
                    else:
                        # TXTファイル
                        raw_targets.extend(uploaded_file.read().decode("utf-8").splitlines())
                        st.session_state['original_df'] = None
                        st.session_state['ip_column_name'] = None
                    
                    if df_orig is not None:
                        st.session_state['original_df'] = df_orig
                        for col in df_orig.columns:
                            sample = df_orig[col].dropna().head(10).astype(str)
                            if any(is_valid_ip(val.strip()) for val in sample):
                                ip_col = col
                                break
                        
                        if ip_col:
                            st.session_state['ip_column_name'] = ip_col
                            raw_targets.extend(df_orig[ip_col].dropna().astype(str).tolist())
                            
                            # --- アップロードデータのプレビュー (空枠の作成) ---
                            st.info(f"📄 ファイル読み込み完了: {len(df_orig)} 行 / IP列: `{ip_col}`")
                            with st.expander("👀 アップロードデータ・プレビュー", expanded=False):
                                preview_container = st.empty() 
                            # ---------------------------------------------
                        else:
                            st.error("ファイル内にIPアドレスの列が見つかりませんでした。")

                except Exception as e:
                    st.error(f"ファイル読み込みエラー: {e}")
    
    cleaned_raw_targets_list = []
    target_freq_counts = {}

    # 生データからすべての空白文字（半角・全角スペース、タブ等）を完全に除去し、空行を排除する
    raw_targets = [re.sub(r'\s+', '', t) for t in raw_targets if t.strip()]

    if raw_targets:
        cleaned_raw_targets_list = [clean_ocr_error_chars(t) for t in raw_targets]
        target_freq_counts = pd.Series(cleaned_raw_targets_list).value_counts().to_dict()
    else:
        target_freq_counts = {}

    targets = []
    invalid_targets_skipped = [] # 無効としてスキップされたターゲットを記録
    ocr_error_chars = set('Iil|OoSsAaBⅡ')
    resolved_dns_map = {} # nslookupの生出力保存用辞書

    def resolve_domain_nslookup(domain):
    
        ips = []
        raw_lines = []
    
        try:
            # システムのリゾルバに依存せず、Google/CloudflareのパブリックDNSを明示的に使用
            resolver = dns.resolver.Resolver(configure=False)
            resolver.nameservers = random.sample(PUBLIC_DNS_SERVERS, 2) + random.sample(PUBLIC_DNS_V6_SERVERS, 1)
            resolver.timeout = 3
            resolver.lifetime = 3

            raw_lines.append(f";; Domain: {domain}")
            raw_lines.append(f";; Resolver: {resolver.nameservers}")

            # --- Aレコード (IPv4) 取得 ---
            try:
                answers_v4 = resolver.resolve(domain, 'A')
                for rdata in answers_v4:
                    ip = rdata.to_text()
                    if ip not in ips:
                        ips.append(ip)
                    raw_lines.append(f"{domain}. \tIN \tA \t{ip}")
            except dns.resolver.NoAnswer:
                raw_lines.append(f";; IPv4 (A) record not found for {domain}")
            except dns.resolver.NXDOMAIN:
                raw_lines.append(f";; Domain {domain} does not exist (NXDOMAIN)")
                return [], "\n".join(raw_lines) # ドメインがないなら終了
            except Exception as e:
                raw_lines.append(f";; IPv4 Query Failed: {str(e)}")

            # --- AAAAレコード (IPv6) 取得 ---
            try:
                answers_v6 = resolver.resolve(domain, 'AAAA')
                for rdata in answers_v6:
                    ip = rdata.to_text()
                    if ip not in ips:
                        ips.append(ip)
                    raw_lines.append(f"{domain}. \tIN \tAAAA \t{ip}")
            except dns.resolver.NoAnswer:
                pass # IPv6がないのは一般的
            except Exception as e:
                raw_lines.append(f";; IPv6 Query Failed: {str(e)}")

            # --- MXレコード (Mail Exchange) 取得 ---
            try:
                # MXレコードは捨てアド特定の生命線であるため、専用の長いライフタイムを設定して取得を試みる
                resolver_mx = dns.resolver.Resolver(configure=False)
                resolver_mx.nameservers = random.sample(PUBLIC_DNS_SERVERS, 3)
                resolver_mx.timeout = 5
                resolver_mx.lifetime = 10
                
                answers_mx = resolver_mx.resolve(domain, 'MX')
                for rdata in answers_mx:
                    mx_target = rdata.exchange.to_text(omit_final_dot=True)
                    mx_pref = rdata.preference
                    raw_lines.append(f"{domain}. \tIN \tMX \t{mx_pref} {mx_target}")
            except dns.resolver.NoAnswer:
                raw_lines.append(f";; MX record not found for {domain}")
            except Exception as e:
                raw_lines.append(f";; MX Query Failed: {str(e)}")

        except Exception as e:
            raw_lines.append(f";; Critical DNS Error: {str(e)}")
    
        return ips, "\n".join(raw_lines)

    for t in raw_targets:
        original_t = t
        is_ocr_error_likely = any(c in ocr_error_chars for c in original_t)
        if is_ocr_error_likely:
            cleaned_t = clean_ocr_error_chars(original_t)
            if is_valid_ip(cleaned_t):
                if cleaned_t not in targets: targets.append(cleaned_t)
                continue
            t = original_t
        
        invalid_ip_chars = set('ghijklmnopqrstuvwxyz')
        has_hyphen = '-' in t
        has_strictly_domain_char = any(c in invalid_ip_chars for c in t.lower())
        is_likely_domain_or_host = has_hyphen or has_strictly_domain_char
    
        if is_valid_ip(t):
            if t not in targets: targets.append(t)
        elif is_likely_domain_or_host:
            # ドメイン形式の厳格チェック
            if is_valid_domain(t):
                # DNS解決を入力時から削除し、単にドメインとしてキューに入れる
                if t not in targets: targets.append(t)
            else:
                invalid_targets_skipped.append(t) # 不正なドメインとして除外
        else:
            cleaned_t_final = clean_ocr_error_chars(t)
            if is_valid_ip(cleaned_t_final):
                if cleaned_t_final not in targets: targets.append(cleaned_t_final)
            else:
                # クリーンアップ後もドメイン形式の厳格チェック
                if is_valid_domain(cleaned_t_final):
                    if cleaned_t_final not in targets: targets.append(cleaned_t_final)
                else:
                    invalid_targets_skipped.append(t) # 不正なドメインとして除外

    # スキップされたターゲットがあれば警告を表示
    if invalid_targets_skipped:
        st.warning(f"⚠️ 以下の入力は「IPアドレス」または「有効なドメイン形式 (例: example.com)」を満たしていないため、検索対象から除外されました: **{', '.join(list(set(invalid_targets_skipped)))}**")

    # --- プレビュー表に判定結果を反映させる (NEW) ---
    if 'preview_container' in locals() and df_orig is not None and ip_col:
        preview_df = df_orig.copy()
        # 除外対象がある場合のみ判定列を追加する
        if invalid_targets_skipped:
            invalid_set = set(invalid_targets_skipped) # ⬅️ 検索高速化のためSet(集合)に変換
            def check_status(val):
                if pd.isna(val): return "➖ 空欄"
                val_str = str(val).strip()
                if val_str in invalid_set:
                    return "⚠️ 除外 (形式エラー)"
                return "✅ 検索対象"
            
            # データフレームの一番左 (インデックス0) に判定列を挿入
            preview_df.insert(0, '📝 判定結果', preview_df[ip_col].apply(check_status))
            
        # プレースホルダーにデータフレームを描画
        preview_container.dataframe(preview_df, width="stretch")

    has_new_targets = (targets != st.session_state.targets_cache)
    
    if has_new_targets or 'target_freq_map' not in st.session_state:
        st.session_state['target_freq_map'] = target_freq_counts
        st.session_state['original_input_list'] = cleaned_raw_targets_list
        if has_new_targets:
            st.session_state['resolved_dns_map'] = {} # 新規入力時はマップをリセットする

    # --- エンジン処理用の振り分け（ドメイン(IP)はIPとして処理させる） ---
    ip_targets = [t for t in targets if is_valid_ip(t)]
    domain_targets = [t for t in targets if not is_valid_ip(t)]

    # --- UI表示用の厳密なカウント & カテゴリ分け ---
    # 1. ドメインから解決されたIP (例: "domain.com (1.2.3.4)")
    count_resolved_ip = sum(1 for t in ip_targets if "(" in t and ")" in t)
    
    # 2. 直接入力されたIPv6
    count_direct_ipv6 = sum(1 for t in ip_targets if not is_ipv4(t) and "(" not in t)
    
    # 3. 直接入力されたIPv4 (全IPターゲット - 解決分 - IPv6)
    count_direct_ipv4 = len(ip_targets) - count_direct_ipv6 - count_resolved_ip
    
    # 4. 純粋なドメインターゲット
    count_domain = len(domain_targets)

    # 合計待機数
    count_pending = len(st.session_state.deferred_ips)

    # 設定エリアをExpanderに格納し、デフォルトで閉じておく
    with st.expander("⚙️ 検索表示・解析オプション (クリックして展開)", expanded=False):
        col_set1, col_set2 = st.columns(2)
        
        # UIの評価順序を制御するため、先に右カラム(col_set2)のチェックボックスを定義する
        with col_set2:
            st.markdown("**解析モード:** (追加の解析オプションを選択)")
            # InternetDBオプション
            use_internetdb_option = st.checkbox("IoTリスク検知 (InternetDBを利用)", value=False, help="Shodan InternetDBを利用して、対象IPの開放ポートや踏み台リスクを検知します。")
            # RDAPオプション
            use_rdap_option = st.checkbox("公式レジストリ情報 (RDAP公式台帳の併用 - 5秒待機)", value=False, help="RDAP(公式台帳)から最新のネットワーク名を取得します。アクセス制限を避けるため処理速度が強制的に低下します。")
            # 逆引き(rDNS)オプション
            use_rdns_option = st.checkbox("IP逆引き (Reverse DNS - dnspython)", value=False, help="対象IPアドレスに対してdnspythonを実行し、ホスト名(PTRレコード)を取得して詳細レポートに追加します。")
            # SecurityTrails Reverse IPオプション
            use_st_reverse_ip = st.checkbox(
                "Reverse IP (SecurityTrails API)", 
                value=False, 
                disabled=not bool(st_api_key), 
                help="SecurityTrails APIを使用し、対象IPに紐づくドメイン群を逆検索します。※APIキーの設定が必要です。"
            )

        with col_set1:
            display_mode = st.radio(
                "**表示モード:** (検索結果の表示形式とAPI使用有無を設定)",
                ("標準モード", "集約モード (IPv4 Group)", "簡易モード (APIなし)"),
                key="display_mode_radio",
                horizontal=False
            )
            st.markdown("---") 
            
            # RDAPまたはrDNSがオンの場合は、ユーザーに設定させずUI上で固定値を明示する
            if use_rdap_option:
                st.info("ℹ️ **RDAP有効時の制限**\n公式台帳のアクセス制限を回避するため、自動的に「単一スレッド / 5秒待機」に固定されます。速度を優先する場合は右側のチェックを外してください。")
                max_workers = 1
                delay_between_requests = 5.0
            elif use_rdns_option:
                st.info("ℹ️ **逆引き(rDNS)有効時の制限**\nDNSクエリの競合を防ぐため、自動的に「単一スレッド / 2秒待機」に固定されます。速度を優先する場合は右側のチェックを外してください。")
                max_workers = 1
                delay_between_requests = 2.0
            else:
                # 1. API 処理モードの選択
                api_mode_options = list(MODE_SETTINGS.keys()) + ["カスタム設定 (任意調整)"]
                api_mode_selection = st.radio(
                    "**API 処理モード:** (速度と安定性のトレードオフ)",
                    api_mode_options,
                    key="api_mode_radio",
                    horizontal=False
                )
                # 2. 変数の確定ロジック (KeyError 回避策)
                if api_mode_selection == "カスタム設定 (任意調整)":
                    st.markdown("---")
                    max_workers = st.slider("並列スレッド数 (同時処理数)", 1, 5, 2, help="数を増やすと速くなりますが、API制限にかかりやすくなります。")
                    delay_between_requests = st.slider("リクエスト間待機時間 (秒)", 0.1, 5.0, 1.5, 0.1, help="値を増やすほど安全ですが、検索に時間がかかります。")
                else:
                    selected_settings = MODE_SETTINGS[api_mode_selection]
                    max_workers = selected_settings["MAX_WORKERS"]
                    delay_between_requests = selected_settings["DELAY_BETWEEN_REQUESTS"]
            
            # 3. 共通定数の設定
            rate_limit_wait_seconds = RATE_LIMIT_WAIT_SECONDS

    mode_mapping = {
        "標準モード": "標準モード (1ターゲット = 1行)",
        "集約モード (IPv4 Group)": "集約モード (IPv4アドレスをISP/国別でグループ化)",
        "簡易モード (APIなし)": "簡易モード (APIなし - セキュリティリンクのみ)"
    }
    current_mode_full_text = mode_mapping[display_mode]

    is_currently_searching = st.session_state.is_searching and not st.session_state.cancel_search
    
    st.markdown("### 📋 実行前ステータス・確認事項")
    
    # --- 公開モード時のみセキュリティ警告を表示 ---
    if IS_PUBLIC_MODE:
        st.warning("""
        **🛡️ セキュリティ上の注意**
        * **テキスト入力推奨**: ファイルアップロードよりも、左側のテキストエリアへの**コピー＆ペースト**の方が、メタデータ（作成者情報など）が含まれないため安全です。
        * **ファイル名に注意**: アップロードする場合は、ファイル名に機密情報（例: `ClientA_Log.txt`）を含めず、`list.txt` などの無機質な名前を使用してください。
        """)

    status_msg = (
        f"**検索対象:** IPアドレス: {count_direct_ipv4}件(v4)・{count_direct_ipv6}件(v6) / "
        f"ドメイン: {count_domain} 件 (正引きIP: {count_resolved_ip}件) / "
        f"待機中: {count_pending}件 / **キャッシュ:** {len(st.session_state.cidr_cache)}件"
    )
    st.info(status_msg)
      
    # 3. 各種APIの精度・制限に関する警告文
    if not pro_api_key:
        st.warning("⚠️ **IPinfo Inactive:** 通常版API(ip-api)を使用するため、ISP判定結果が正確ではない可能性があります。")
    else:
        st.success("🔑 **IPinfo Pro Active:** 高精度なISP情報・地理位置を取得します。")

    if not vpnapi_key:
        st.warning("⚠️ **VPNAPI.io Inactive:** 未設定時はTorノードのみを検知し、それ以外のプロキシ/VPN判定は空欄となります。高精度な判定が必要な場合はAPIキーを設定してください。")
    else:
        st.success("🕵️ **VPNAPI.io Evidence Active:** 不審判定時に自動で匿名通信判定結果を取得します。")

    if not use_internetdb_option:
        st.caption("※ **IoT Check Inactive:** IoT/脆弱性リスク検知はスキップされます。")

    st.markdown("<br>", unsafe_allow_html=True) # ボタンとの間に少し余白を作る

    # 4. 実行ボタン
    is_currently_searching = st.session_state.is_searching and not st.session_state.cancel_search
    total_ip_targets_for_display = len(ip_targets) + len(st.session_state.deferred_ips)

    if is_currently_searching:
        if st.button("❌ 検索を中止する", type="secondary", width="stretch"):
            st.session_state.cancel_search = True
            st.session_state.is_searching = False
            st.session_state.deferred_ips = {}
            st.rerun()
    else:
        # ボタンのテキストを変更し、警告を読んだことを意識させる
        execute_search = st.button(
        "🚀 上記の確認事項を了承して検索を開始する",
        type="primary",
        width="stretch",
        disabled=(len(targets) == 0 and len(st.session_state.deferred_ips) == 0)
    )

    if ('execute_search' in locals() and execute_search and (has_new_targets or len(st.session_state.deferred_ips) > 0)) or is_currently_searching:
        
        if ('execute_search' in locals() and execute_search and has_new_targets and len(targets) > 0):
            # 新規検索時に古い巨大なデータを明示的に解放し、状態をリセットする
            reset_search_state()
            st.session_state.targets_cache = targets
            st.rerun() 
            
        elif is_currently_searching:
            targets = st.session_state.targets_cache
            domain_targets = [t for t in targets if not is_valid_ip(t)]

            st.subheader("⏳ 処理中...")
            
            # メインスレッドを占有しないよう、検索開始直後に専用スレッドで並列DNS解決を一括実行する
            unresolved_domains = [d for d in domain_targets if d not in st.session_state.get('resolved_dns_map', {})]
            if unresolved_domains:
                with st.spinner(f"⏳ {len(unresolved_domains)}件のドメインを並列で名前解決中... (並列数: {max_workers})"):
                    def resolve_and_map(domain):
                        ips, raw = resolve_domain_nslookup(domain)
                        return domain, ips, raw
                    
                    # DNSクエリ(UDP)によるルーターのNAT溢れを防ぐため、ユーザー設定のmax_workersに同期させる
                    with ThreadPoolExecutor(max_workers=max_workers) as dns_executor:
                        dns_results = list(dns_executor.map(resolve_and_map, unresolved_domains))
                        
                    for domain, ips, raw in dns_results:
                        st.session_state.resolved_dns_map[domain] = {'ips': ips, 'raw': raw}
                        for resolved_ip in ips:
                            combined_t = f"{domain} ({resolved_ip})"
                            if combined_t not in targets: 
                                targets.append(combined_t)
                    
                    # DNS解決済みのターゲットリストでキャッシュを最新状態に上書き
                    st.session_state.targets_cache = targets

            # DNS並列解決が完了した後、改めて全体のIPターゲットを抽出してキューに流す
            ip_targets = [t for t in targets if is_valid_ip(t)]
            total_targets = len(targets)
            total_ip_api_targets = len(ip_targets)
            
            ip_targets_to_process = [ip for ip in ip_targets if ip not in st.session_state.finished_ips]
            
            current_time = time.time()
            ready_to_retry_ips = []
            deferred_ips_new = {}
            for ip, defer_time in st.session_state.deferred_ips.items():
                if current_time >= defer_time:
                    ready_to_retry_ips.append(ip)
                else:
                    deferred_ips_new[ip] = defer_time
            
            st.session_state.deferred_ips = deferred_ips_new
            
            immediate_ip_queue_unique = []
            for ip in ip_targets_to_process:
                if ip not in st.session_state.deferred_ips and ip not in immediate_ip_queue_unique:
                    immediate_ip_queue_unique.append(ip)

            immediate_ip_queue = immediate_ip_queue_unique
            immediate_ip_queue.extend(ready_to_retry_ips)
            
            is_single_input = (len(cleaned_raw_targets_list) == 1)
            if "簡易" in current_mode_full_text:
                if not st.session_state.raw_results:
                    results_list = []
                    for t in targets:
                        results_list.append(get_simple_mode_details(t))
                    st.session_state.raw_results = results_list
                    st.session_state.finished_ips.update(targets)
                    st.session_state.is_searching = False
                    st.rerun()

            else:
                if not any(res['ISP'] == 'Domain/Host' for res in st.session_state.raw_results) and domain_targets:
                    for d in domain_targets:
                        dns_data = st.session_state.get('resolved_dns_map', {}).get(d, {})
                        ns_raw = dns_data.get('raw', '') if isinstance(dns_data, dict) else str(dns_data)
                        res_domain = get_domain_details(d, ns_raw, st_api_key, st_start_date, st_end_date, is_single_target=is_single_input)
                        
                        heavy_keys = ['RDAP_JSON', 'VPNAPI_JSON', 'IPINFO_JSON', 'DOMAIN_RDAP_JSON', 'ST_JSON', 'RDNS_DATA', 'ST_REVERSE_IP_JSON', 'DOMAIN_WHOIS_TEXT', 'IP_WHOIS_TEXT']
                        ip_val = res_domain['Target_IP']
                        st.session_state.detailed_data[ip_val] = {k: res_domain.pop(k) for k in heavy_keys if k in res_domain}
                        
                        st.session_state.raw_results.append(res_domain)
                    st.session_state.finished_ips.update(domain_targets)

                prog_bar_container = st.empty()
                status_text_container = st.empty()
                summary_container = st.empty() 

                if immediate_ip_queue:
                    cidr_cache_snapshot = st.session_state.cidr_cache.copy() 
                    learned_isps_snapshot = st.session_state.learned_proxy_isps.copy()
                    
                    # --- 各種オプション有効時の動的負荷調整 (安全装置) ---
                    current_max_workers = max_workers
                    current_delay = delay_between_requests
                    
                    if use_rdap_option:
                        # RDAPエンドポイントの厳格なアクセス制限(429エラー)を回避するため強制保護
                        current_max_workers = 1
                        if current_delay < 5.0:
                            current_delay = 5.0
                        st.info("ℹ️ RDAP公式台帳のアクセス制限を回避するため、安全モード（シングルスレッド/最低5秒待機）で実行中...")
                    elif use_rdns_option:
                        # DNSクエリの競合とタイムアウトを防ぐため強制的にシングルスレッド化
                        current_max_workers = 1 
                        if current_delay < 2.0:
                            current_delay = 2.0
                        st.info("ℹ️ 逆引き精度向上のため、負荷調整モード（シングルスレッド/最低2秒待機）で実行中...")

                    with ThreadPoolExecutor(max_workers=current_max_workers) as executor:
                        future_to_ip = {
                            executor.submit(
                                get_ip_details_from_api, 
                                ip, 
                                cidr_cache_snapshot, 
                                learned_isps_snapshot, 
                                current_delay,
                                rate_limit_wait_seconds,
                                tor_nodes,
                                cloud_ip_data,
                                use_rdap_option,
                                use_internetdb_option,
                                use_rdns_option,
                                use_st_reverse_ip,
                                pro_api_key,
                                vpnapi_key,
                                st_api_key,
                                st_start_date,
                                st_end_date,
                                use_st_rev_fetchall,
                                is_single_input
                            ): ip for ip in immediate_ip_queue
                        }
                        remaining = set(future_to_ip.keys())

                        # UI更新用のタイマー初期化
                        last_ui_update_time = time.time()
                        last_backup_time = time.time() # バックアップ用タイマー
                        
                        while remaining and not st.session_state.cancel_search:
                            done, remaining = wait(remaining, timeout=0.1, return_when=FIRST_COMPLETED)
                            
                            # タスクが完了した時のみ画面更新処理を行う
                            if done: 
                                for f in done:
                                    res_tuple = f.result()
                                    res = res_tuple[0]
                                    new_cache_entry = res_tuple[1] if len(res_tuple) > 1 else None
                                    new_learned_isp = res_tuple[2] if len(res_tuple) > 2 else None
                                    ip = res['Target_IP']
                                    
                                    if new_cache_entry:
                                        st.session_state.cidr_cache.update(new_cache_entry)
                                    
                                    # メインスレッド側で学習済みリストを安全に更新
                                    if new_learned_isp:
                                        st.session_state.learned_proxy_isps.update(new_learned_isp)
                                    if res.get('Status', '').startswith('Success'):
                                        heavy_keys = ['RDAP_JSON', 'VPNAPI_JSON', 'IPINFO_JSON', 'DOMAIN_RDAP_JSON', 'ST_JSON', 'RDNS_DATA', 'ST_REVERSE_IP_JSON', 'DOMAIN_WHOIS_TEXT', 'IP_WHOIS_TEXT']
                                        st.session_state.detailed_data[ip] = {k: res.pop(k) for k in heavy_keys if k in res}
                                        
                                        st.session_state.raw_results.append(res)
                                        st.session_state.finished_ips.add(ip)
                                    elif res.get('Defer_Until'):
                                        st.session_state.deferred_ips[ip] = res['Defer_Until']
                                    else:
                                        heavy_keys = ['RDAP_JSON', 'VPNAPI_JSON', 'IPINFO_JSON', 'DOMAIN_RDAP_JSON', 'ST_JSON', 'RDNS_DATA', 'ST_REVERSE_IP_JSON', 'DOMAIN_WHOIS_TEXT', 'IP_WHOIS_TEXT']
                                        st.session_state.detailed_data[ip] = {k: res.pop(k) for k in heavy_keys if k in res}
                                        
                                        st.session_state.raw_results.append(res)
                                        st.session_state.finished_ips.add(ip)

                                current_time_for_ui = time.time()
                                is_last_item = not remaining and not st.session_state.deferred_ips

                                if total_ip_api_targets > 0 and (current_time_for_ui - last_ui_update_time > 1.5 or is_last_item):
                                    last_ui_update_time = current_time_for_ui # タイマーをリセット
                                    
                                    processed_api_ips_count = len([ip for ip in st.session_state.finished_ips if is_valid_ip(ip)])
                                    pct = int(processed_api_ips_count / total_ip_api_targets * 100)
                                    elapsed_time = time.time() - st.session_state.search_start_time
                                    eta_seconds = 0
                                    if processed_api_ips_count > 0:
                                        rate = processed_api_ips_count / elapsed_time
                                        remaining_count = total_ip_api_targets - processed_api_ips_count
                                        eta_seconds = math.ceil(remaining_count / rate)
                                    
                                    eta_display = "計算中..."
                                    if eta_seconds > 0:
                                        minutes = int(eta_seconds // 60)
                                        seconds = int(eta_seconds % 60)
                                        eta_display = f"{minutes:02d}分{seconds:02d}秒"
                                        
                                    # withを使わずに直接コンテナを上書きしてチラつきを防ぐ
                                    prog_bar_container.progress(pct)
                                    status_text_container.info(f"**⏳ 処理中... ({pct}%)** | 完了: {processed_api_ips_count}/{total_ip_api_targets} | ⏸️ 保留: {len(st.session_state.deferred_ips)} | 📦 キャッシュ: {len(st.session_state.cidr_cache)} | ⏱️ 残り: {eta_display}")
                                    
                                    isp_df, country_df, freq_df, country_all_df, isp_full_df, country_full_df, freq_full_df, proxy_df = summarize_in_realtime(st.session_state.raw_results)
                                    
                                    # empty()による全消去を廃止し、直接上書きさせることで点滅を防ぐ
                                    with summary_container.container():
                                        draw_summary_content(isp_df, country_df, freq_df, country_all_df, proxy_df, "📊 リアルタイム分析") 
                                        
                                    # 10秒ごとにディスクへセッションをバックアップする
                                    if current_time_for_ui - last_backup_time > 10.0 or is_last_item:
                                        save_recovery_data()
                                        last_backup_time = current_time_for_ui

                            if not remaining and not st.session_state.deferred_ips:
                                break
                            
                            if st.session_state.deferred_ips:
                                # 強制再起動ではなく、未実行のタスクをキャンセルしてループを安全に脱出する
                                for f in remaining:
                                    f.cancel()
                                break  
                            
                            time.sleep(0.5) 
                            
                        if total_ip_api_targets > 0 and not st.session_state.deferred_ips:
                            processed_api_ips_count = len([ip for ip in st.session_state.finished_ips if is_valid_ip(ip)])
                            final_pct = int(processed_api_ips_count / total_ip_api_targets * 100)
                            with prog_bar_container:
                                st.progress(final_pct)
                            with status_text_container:
                                st.success(f"**✅ 処理完了 (100%)** | 完了: {processed_api_ips_count}/{total_ip_api_targets} | 📦 キャッシュ: {len(st.session_state.cidr_cache)}")
                        
                if len(st.session_state.finished_ips) == total_targets and not st.session_state.deferred_ips:
                    st.session_state.is_searching = False
                    clear_recovery_data() # 正常完了時はバックアップを消去
                    st.info("✅ 全ての検索が完了しました。")
                    st.rerun()
                
                elif st.session_state.deferred_ips and not st.session_state.cancel_search:
                    next_retry_time = min(st.session_state.deferred_ips.values())
                    wait_time = max(1, int(next_retry_time - time.time()))
                    
                    prog_bar_container.empty()
                    status_text_container.empty()
                    st.warning(f"⚠️ **ネットワーク切断、またはAPI制限を検知しました。** 保留中の **{len(st.session_state.deferred_ips)}** 件のターゲットは通信回復を待ち、**{wait_time}** 秒後に自動で再試行されます。")
                    time.sleep(min(5, wait_time)) 
                    st.rerun()

                elif st.session_state.cancel_search:
                    prog_bar_container.empty()
                    status_text_container.empty()
                    st.warning("検索がユーザーによって中止されました。")
                    st.session_state.is_searching = False
                    st.rerun()


    # --- 結果表示 ---
    if st.session_state.raw_results or st.session_state.deferred_ips:
        res = st.session_state.raw_results
        
        if st.session_state.get('debug_summary'):
            with st.expander("🛠️ デバッグ情報 (集計データ確認用)", expanded=False):
                st.markdown("**API 処理モード設定**")
                st.write(f"MAX_WORKERS: {max_workers}")
                st.write(f"DELAY_BETWEEN_REQUESTS: {delay_between_requests}")
                st.markdown("---")
                st.json(st.session_state['debug_summary'].get('country_code_counts', {}))
                st.json(st.session_state['debug_summary'].get('country_all_df', []))
                st.markdown("---")
                st.json(st.session_state.get('cidr_cache', {}))

        
        successful_results = [r for r in res if r['Status'].startswith('Success') or r['Status'].startswith('Aggregated')]
        error_results = [r for r in res if not (r['Status'].startswith('Success') or r['Status'].startswith('Aggregated'))]
        
        for ip, defer_time in st.session_state.deferred_ips.items():
            status = f"Pending (Retry in {max(0, int(defer_time - time.time()))}s)"
            error_results.append({
                'Target_IP': ip, 'ISP': 'N/A', 'Country': 'N/A', 'CountryCode': 'N/A', 'RIR_Link': get_authoritative_rir_link(ip, 'N/A'),
                'Secondary_Security_Links': create_secondary_links(ip), 
                'Status': status
            })
        
        if "集約" in current_mode_full_text:
            display_res = group_results_by_isp(successful_results)
            display_res.extend(error_results)
        else:
            display_res = successful_results + error_results
            target_order = {ip: i for i, ip in enumerate(targets)}
            display_res.sort(key=lambda x: target_order.get(get_copy_target(x['Target_IP']), float('inf')))

        display_results(display_res, current_mode_full_text, display_mode, use_rdap_option, pro_api_key, vpnapi_key, st_api_key, use_rdns_option)
        
        if not st.session_state.is_searching or st.session_state.cancel_search:
            isp_df, country_df, freq_df, country_all_df, isp_full_df, country_full_df, freq_full_df, proxy_df = summarize_in_realtime(st.session_state.raw_results)
            st.markdown("---")
            draw_summary_content(isp_df, country_df, freq_df, country_all_df, proxy_df, "✅ 集計結果")

            # --- 全入力順・全件ベースのデータフレーム構築 ---
            df_for_analysis = pd.DataFrame()
            
            # マッチング精度を高めるための多重キー辞書の構築
            result_lookup = {}
            for r in st.session_state.raw_results:
                target = r.get('Target_IP', '')
                actual = extract_actual_ip(target)
                result_lookup[target] = r
                if actual and actual != target:
                    result_lookup[actual] = r

            def get_result_info(raw_ip_str):
                if pd.isna(raw_ip_str): return {}
                val = str(raw_ip_str).strip()
                cleaned = clean_ocr_error_chars(val)
                actual = extract_actual_ip(cleaned)
                # 実IP、クリーンIP、生文字列の順で一致する結果を探す
                return result_lookup.get(actual) or result_lookup.get(cleaned) or result_lookup.get(val) or {}

            full_input_list = st.session_state.get('original_input_list', [])

            if full_input_list:
                if st.session_state.get('original_df') is not None:
                    # 元のアップロードデータ(CSV/Excel)が存在する場合、その行構造(時間など)を完全維持する
                    df_for_analysis = st.session_state['original_df'].copy()
                    ip_col = st.session_state['ip_column_name']
                    
                    df_for_analysis['国名'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('Country_JP', 'N/A'))
                    df_for_analysis['Whois結果（元データ）'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('ISP', 'N/A'))
                    df_for_analysis['Whois結果（日本語名称）'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('ISP_JP', 'N/A'))
                    df_for_analysis['RDAP結果（元データ）'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('RDAP_Name_Raw', 'N/A'))
                    df_for_analysis['RDAP結果（日本語名称）'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('RDAP_JP', 'N/A'))
                    df_for_analysis['プロキシ種別'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('Proxy_Type', ''))
                    df_for_analysis['IoTリスク'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('IoT_Risk', 'N/A'))
                    df_for_analysis['ステータス'] = df_for_analysis[ip_col].map(lambda x: get_result_info(x).get('Status', 'N/A'))
                else:
                    # テキスト貼り付けの場合
                    temp_rows = []
                    for t in full_input_list:
                        info = get_result_info(t)
                        temp_rows.append({
                            '対象IP/Domain': t,
                            '国名': info.get('Country_JP', 'N/A'),
                            'Whois結果（元データ）': info.get('ISP', 'N/A'),
                            'Whois結果（日本語名称）': info.get('ISP_JP', 'N/A'),
                            'RDAP結果（元データ）': info.get('RDAP_Name_Raw', 'N/A'),
                            'RDAP結果（日本語名称）': info.get('RDAP_JP', 'N/A'),
                            'プロキシ種別': info.get('Proxy_Type', ''),
                            'IoTリスク': info.get('IoT_Risk', 'N/A'),
                            'ステータス': info.get('Status', 'N/A')
                        })
                    df_for_analysis = pd.DataFrame(temp_rows)

            # 💡 マスターデータ（Excel/全件CSV用）から無効オプション列を削除
            if not df_for_analysis.empty:
                master_cols_to_drop = []
                if not use_rdap_option:
                    master_cols_to_drop.extend(['RDAP結果（元データ）', 'RDAP結果（日本語名称）'])
                if not use_internetdb_option:
                    master_cols_to_drop.append('IoTリスク')
                
                if master_cols_to_drop:
                    df_for_analysis = df_for_analysis.drop(columns=[c for c in master_cols_to_drop if c in df_for_analysis.columns], errors='ignore')

            # --- クロス分析 (画面表示) ---
            if not df_for_analysis.empty:
                st.markdown("---")
                # 出力用マスターデータを汚染しないよう、一時的なコピーを作成
                df_for_render = df_for_analysis.copy()
                if st.session_state.get('ip_column_name') and st.session_state['ip_column_name'] in df_for_render.columns:
                    df_for_render['Target_IP'] = df_for_render[st.session_state['ip_column_name']].astype(str)
                elif '対象IP/Domain' in df_for_render.columns:
                    df_for_render['Target_IP'] = df_for_render['対象IP/Domain'].astype(str)
                
                render_merged_analysis(df_for_render)

            # --- UI改善：ダウンロードセンター ---
            st.markdown("---")
            st.markdown("### 📥 レポート ＆ データ出力")
            
            base_fname = st.session_state.get('base_filename', 'WhoisSearchResult')
            ts = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
            file_prefix = f"{base_fname}_{ts}"
            
            main_col1, main_col2 = st.columns(2)
            with main_col1:
                st.info("📊 **分析マスター (全入力順)**\n\nアップロードされた全行に基づき、ISP・RDAP・国別などの集計表とグラフを生成します。")
                if not df_for_analysis.empty:
                    time_cols = [c for c in df_for_analysis.columns if any(k in c.lower() for k in ['date', 'time', 'jst'])]
                    selected_time_col = None
                    if time_cols:
                        selected_time_col = st.selectbox("時間分析に使用する列:", df_for_analysis.columns, index=df_for_analysis.columns.get_loc(time_cols[0]), key="time_col_selector_final")
                    
                    with st.spinner("⏳ Excelレポートを生成中..."):
                        excel_advanced = create_advanced_excel(df_for_analysis, selected_time_col)
                    
                    excel_filename = f"{file_prefix}_MasterReport.xlsx"
                    if IS_PUBLIC_MODE:
                        st.download_button(
                            label="📥 Excelレポート (全入力順・グラフ付き) を保存",
                            data=excel_advanced,
                            file_name=excel_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            width="stretch",
                            type="primary"
                        )
                    else:
                        render_local_save_ui(
                            "💾 Excelレポートをローカル保存 (容量無制限)", 
                            excel_filename, excel_advanced, "master_excel", "primary"
                        )
                else:
                    st.button("データなし", disabled=True, width="stretch")

            with main_col2:
                st.success("🌐 **全件グラフ HTMLレポート**\n\nブラウザで閲覧・印刷可能なグラフィカルな分析レポートです。")
                with st.spinner("⏳ HTMLレポートを生成中..."):
                    html_report = generate_full_report_html(isp_full_df, country_full_df, freq_full_df)
                
                html_summary_filename = f"{file_prefix}_Summary.html"
                if IS_PUBLIC_MODE:
                    st.download_button(
                        label="📥 HTMLレポート (閲覧・印刷用) を表示",
                        data=html_report,
                        file_name=html_summary_filename,
                        mime="text/html",
                        width="stretch"
                    )
                else:
                    render_local_save_ui(
                        "💾 HTMLレポートをローカル保存", 
                        html_summary_filename, html_report, "master_html", "secondary"
                    )

            with st.expander("🛠️ システム連携用・RAWデータ ＆ 脅威インテリジェンス出力"):
                st.caption("SIEM（セキュリティログ監視）への取り込みや、データベース連携に利用してください。")
                sub_tab1, sub_tab2, sub_tab3 = st.tabs(["📄 検索結果リスト (CSV/Excel)", "📈 統計・カウントデータ", "🛡️ STIX 2.1 (SOC/MISP連携用)"])
                
                with sub_tab1:
                    c1, c2 = st.columns(2)
                    csv_display = pd.DataFrame(display_res).astype(str)
                    rename_map = {
                        'Target_IP': 'IPアドレス',
                        'Country_JP': '国名', 
                        'ISP_API_Raw': 'Whois(元データ)',
                        'ISP_JP': 'Whois(日本語名)',
                        'RDAP_Name_Raw': 'RDAP(元データ)',
                        'RDAP_JP': 'RDAP(日本語名)',
                        'Proxy_Type': 'Proxy種別',
                        'IoT_Risk': 'IoTリスク',
                        'Status': 'ステータス'
                    }
                    csv_display = csv_display.rename(columns=rename_map)
                    
                    # 💡 ダウンロード用(画面表示順)から無効オプション列を削除
                    display_cols_to_drop = []
                    if not use_rdap_option:
                        display_cols_to_drop.extend(['RDAP(元データ)', 'RDAP(日本語名)'])
                    if not use_internetdb_option:
                        display_cols_to_drop.append('IoTリスク')
                        
                    if display_cols_to_drop:
                        csv_display = csv_display.drop(columns=[c for c in display_cols_to_drop if c in csv_display.columns], errors='ignore')

                    with c1:
                        st.markdown("**画面表示順 (現在の並び)**")
                        st.download_button("CSV形式", csv_display.to_csv(index=False).encode('utf-8-sig'), f"{file_prefix}_Display.csv", "text/csv", key="csv_display_btn", width="stretch")
                        st.download_button("Excel形式", convert_df_to_excel(csv_display), f"{file_prefix}_Display.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="excel_display_btn", width="stretch")
                    
                    # 元のアップロードデータ(時間や他カラムを含む)を完全に維持している df_for_analysis をそのまま出力に使用する
                    if not df_for_analysis.empty:
                        csv_full = df_for_analysis.astype(str)
                    else:
                        csv_full = pd.DataFrame() # 空の場合のフォールバック
                        
                    with c2:
                        st.markdown("**全データ (入力した順番)**")
                        st.download_button("CSV形式", csv_full.to_csv(index=False).encode('utf-8-sig'), f"{file_prefix}_Full.csv", "text/csv", key="csv_full_btn", width="stretch")
                        st.download_button("Excel形式", convert_df_to_excel(csv_full), f"{file_prefix}_Full.xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", key="excel_full_btn", width="stretch")

                with sub_tab2:
                    sc1, sc2, sc3 = st.columns(3)
                    with sc1:
                        st.download_button("🎯 ターゲット別件数 (CSV)", freq_full_df.to_csv(index=False).encode('utf-8-sig'), f"{file_prefix}_Freq.csv", "text/csv", key="btn_freq_csv", width="stretch")
                    with sc2:
                        st.download_button("🏢 ISP別件数 (CSV)", isp_full_df.to_csv(index=False).encode('utf-8-sig'), f"{file_prefix}_ISP.csv", "text/csv", key="btn_isp_csv", width="stretch")
                    with sc3:
                        st.download_button("🌍 国別件数 (CSV)", country_full_df.to_csv(index=False).encode('utf-8-sig'), f"{file_prefix}_Country.csv", "text/csv", key="btn_country_csv", width="stretch")

                with sub_tab3:
                    st.info("**STIX (Structured Threat Information Expression) 2.1 形式**\n\n調査結果を、世界標準の脅威インテリジェンス・フォーマット (JSON形式) で出力します。SIEMへのIoC（侵害指標）の取り込みや、MISPへのインポートにそのまま使用できます。")
                    
                    stix_data = generate_stix2_bundle(display_res)
                    stix_filename = f"{file_prefix}_STIX.json"
                    
                    if IS_PUBLIC_MODE:
                        st.download_button(
                            label="STIX 2.1 Bundle (JSON) をダウンロード",
                            data=stix_data.encode('utf-8'),
                            file_name=stix_filename,
                            mime="application/json",
                            width="stretch",
                            type="primary"
                        )
                    else:
                        render_local_save_ui(
                            "STIX 2.1 Bundle をローカル保存", 
                            stix_filename, stix_data.encode('utf-8'), "stix_json", "primary"
                        )
                
if __name__ == "__main__":
    main()
